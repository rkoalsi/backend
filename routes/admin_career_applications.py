from fastapi import (
    APIRouter,
    HTTPException,
    Query,
)
from fastapi.responses import JSONResponse, StreamingResponse
from ..config.root import get_database, serialize_mongo_document
from bson.objectid import ObjectId
from typing import Optional
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()
db = get_database()


@router.get("")
def get_career_applications(
    page: int = Query(0, ge=0, description="0-based page index"),
    limit: int = Query(10, ge=1, description="Number of items per page"),
    career_id: Optional[str] = Query(None, description="Filter by career ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
):
    try:
        match_statement = {}
        if career_id:
            match_statement["career_id"] = ObjectId(career_id)
        if status:
            match_statement["status"] = status

        pipeline = [
            {"$match": match_statement},
            {"$sort": {"created_at": -1}},
            {"$skip": page * limit},
            {"$limit": limit},
        ]

        total_count = db.career_applications.count_documents(match_statement)
        cursor = db.career_applications.aggregate(pipeline)
        applications = [serialize_mongo_document(doc) for doc in cursor]
        total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1

        if page > total_pages and total_pages != 0:
            raise HTTPException(status_code=400, detail="Page number out of range")
        return {
            "career_applications": applications,
            "total_count": total_count,
            "page": page,
            "per_page": limit,
            "total_pages": total_pages,
        }
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router.get("/report")
def download_career_applications_report(
    career_id: Optional[str] = Query(None, description="Filter by career ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
):
    try:
        match_statement = {}
        if career_id:
            match_statement["career_id"] = career_id
        if status:
            match_statement["status"] = status

        pipeline = [
            {"$match": match_statement},
            {"$sort": {"created_at": -1}},
        ]

        applications = list(db.career_applications.aggregate(pipeline))

        # Build career_id -> title lookup
        career_ids = list({app.get("career_id", "") for app in applications if app.get("career_id")})
        career_map = {}
        if career_ids:
            for career in db.careers.find({"_id": {"$in": [ObjectId(cid) for cid in career_ids]}}, {"title": 1}):
                career_map[str(career["_id"])] = career.get("title", "")

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Career Applications"

        headers = [
            "Applicant Name",
            "Email",
            "Phone",
            "Career",
            "Current Location",
            "Total Experience",
            "Relevant Experience",
            "Current Company",
            "Current Designation",
            "Current CTC",
            "Expected CTC",
            "Notice Period",
            "Preferred Location",
            "LinkedIn",
            "Resume URL",
            "Available for Interview (7 days)",
            "Applied Before",
            "Status",
            "Applied On",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, app in enumerate(applications, 2):
            ws.cell(row=row_idx, column=1, value=app.get("applicant_name", ""))
            ws.cell(row=row_idx, column=2, value=app.get("applicant_email", ""))
            ws.cell(row=row_idx, column=3, value=app.get("applicant_phone", ""))
            ws.cell(row=row_idx, column=4, value=career_map.get(app.get("career_id", ""), app.get("career_id", "")))
            ws.cell(row=row_idx, column=5, value=app.get("current_location", ""))
            ws.cell(row=row_idx, column=6, value=app.get("total_experience", ""))
            ws.cell(row=row_idx, column=7, value=app.get("relevant_experience", ""))
            ws.cell(row=row_idx, column=8, value=app.get("current_company", ""))
            ws.cell(row=row_idx, column=9, value=app.get("current_designation", ""))
            ws.cell(row=row_idx, column=10, value=app.get("current_ctc", ""))
            ws.cell(row=row_idx, column=11, value=app.get("expected_ctc", ""))
            ws.cell(row=row_idx, column=12, value=app.get("notice_period", ""))
            ws.cell(row=row_idx, column=13, value=app.get("preferred_location", ""))
            ws.cell(row=row_idx, column=14, value=app.get("linkedin_url", ""))
            ws.cell(row=row_idx, column=15, value=app.get("resume_url", ""))
            ws.cell(row=row_idx, column=16, value=app.get("available_for_interview", ""))
            ws.cell(row=row_idx, column=17, value=app.get("applied_before", ""))
            ws.cell(row=row_idx, column=18, value=app.get("status", "pending"))
            created_at = app.get("created_at", "")
            if created_at:
                try:
                    created_at = datetime.fromisoformat(str(created_at)).strftime("%Y-%m-%d %H:%M")
                except Exception:
                    created_at = str(created_at)
            ws.cell(row=row_idx, column=19, value=created_at)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        filename = f"career_applications_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router.get("/{application_id}")
def get_career_application(application_id: str):
    try:
        try:
            obj_id = ObjectId(application_id)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid application_id format")

        doc = db.career_applications.find_one({"_id": obj_id})
        if not doc:
            raise HTTPException(status_code=404, detail="Application not found")

        return serialize_mongo_document(doc)
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router.put("/{application_id}")
def update_career_application(application_id: str, application: dict):
    try:
        try:
            obj_id = ObjectId(application_id)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid application_id format")

        existing_doc = db.career_applications.find_one({"_id": obj_id})
        if not existing_doc:
            raise HTTPException(status_code=404, detail="Application not found")

        update_data = {k: v for k, v in application.items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=400, detail="No update data provided")

        result = db.career_applications.update_one(
            {"_id": obj_id}, {"$set": update_data}
        )

        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Application not found")

        updated_doc = db.career_applications.find_one({"_id": obj_id})
        return serialize_mongo_document(updated_doc)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Internal server error")


@router.delete("/{application_id}")
def delete_career_application(application_id: str):
    try:
        try:
            obj_id = ObjectId(application_id)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid application_id format")

        result = db.career_applications.delete_one({"_id": obj_id})
        if result.deleted_count == 1:
            return {"detail": "Application deleted successfully"}
        else:
            raise HTTPException(status_code=404, detail="Application not found")
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
