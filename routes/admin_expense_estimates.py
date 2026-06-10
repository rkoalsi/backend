from fastapi import APIRouter, Depends, HTTPException, Request, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
import datetime, io, openpyxl, calendar
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from ..config.root import get_database, serialize_mongo_document
from ..config.auth import get_current_user
from .notifications import create_notification
from .expense_estimates import (
    APPROVER_CHAIN,
    _send_email,
    _get_user_by_email,
    _notify_salesperson,
    _current_user_id,
    _get_estimate_or_404,
    _compute_actual_totals,
)

router = APIRouter()
db = get_database()


def _approver_for_status(status: str) -> dict | None:
    for a in APPROVER_CHAIN:
        if a["stage"] == status:
            return a
    return None


# ── list + single ──────────────────────────────────────────────────────────────

@router.get("")
def list_all_estimates(
    page: int = 0,
    limit: int = 20,
    status: str = Query(None),
    salesperson_id: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    current_user: dict = Depends(get_current_user),
):
    query = {}
    if status:
        query["status"] = status
    if salesperson_id:
        query["created_by"] = ObjectId(salesperson_id)
    if start_date:
        query.setdefault("travel_start_date", {})["$gte"] = start_date
    if end_date:
        query.setdefault("travel_start_date", {})["$lte"] = end_date

    total = db.expense_estimates.count_documents(query)
    docs = list(
        db.expense_estimates.find(query)
        .sort("created_at", -1)
        .skip(page * limit)
        .limit(limit)
    )
    return {
        "estimates": serialize_mongo_document(docs),
        "total_count": total,
        "total_pages": max(1, -(-total // limit)),
    }


@router.get("/{estimate_id}")
def get_estimate(estimate_id: str, current_user: dict = Depends(get_current_user)):
    return serialize_mongo_document(_get_estimate_or_404(estimate_id))


@router.delete("/{estimate_id}")
def delete_estimate(estimate_id: str, current_user: dict = Depends(get_current_user)):
    role = (current_user.get("data") or current_user).get("role", "")
    if role not in ("admin", "sales_admin"):
        raise HTTPException(status_code=403, detail="Only admin or sales_admin can delete estimates")
    _get_estimate_or_404(estimate_id)
    db.expense_estimates.delete_one({"_id": ObjectId(estimate_id)})
    return {"detail": "Deleted"}


# ── approve ────────────────────────────────────────────────────────────────────

@router.post("/{estimate_id}/approve")
async def approve_estimate(
    estimate_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    est = _get_estimate_or_404(estimate_id)
    caller_email = current_user.get("data", {}).get("email") or current_user.get("email", "")
    current_status = est["status"]

    approver = _approver_for_status(current_status)
    if not approver:
        raise HTTPException(status_code=400, detail=f"Estimate cannot be approved from status '{current_status}'")
    if caller_email != approver["email"]:
        raise HTTPException(status_code=403, detail="You are not the designated approver for this stage")

    body = await request.json()
    remarks = body.get("remarks", "")
    now = datetime.datetime.utcnow()
    next_status = approver["next_status"]

    approval_field_map = {
        "Pending Review": ("rahul_approved_at", "rahul_remarks"),
        "Pending Second Review": ("amit_approved_at", "amit_remarks"),
        "Pending Payment": ("yogesh_processed_at", "yogesh_remarks"),
    }
    ts_field, rem_field = approval_field_map[current_status]

    update = {
        "status": next_status,
        ts_field: now,
        rem_field: remarks,
        "updated_at": now,
    }

    # Pending Payment → Draft: also capture advance_released flag
    if current_status == "Pending Payment":
        advance_released = body.get("advance_released", False)
        approved_total = float(body.get("approved_total") or est.get("estimated_total") or 0)
        update["yogesh_advance_released"] = advance_released
        update["approved_total"] = approved_total

    db.expense_estimates.update_one({"_id": ObjectId(estimate_id)}, {"$set": update})
    updated = db.expense_estimates.find_one({"_id": ObjectId(estimate_id)})

    sp_name = est.get("created_by_name", "")
    trip_date = (est.get("travel_start_date") or "")[:10]

    if next_status == "Pending Second Review":
        next_approver = APPROVER_CHAIN[1]
        subject = f"Expense Estimate Approved – Action Required ({sp_name})"
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
            <h2>Expense Estimate – Second Review Required</h2>
            <p>Expense estimate from <b>{sp_name}</b> (trip: {trip_date}) has been approved by {approver['label']}
            and is now awaiting your review.</p>
            <ul>
                <li><b>Estimated Total:</b> ₹{est.get('estimated_total', 0):,.2f}</li>
                <li><b>Advance Requested:</b> ₹{est.get('advance_requested', 0):,.2f}</li>
            </ul>
            <p>Please log in to review and approve.</p>
        </div>"""
        background_tasks.add_task(_send_email, next_approver["email"], subject, html)
        next_user = _get_user_by_email(next_approver["email"])
        if next_user:
            create_notification(db, str(next_user["_id"]), "expense_submitted", subject,
                f"Expense estimate from {sp_name} needs your approval", f"/admin/expense_estimates")

    elif next_status == "Pending Payment":
        next_approver = APPROVER_CHAIN[2]
        subject = f"Expense Estimate Cleared for Payment – {sp_name}"
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
            <h2>Expense Estimate – Payment Processing Required</h2>
            <p>Expense estimate from <b>{sp_name}</b> (trip: {trip_date}) has been approved and is ready
            for advance payment processing.</p>
            <ul>
                <li><b>Estimated Total:</b> ₹{est.get('estimated_total', 0):,.2f}</li>
                <li><b>Advance Requested:</b> ₹{est.get('advance_requested', 0):,.2f}</li>
            </ul>
        </div>"""
        background_tasks.add_task(_send_email, next_approver["email"], subject, html)
        next_user = _get_user_by_email(next_approver["email"])
        if next_user:
            create_notification(db, str(next_user["_id"]), "expense_submitted", subject,
                f"Process advance for {sp_name}'s trip on {trip_date}", f"/admin/expense_estimates")

    elif next_status == "Draft":
        advance_msg = "Advance has been released." if update.get("yogesh_advance_released") else "Your trip expenses are approved."
        background_tasks.add_task(
            _notify_salesperson, updated, "expense_advance_released",
            f"Expense Estimate Approved – {approver['label']}",
            f"Your expense estimate for the trip on {trip_date} has been fully approved. {advance_msg} "
            "Please submit your actual expenses after you return.",
        )

    if next_status == "Pending Second Review":
        background_tasks.add_task(
            _notify_salesperson, updated, "expense_approved_stage",
            f"Expense Estimate Approved – Stage 1",
            f"Your expense estimate for trip on {trip_date} has been approved by {approver['label']} and forwarded for second review.",
        )
    elif next_status == "Pending Payment":
        background_tasks.add_task(
            _notify_salesperson, updated, "expense_approved_stage",
            f"Expense Estimate Approved – Stage 2",
            f"Your expense estimate for trip on {trip_date} has been approved by {approver['label']} and is now being processed for advance payment.",
        )

    return serialize_mongo_document(updated)


# ── reject ─────────────────────────────────────────────────────────────────────

@router.post("/{estimate_id}/reject")
async def reject_estimate(
    estimate_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    est = _get_estimate_or_404(estimate_id)
    caller_email = current_user.get("data", {}).get("email") or current_user.get("email", "")
    current_status = est["status"]

    approver = _approver_for_status(current_status)
    if not approver:
        raise HTTPException(status_code=400, detail=f"Estimate cannot be rejected from status '{current_status}'")
    if caller_email != approver["email"]:
        raise HTTPException(status_code=403, detail="You are not the designated approver for this stage")

    body = await request.json()
    reason = body.get("reason", "")
    if not reason:
        raise HTTPException(status_code=400, detail="Rejection reason is required")

    db.expense_estimates.update_one(
        {"_id": ObjectId(estimate_id)},
        {"$set": {"status": "Rejected", "rejection_reason": reason, "updated_at": datetime.datetime.utcnow()}},
    )
    updated = db.expense_estimates.find_one({"_id": ObjectId(estimate_id)})

    trip_date = (est.get("travel_start_date") or "")[:10]
    background_tasks.add_task(
        _notify_salesperson, updated, "expense_rejected",
        "Expense Estimate Rejected",
        f"Your expense estimate for trip on {trip_date} was rejected by {approver['label']}. Reason: {reason}",
    )

    return serialize_mongo_document(updated)


# ── complete settlement ────────────────────────────────────────────────────────

@router.post("/{estimate_id}/complete")
async def complete_settlement(
    estimate_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    est = _get_estimate_or_404(estimate_id)
    if est["status"] != "Submitted":
        raise HTTPException(status_code=400, detail="Settlement can only be completed when status is 'Submitted'")

    body = await request.json()
    approved_total = float(body.get("approved_total") or est.get("actual_total") or 0)
    advance = float(est.get("advance_requested") or 0)
    reimburse = max(0.0, approved_total - advance)
    returned = max(0.0, advance - approved_total)

    db.expense_estimates.update_one(
        {"_id": ObjectId(estimate_id)},
        {"$set": {
            "status": "Completed",
            "approved_total": approved_total,
            "amount_to_reimburse": reimburse,
            "amount_to_return": returned,
            "updated_at": datetime.datetime.utcnow(),
        }},
    )
    updated = db.expense_estimates.find_one({"_id": ObjectId(estimate_id)})

    msg = f"Reimbursement of ₹{reimburse:,.2f} will be processed." if reimburse > 0 else f"Please return ₹{returned:,.2f}." if returned > 0 else "Settlement complete — no amount to exchange."
    background_tasks.add_task(
        _notify_salesperson, updated, "expense_approved_stage",
        "Expense Settlement Completed",
        f"Your expense report has been settled. {msg}",
    )

    return serialize_mongo_document(updated)


# ── Customer analytics helpers for Excel report ───────────────────────────────

def _get_fy_bounds(fy_start_year: int):
    """Return (start_str, end_str) for a financial year starting April of fy_start_year."""
    start = f"{fy_start_year}-04-01"
    end = f"{fy_start_year + 1}-03-31"
    return start, end


def _get_fy_label(fy_start: int) -> str:
    return f"FY {fy_start}-{str(fy_start + 1)[2:]}"


def _customer_stats_for_report(customer_id: str) -> dict:
    """
    Return sales and outstanding balance from the invoices collection.
    - current_yr_sales  : current FY (e.g. 2026-27)
    - last_fy_sales     : last FY    (e.g. 2025-26) — "Last Financial Year" card in analytics
    - prev_fy_sales     : prev FY    (e.g. 2024-25) — "Previous Financial Year" card in analytics
    customer_id may be a 24-char MongoDB _id string or a Zoho contact_id string.
    """
    now = datetime.datetime.now()
    cur_month = now.month
    cur_fy_start = now.year if cur_month >= 4 else now.year - 1
    last_fy_start = cur_fy_start - 1
    prev_fy_start = cur_fy_start - 2

    # Resolve MongoDB _id → Zoho contact_id if needed
    zoho_id = customer_id
    if customer_id and len(customer_id) == 24:
        try:
            cust = db.customers.find_one({"_id": ObjectId(customer_id)}, {"contact_id": 1})
            if cust and cust.get("contact_id"):
                zoho_id = cust["contact_id"]
            else:
                return {"current_yr_sales": 0, "last_fy_sales": 0, "prev_fy_sales": 0, "outstanding_balance": 0}
        except Exception:
            pass

    def fy_cond(fy_s):
        return {"$or": [
            {"$and": [{"$eq": ["$parsedYear", fy_s]}, {"$gte": ["$parsedMonth", 4]}]},
            {"$and": [{"$eq": ["$parsedYear", fy_s + 1]}, {"$lte": ["$parsedMonth", 3]}]},
        ]}

    pipeline = [
        {"$match": {"customer_id": zoho_id, "status": {"$nin": ["void", "draft"]}}},
        {"$addFields": {
            "parsedYear": {"$year": {"$dateFromString": {"dateString": "$date", "onError": None}}},
            "parsedMonth": {"$month": {"$dateFromString": {"dateString": "$date", "onError": None}}},
        }},
        {"$group": {
            "_id": None,
            "currentFYSales": {"$sum": {"$cond": [fy_cond(cur_fy_start), {"$ifNull": ["$total", 0]}, 0]}},
            "lastFYSales":    {"$sum": {"$cond": [fy_cond(last_fy_start), {"$ifNull": ["$total", 0]}, 0]}},
            "prevFYSales":    {"$sum": {"$cond": [fy_cond(prev_fy_start), {"$ifNull": ["$total", 0]}, 0]}},
            "outstandingBalance": {"$sum": {"$cond": [
                {"$in": ["$status", ["sent", "overdue", "partially_paid"]]},
                {"$ifNull": ["$balance", 0]}, 0,
            ]}},
        }},
    ]

    try:
        result = list(db.invoices.aggregate(pipeline, allowDiskUse=True))
        if result:
            r = result[0]
            return {
                "current_yr_sales": round(r.get("currentFYSales") or 0, 2),
                "last_fy_sales": round(r.get("lastFYSales") or 0, 2),
                "prev_fy_sales": round(r.get("prevFYSales") or 0, 2),
                "outstanding_balance": round(r.get("outstandingBalance") or 0, 2),
            }
    except Exception as e:
        print(f"[expense report] customer stats lookup failed for {customer_id}: {e}")

    return {"current_yr_sales": 0, "last_fy_sales": 0, "prev_fy_sales": 0, "outstanding_balance": 0}


# ── Excel report ───────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_report(est: dict) -> io.BytesIO:
    # Enrich existing-customer visits with live FY sales + outstanding balance
    now_dt = datetime.datetime.now()
    _cur_fy_start = now_dt.year if now_dt.month >= 4 else now_dt.year - 1
    _last_fy_start = _cur_fy_start - 1
    _prev_fy_start = _cur_fy_start - 2
    last_fy_label = _get_fy_label(_last_fy_start)   # e.g. "FY 2025-26"
    prev_fy_label = _get_fy_label(_prev_fy_start)   # e.g. "FY 2024-25"

    enriched_visits = []
    for visit in est.get("customer_visits", []):
        cid = visit.get("customer_id")
        if cid and visit.get("customer_type") == "existing":
            stats = _customer_stats_for_report(cid)
            visit = {**visit, **stats}
        enriched_visits.append(visit)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Column widths
    col_widths = {1: 6, 2: 14, 3: 16, 4: 30, 5: 28, 6: 14, 7: 14, 8: 12, 9: 12, 10: 14, 11: 16, 12: 30, 13: 16, 14: 30, 15: 18}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=10)
    label_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    section_fill = _header_fill("DDEEFF")
    header_fill = _header_fill("BDD7EE")
    label_fill = _header_fill("F2F2F2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def merge_write(row, col_start, col_end, value, font=None, fill=None, align=None):
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        cell.border = _thin_border()
        return cell

    def write(row, col, value, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        cell.border = _thin_border()
        return cell

    # ── Title ──────────────────────────────────────────────────────────────────
    r = 1
    merge_write(r, 1, 14, "SALES Team – Outstation TRAVEL & Food EXPENSE REPORT", title_font, _header_fill("1F4E79"), center)
    ws.cell(row=r, column=1).font = Font(bold=True, size=13, color="FFFFFF")
    r += 1
    merge_write(r, 1, 14, "Submit within 7 days of return", label_font, label_fill, left)

    # ── Section 1 ──────────────────────────────────────────────────────────────
    r += 1
    merge_write(r, 1, 14, "  SECTION 1 – EMPLOYEE & TRIP INFORMATION", section_font, section_fill, left)
    ws.row_dimensions[r].height = 18

    def label_value_pair(row, label_col, label, val_col, val_end, value):
        write(row, label_col, label, label_font, label_fill, left)
        ws.merge_cells(start_row=row, start_column=val_col, end_row=row, end_column=val_end)
        write(row, val_col, value, normal_font, align=left)

    r += 1
    label_value_pair(r, 1, "Employee Name", 2, 6, est.get("created_by_name", ""))
    label_value_pair(r, 7, "Travel Start Date", 8, 9, _fmt_date(est.get("travel_start_date")))
    r += 1
    label_value_pair(r, 1, "Employee ID", 2, 6, est.get("employee_id", ""))
    label_value_pair(r, 7, "Travel End Date", 8, 9, _fmt_date(est.get("travel_end_date")))
    r += 1
    label_value_pair(r, 1, "Designation", 2, 6, est.get("designation", ""))
    travel_start = _parse_date(est.get("travel_start_date"))
    travel_end = _parse_date(est.get("travel_end_date"))
    days = ((travel_end - travel_start).days + 1) if travel_start and travel_end else ""
    label_value_pair(r, 7, "Number of Travel Days", 8, 9, days)
    r += 1
    label_value_pair(r, 1, "Department", 2, 6, est.get("department", ""))
    label_value_pair(r, 7, "Purpose of Trip", 8, 14, est.get("purpose_of_trip", ""))
    r += 1
    label_value_pair(r, 1, "Reporting Manager", 2, 6, est.get("reporting_manager", ""))
    label_value_pair(r, 7, "Locations Visited", 8, 14, est.get("locations_visited", ""))
    r += 1
    label_value_pair(r, 1, "Current Location", 2, 6, est.get("current_location", ""))
    label_value_pair(r, 7, "Mode of Travel", 8, 14, est.get("mode_of_travel", ""))

    # ── Section 2B: Visit Summary ──────────────────────────────────────────────
    r += 2
    merge_write(r, 1, 14, "  SECTION 2B – VISIT SUMMARY", section_font, section_fill, left)
    ws.row_dimensions[r].height = 18
    r += 1
    hdrs_2b = ["Metric", "Planned Count", "Actual Count", "Achievement %",
               "Last Visit Potential Customers", "Onboarded", "Orders Received", "Conversion Rate %", "Remarks"]
    for ci, h in enumerate(hdrs_2b, 1):
        write(r, ci, h, header_font, header_fill, center)
    r += 1
    planned_total = int(est.get("planned_existing_visits") or 0) + int(est.get("planned_new_visits") or 0)
    actual_total = int(est.get("actual_existing_visits") or 0) + int(est.get("actual_new_visits") or 0)
    ach_total = f"{actual_total/planned_total*100:.1f}%" if planned_total else "-"
    for col, val in enumerate([
        "No. of Visits", planned_total, actual_total, ach_total, "-", "-", "-", "-", ""
    ], 1):
        write(r, col, val, normal_font, align=center)
    r += 1
    p_ex = int(est.get("planned_existing_visits") or 0)
    a_ex = int(est.get("actual_existing_visits") or 0)
    ach_ex = f"{a_ex/p_ex*100:.1f}%" if p_ex else "-"
    for col, val in enumerate(["Existing Customers", p_ex, a_ex, ach_ex, "", "", "", "", ""], 1):
        write(r, col, val, normal_font, align=center)
    r += 1
    p_new = int(est.get("planned_new_visits") or 0)
    a_new = int(est.get("actual_new_visits") or 0)
    ach_new = f"{a_new/p_new*100:.1f}%" if p_new else "-"
    for col, val in enumerate(["New / Prospect Customers", p_new, a_new, ach_new, "", "", "", "", ""], 1):
        write(r, col, val, normal_font, align=center)

    # ── Section 3: Itemised Expenses ──────────────────────────────────────────
    r += 2
    merge_write(r, 1, 14, "  SECTION 3 – ITEMISED EXPENSE DETAILS", section_font, section_fill, left)
    ws.row_dimensions[r].height = 18

    def write_expense_table(ws, start_row, items, table_label):
        r = start_row
        merge_write(r, 1, 14, table_label, label_font, label_fill, left)
        r += 1
        hdrs = ["SL No", "Date", "Expense Type", "Particulars / Description", "Location / Route",
                "Amount (₹)", "Bill Status", "Bill No.", "Tax (GST)", "Net Amount", "Approved Amount", "Remarks", "DA (₹)", "DA Date", "Bill"]
        for ci, h in enumerate(hdrs, 1):
            write(r, ci, h, header_font, header_fill, center)
            ws.row_dimensions[r].height = 30
        r += 1
        for i, item in enumerate(items, 1):
            row_data = [
                i,
                _fmt_date(item.get("date")),
                item.get("expense_type", ""),
                item.get("description", ""),
                item.get("location_route", ""),
                float(item.get("amount") or 0),
                item.get("bill_status", ""),
                item.get("bill_no", ""),
                float(item.get("tax_gst") or 0),
                float(item.get("amount") or 0) + float(item.get("tax_gst") or 0),
                float(item.get("approved_amount") or 0),
                item.get("remarks", ""),
                float(item.get("daily_allowance") or 0),
                _fmt_date(item.get("da_date")),
            ]
            for ci, val in enumerate(row_data, 1):
                write(r, ci, val, normal_font, align=center)
            bill_url = item.get("bill_url", "")
            if bill_url:
                bc = ws.cell(row=r, column=15, value="View Bill")
                bc.hyperlink = bill_url
                bc.font = Font(color="0563C1", underline="single", size=10)
                bc.border = _thin_border()
                bc.alignment = center
            else:
                write(r, 15, "", normal_font, align=center)
            r += 1
        # Subtotals
        total_amt = sum(float(i.get("amount") or 0) for i in items)
        total_gst = sum(float(i.get("tax_gst") or 0) for i in items)
        total_net = sum(float(i.get("amount") or 0) + float(i.get("tax_gst") or 0) for i in items)
        total_approved = sum(float(i.get("approved_amount") or 0) for i in items)
        total_da = sum(float(i.get("daily_allowance") or 0) for i in items)
        subtotal_row = ["SUBTOTALS", "", "", "", "", total_amt, "", "", total_gst, total_net, total_approved, "", total_da, ""]
        for ci, val in enumerate(subtotal_row, 1):
            write(r, ci, val, Font(bold=True, size=10), label_fill, center)
        r += 1
        # Breakdown by category
        merge_write(r, 1, 14, "  Expense Breakdown by Category", label_font, label_fill, left)
        r += 1
        travel_sum = sum(float(i.get("amount") or 0) for i in items if i.get("expense_type") == "Travel")
        stay_sum = sum(float(i.get("amount") or 0) for i in items if i.get("expense_type") == "Stay")
        for label, val in [
            ("Travel (booked from office)", travel_sum),
            ("Stay (booked from office)", stay_sum),
            ("Other / DA (Food & local travel)", total_da),
        ]:
            write(r, 1, label, label_font, label_fill, left)
            write(r, 2, val, normal_font, align=center)
            r += 1
        return r

    expense_items = est.get("expense_items", [])
    r = write_expense_table(ws, r + 1, expense_items, "Estimated Expenses")

    actual_items = est.get("actual_expense_items", [])
    if actual_items:
        r += 1
        r = write_expense_table(ws, r, actual_items, "Actual Expenses")

    # ── Section 4: Settlement Summary ─────────────────────────────────────────
    r += 1
    merge_write(r, 1, 14, "  SECTION 4 – ADVANCE & SETTLEMENT SUMMARY", section_font, section_fill, left)
    ws.row_dimensions[r].height = 18
    r += 1
    summary_rows = [
        ("Estimated Total Expense", est.get("estimated_total", 0)),
        ("Advance Amount Requested", est.get("advance_requested", 0)),
        ("Actual Total Expense", est.get("actual_total", 0)),
        ("Approved Total Expense", est.get("approved_total", 0)),
        ("Amount to be Reimbursed / Refunded", est.get("amount_to_reimburse", 0)),
        ("Amount to be Returned by Employee", est.get("amount_to_return", 0)),
    ]
    for label, val in summary_rows:
        write(r, 1, label, label_font, label_fill, left)
        write(r, 2, float(val or 0), normal_font, align=center)
        r += 1

    # ── Section 2A: Customer Visit Log ────────────────────────────────────────
    r += 1
    merge_write(r, 1, 14, "  SECTION 2A – CUSTOMER VISIT LOG", section_font, section_fill, left)
    ws.row_dimensions[r].height = 18
    r += 1
    visit_hdrs = ["Date", "Customer Name", "City", "Customer Status",
                  "Current Yr Sales", f"{last_fy_label} Sales", f"{prev_fy_label} Sales",
                  "Outstanding Balance", "Purpose of Visit", "Outcome / Next Action",
                  "Follow-up Date", "Order Value", "Notes"]
    for ci, h in enumerate(visit_hdrs, 1):
        write(r, ci, h, header_font, header_fill, center)
    ws.row_dimensions[r].height = 30
    r += 1
    total_curr = total_last_fy = total_prev_fy = total_outstanding = total_order = 0.0
    for visit in enriched_visits:
        curr = float(visit.get("current_yr_sales") or 0)
        last_fy = float(visit.get("last_fy_sales") or 0)
        prev_fy = float(visit.get("prev_fy_sales") or 0)
        outstanding = float(visit.get("outstanding_balance") or 0)
        order_val = float(visit.get("order_value") or 0)
        total_curr += curr
        total_last_fy += last_fy
        total_prev_fy += prev_fy
        total_outstanding += outstanding
        total_order += order_val
        name = visit.get("customer_name") or visit.get("potential_customer_name", "")
        row_data = [
            _fmt_date(visit.get("date")),
            name,
            visit.get("city", ""),
            visit.get("customer_status", ""),
            curr, last_fy, prev_fy, outstanding,
            visit.get("purpose_of_visit", ""),
            visit.get("outcome", ""),
            _fmt_date(visit.get("follow_up_date")),
            order_val,
            visit.get("notes", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            write(r, ci, val, normal_font, align=center)
        r += 1
    # Totals row
    total_row = ["TOTAL", "", "", "", total_curr, total_last_fy, total_prev_fy, total_outstanding, "", "", "", total_order, ""]
    for ci, val in enumerate(total_row, 1):
        write(r, ci, val, Font(bold=True, size=10), label_fill, center)

    # Status block
    r += 2
    merge_write(r, 1, 4, f"Status: {est.get('status', '')}", label_font, label_fill, left)
    r += 1
    for approver in [
        ("First Review (Rahul)", est.get("rahul_approved_at"), est.get("rahul_remarks")),
        ("Second Review (Amit)", est.get("amit_approved_at"), est.get("amit_remarks")),
        ("Payment (Yogesh)", est.get("yogesh_processed_at"), est.get("yogesh_remarks")),
    ]:
        label, ts, remarks = approver
        val = f"Approved on {_fmt_date(ts)}" if ts else "Pending"
        if remarks:
            val += f" — {remarks}"
        write(r, 1, label, label_font, label_fill, left)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=14)
        write(r, 2, val, normal_font, align=left)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fmt_date(val) -> str:
    if not val:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%d-%b-%Y")
    if isinstance(val, str):
        try:
            return datetime.datetime.fromisoformat(val).strftime("%d-%b-%Y")
        except Exception:
            return val
    return str(val)


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, str):
        try:
            return datetime.datetime.fromisoformat(val)
        except Exception:
            return None
    return None


@router.get("/{estimate_id}/report")
def download_report(estimate_id: str, current_user: dict = Depends(get_current_user)):
    est = _get_estimate_or_404(estimate_id)
    buf = _write_report(est)
    name = est.get("created_by_name", "salesperson").replace(" ", "_")
    trip = (est.get("travel_start_date") or "")[:10]
    filename = f"Expense_Report_{name}_{trip}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
