import io, requests, os, smtplib
import pandas as pd
from functools import lru_cache
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from dotenv import load_dotenv
from datetime import date, datetime
from ..config.whatsapp import send_whatsapp

load_dotenv()
org_id = os.getenv("ORG_ID")
PURCHASE_ORDER_URL = os.getenv("PURCHASE_ORDER_URL")
PURCHASE_URL = os.getenv("PURCHASE_URL")
ITEM_URL = os.getenv("ITEM_URL")
INVENTORY_URL = os.getenv("INVENTORY_URL")
BOOKS_URL = os.getenv("BOOKS_URL")
clientId = os.getenv("CLIENT_ID")
clientSecret = os.getenv("CLIENT_SECRET")
grantType = os.getenv("GRANT_TYPE")
inventory_refresh_token = os.getenv("INVENTORY_REFRESH_TOKEN")
books_refresh_token = os.getenv("BOOKS_REFRESH_TOKEN")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = os.getenv("SMTP_PORT")
SENDER_EMAIL = os.getenv("SMTP_USERNAME")  # Use your email
SENDER_PASSWORD = os.getenv("SMTP_PASSWORD")  # Use your email app password


def validate_file(file) -> dict:
    """
    Checks if the given Excel file contains both 'PL' and 'CI' sheets.

    Args:
        file (BytesIO): In-memory file data (from form data).

    Returns:
        dict: Response indicating whether the required sheets are present or not.
    """
    try:
        # Load the workbook from the file
        wb = load_workbook(file)

        # Get all sheet names
        sheet_names = wb.sheetnames

        # Check if both 'PL' and 'CI' sheets are present
        if "PL" not in sheet_names or "CI" not in sheet_names:
            missing_sheets = []
            if "PL" not in sheet_names:
                missing_sheets.append("PL")
            if "CI" not in sheet_names:
                missing_sheets.append("CI")

            # Return an error response if any sheet is missing
            return {
                "status": "error",
                "message": f"Missing sheets: {', '.join(missing_sheets)}",
            }

        # If both sheets are found, return a success response
        return {"status": "success", "message": "Both PL and CI sheets are present."}

    except Exception as e:
        # Handle any other errors (e.g., invalid file format)
        return {"status": "error", "message": f"An error occurred: {str(e)}"}


def send_email(subject, body, email, cc):
    """Send email with multiple in-memory attachments and CC."""
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = email
    msg["Subject"] = subject
    if cc:
        msg["Cc"] = cc  # Add CC header to the email
    msg.attach(MIMEText(body, "plain"))

    # Combine primary recipient and CC recipients for sending
    recipient_list = [email] + [cc_email.strip() for cc_email in cc.split(",")]

    # Send the email
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()  # Encrypt connection
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, recipient_list, msg.as_string())
        server.quit()
        print(f"Email sent to {email} with CC: {cc}")
    except Exception as e:
        print(f"Failed to send email: {e}")


def send_email_with_attachments_in_memory(workbook, subject, body, filename, email):
    """Send email with multiple in-memory attachments."""
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # Attach each in-memory workbook
    part = MIMEBase("application", "octet-stream")
    part.set_payload(workbook)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}.xlsx"')
    msg.attach(part)

    # Send the email
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()  # Encrypt connection
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, email, msg.as_string())
        server.quit()
        print(f"Email sent to {email} with in-memory attachments.")
    except Exception as e:
        print(f"Failed to send email: {e}")


def get_access_token(tkn: str):
    r = None
    access_token = ""
    if tkn == "inventory":
        r = requests.post(
            INVENTORY_URL.format(
                clientId=clientId,
                clientSecret=clientSecret,
                grantType=grantType,
                inventory_refresh_token=inventory_refresh_token,
            )
        )
    elif tkn == "books":
        r = requests.post(
            BOOKS_URL.format(
                clientId=clientId,
                clientSecret=clientSecret,
                grantType=grantType,
                books_refresh_token=books_refresh_token,
            )
        )
    else:
        print("missing token type")
        return
    access_token = str(r.json().get("access_token", ""))
    print(f"Got {tkn.capitalize()} Access Token: {access_token[-4:]}")
    return access_token


company_name = "Pettingzoo"


def save_combined_sheet(matched_ci, unmatched_ci, matched_pl, unmatched_pl):
    """
    Saves four DataFrames to two sheets in a combined Excel file in memory.

    Args:
        matched_ci (pandas.DataFrame): DataFrame containing matched CI data.
        unmatched_ci (pandas.DataFrame): DataFrame containing unmatched CI data.
        matched_pl (pandas.DataFrame): DataFrame containing matched PL data.
        unmatched_pl (pandas.DataFrame): DataFrame containing unmatched PL data.

    Returns:
        bytes: The combined Excel file content in memory.
    """

    try:
        # Create a workbook
        wb = Workbook()

        # Sheet 1: Write matched and unmatched CI data
        ws1 = wb.active
        ws1.title = "CI Data"

        # Add "Matched CI" title
        ws1.append(["Matched CI"])
        for row in dataframe_to_rows(matched_ci, index=None, header=True):
            ws1.append(row)

        # Add a gap of two rows
        ws1.append([])
        ws1.append([])

        # Add "Unmatched CI" title
        ws1.append(["Unmatched CI"])
        for row in dataframe_to_rows(unmatched_ci, index=None, header=True):
            ws1.append(row)

        # Sheet 2: Write matched and unmatched PL data
        ws2 = wb.create_sheet(title="PL Data")

        # Add "Matched PL" title
        ws2.append(["Matched PL"])
        for row in dataframe_to_rows(matched_pl, index=None, header=True):
            ws2.append(row)

        # Add a gap of two rows
        ws2.append([])
        ws2.append([])

        # Add "Unmatched PL" title
        ws2.append(["Unmatched PL"])
        for row in dataframe_to_rows(unmatched_pl, index=None, header=True):
            ws2.append(row)

        # Create an in-memory buffer
        output_buffer = io.BytesIO()

        # Save the workbook to the buffer
        wb.save(output_buffer)

        # Reset the buffer position to the beginning
        output_buffer.seek(0)

        return output_buffer.getvalue()

    except Exception as e:
        print(f"Error saving combined sheet: {e}")
        return None


@lru_cache(maxsize=None)
def compare_strings(s1, s2):
    # remove whitespace, double spaces, hypens and brackets
    s1 = str(s1).replace(",", "").replace(" ", "").replace("--", "").casefold()
    s2 = str(s2).replace(",", "").replace(" ", "").replace("--", "").casefold()
    # compare strings
    if s1 == s2:
        return True
    else:
        return False


def extract_table_data(file_path, sheet_name):
    """
    Extracts table data from a given sheet using pandas.
    Assumes that the table starts after the specified start_row.
    """
    # Read the sheet starting from the specified row
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # Drop rows that are entirely NaN
    df = df.dropna(how="all")

    # Optionally, reset the index
    df.reset_index(drop=True, inplace=True)

    return df


def get_purchase_orders(items):
    po = []
    access_token = get_access_token("books")
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    p = requests.get(
        url=PURCHASE_URL.format(org_id=org_id, search_text=company_name, page=1),
        headers=headers,
    )
    po.extend([x for x in p.json()["purchaseorders"] if x["status"] != "draft"])
    has_more_pages = bool(p.json()["page_context"]["has_more_page"])
    if has_more_pages:
        print("Purchase orders have More Pages")
        page = 2
        while True:
            p = requests.get(
                url=PURCHASE_URL.format(
                    org_id=org_id, search_text=company_name, page=page
                ),
                headers=headers,
            )
            response = p.json()
            po.extend([x for x in p.json()["purchaseorders"] if x["status"] != "draft"])
            has_more = bool(p.json()["page_context"]["has_more_page"])
            page += 1
            if not has_more:
                break
    else:
        print("No More Purchase Orders")
    found_items = []
    found_names = set()  # Use a set for quick lookups
    print("Processing POs")
    # Process each purchase order
    for purchase_order in po:
        po_id = purchase_order.get("purchaseorder_id")
        if not po_id:
            continue  # Skip if purchaseorder_id is missing

        # Fetch detailed purchase order items
        response = requests.get(
            url=PURCHASE_ORDER_URL.format(org_id=org_id, purchase_order_id=po_id),
            headers=headers,
        )
        purchase_order_items = (
            response.json().get("purchaseorder", {}).get("line_items", [])
        )

        # Check for matching items in line_items
        for item in items:
            for line_item in purchase_order_items:
                item_name = line_item.get("name")
                rate = line_item.get("rate")

                if (
                    compare_strings(item_name, item["name"])
                    and item_name not in found_names
                ):
                    found_items.append({"rate": rate, "name": item_name})
                    found_names.add(item_name)
                    break
    for item in items:
        if item["name"] not in found_names:
            found_items.append({"rate": 0, "name": item["name"]})
    print("Done Processing POs")
    return found_items


def process_upload(input_file, email):
    access_token = get_access_token("books")
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    # Extract table data from both sheets
    input_file.seek(0)
    input_file = io.BytesIO(input_file.read())

    pl_sheet = extract_table_data(input_file, "PL")
    ci_sheet = extract_table_data(input_file, "CI")
    print(len(pl_sheet["Name"]))
    pl_data = [x for x in pl_sheet["Name"]]

    ci_data = [
        {
            "name": row["Name"],
            "hsn": str(int(row["HSN"])),
            "price": row["Price"],
        }
        for _, row in ci_sheet.iterrows()
        if isinstance(row["Name"], str)
    ]
    # Check if data is matching with the data on zoho and print a list of all found items, and not found items
    matched_pl, unmatched_pl, matched_ci, unmatched_ci = [], [], [], []
    data = get_purchase_orders(ci_data)
    print("Processing PL Data", len(pl_data))
    # fetch all items from PL sheet on zoho
    for item in pl_data:
        x = requests.get(
            url=ITEM_URL.format(org_id=org_id, search_text=item), headers=headers
        )
        product = x.json()["items"]
        if len(product) > 0:
            product = product[0]
            product_name = product.get("item_name")
            if compare_strings(item, product_name):
                matched_pl.append({"name": product_name})
            else:
                unmatched_pl.append({"name": product_name})
    print("Done Processing PL Data")
    print("Processing CI Data", len(ci_data))
    for item in ci_data:
        name = item.get("name")
        code = item.get("hsn")
        price = item.get("price")
        x = requests.get(
            url=ITEM_URL.format(org_id=org_id, search_text=name), headers=headers
        )
        product = x.json()["items"]
        if len(product) > 0:
            product = product[0]
            product_name = product.get("item_name")
            product_code = product.get("hsn_or_sac")
            product_price = next(
                (
                    entry["rate"]
                    for entry in data
                    if compare_strings(entry["name"], product_name)
                ),
                0,
            )
            if (
                compare_strings(name, product_name)
                and compare_strings(code, product_code)
                and compare_strings(price, product_price)
            ):
                matched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                    }
                )
            else:
                reasons = []
                if not compare_strings(name, product_name):
                    reasons.append(f"Name {name} not matched with {product_name} ")
                if not compare_strings(code, product_code):
                    reasons.append(f"HSN {code} not matched with {product_code} ")
                if not compare_strings(price, product_price):
                    reasons.append(f"Price {price} not matched with {product_price} ")

                reason = "; ".join(reasons)
                unmatched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                        "reason": reason,
                    }
                )
        else:
            reason = f"{name} Not found in zoho"
            unmatched_ci.append(
                {
                    "name": name,
                    "hsn": code,
                    "price": price,
                    "reason": reason,
                }
            )

    print("Done Processing CI Data")

    matched_pl_df = pd.DataFrame(sorted(matched_pl, key=lambda x: str(x["name"])))
    unmatched_pl_df = pd.DataFrame(sorted(unmatched_pl, key=lambda x: str(x["name"])))
    matched_ci_df = pd.DataFrame(sorted(matched_ci, key=lambda x: int(x["hsn"])))
    unmatched_ci_df = pd.DataFrame(sorted(unmatched_ci, key=lambda x: int(x["hsn"])))

    workbook = save_combined_sheet(
        matched_ci_df, unmatched_ci_df, matched_pl_df, unmatched_pl_df
    )
    filename = "CI & PL Data"
    subject = "CI & PL Workbook"
    body = "Please find the attached CI and PL verification files."
    send_email_with_attachments_in_memory(workbook, subject, body, filename, email)
    return 1


def notify_all_salespeople(db, template, params):
    all_salespeople = db.users.find({"status": "active", "role": "sales_person"})
    for salesperson in all_salespeople:
        phone = salesperson.get("phone")
        name = salesperson.get("first_name")
        template_doc = {**template}
        parameters = {"name": name, **params}
        if phone != "":
            x = send_whatsapp(phone, template_doc, parameters)
            print(x)
    pass


def notify_all_sales_admins(db, template, params):
    all_sales_admins = db.users.find({"status": "active", "role": "sales_admin"})
    for sales_admin in all_sales_admins:
        phone = sales_admin.get("phone")
        name = sales_admin.get("first_name")
        template_doc = {**template}
        parameters = {"name": name, **params}
        if phone != "":
            send_whatsapp(phone, template_doc, parameters)
    pass


def notify_office_coordinator_and_sales_admins(db, template, params):
    office_coordinator = db.users.find_one(
        {"email": "pupscribeoffcoordinator@gmail.com"}
    )
    sales_admin = db.users.find_one({"email": "barksalesamit@gmail.com"})
    for person in [office_coordinator, sales_admin]:
        phone = person.get("phone")
        name = person.get("first_name")
        template_doc = {**template}
        parameters = {"name": name, **params}
        if phone != "":
            send_whatsapp(phone, template_doc, parameters)
    pass


def notify_sales_admin(db, template, params):
    sales_admin = db.users.find_one({"email": "barksalesamit@gmail.com"})
    phone = sales_admin.get("phone")
    name = sales_admin.get("first_name")
    template_doc = {**template}
    parameters = {"admin_name": name, **params}
    if phone != "":
        send_whatsapp(phone, template_doc, parameters)
    pass


def notify_person(template, params, person):
    phone = person.get("phone")
    name = person.get("first_name")
    template_doc = {**template}
    parameters = {"sales_person_name": name, **params}
    if phone != "":
        send_whatsapp(phone, template_doc, parameters)
    pass


# Lightweight field set for listing overdue invoices. Deliberately EXCLUDES
# line_items (which can be very large and would make us drag the whole
# collection's item arrays over the wire just to sort/count/paginate). Callers
# that need line_items (e.g. the admin drawer's Products table) should fetch
# them for the current page's invoices only.
_OVERDUE_INVOICE_PROJECTION = {
    "invoice_id": 1,
    "invoice_number": 1,
    "date": 1,
    "due_date": 1,
    "status": 1,
    "customer_id": 1,
    "customer_name": 1,
    "total": 1,
    "balance": 1,
    "cf_sales_person": 1,
    "salesperson_name": 1,
    "created_by_name": 1,
    "created_at": 1,
    "invoice_url": 1,
}


def fetch_overdue_invoices(db, extra_query: dict = None, projection: dict = None):
    """
    Return all unpaid, overdue (<=365 days), non-void invoices with a
    computed `overdue_by_days` field, sorted ascending by that field.

    due_date is stored inconsistently as either a string or a real Date
    across the collection. A single query with `$or` on `{$type: 'string'}`
    / `{$type: 'date'}` makes Mongo's planner pick one weak index for both
    branches (observed 8-50s on this cluster). Splitting into two
    homogeneous-type queries, each hinted to its own best index, and merging
    client-side is 5-10x faster in practice.

    By default only a lightweight projection is fetched (no line_items) so we
    don't pull the entire matched set's item arrays across the wire just to
    sort and paginate. Pass an explicit `projection` to override.
    """
    today_str = date.today().isoformat()
    today_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    status_filter = {"status": {"$nin": ["paid", "void"]}}
    proj = projection if projection is not None else _OVERDUE_INVOICE_PROJECTION

    q_date = {"due_date": {"$type": "date", "$lt": today_date}, **status_filter}
    q_str = {"due_date": {"$type": "string", "$lt": today_str}, **status_filter}
    if extra_query:
        q_date = {"$and": [q_date, extra_query]}
        q_str = {"$and": [q_str, extra_query]}

    docs = list(db.invoices.find(q_date, proj).hint("due_date_1_status_1"))
    docs += list(db.invoices.find(q_str, proj).hint("status_1"))

    now = datetime.now()
    result = []
    for doc in docs:
        due = doc.get("due_date")
        if isinstance(due, str):
            try:
                due = datetime.fromisoformat(due)
            except ValueError:
                continue
        overdue_by_days = (now - due).days
        if overdue_by_days > 365:
            continue
        doc["overdue_by_days"] = overdue_by_days
        result.append(doc)

    result.sort(key=lambda d: d["overdue_by_days"])
    return result


# Minimal fields shown for each credit note associated with an invoice.
_ASSOCIATED_CREDIT_NOTE_PROJECTION = {
    "creditnote_id": 1,
    "creditnote_number": 1,
    "invoice_id": 1,
    "customer_id": 1,
    "date": 1,
    "status": 1,
    "total": 1,
    "balance": 1,
}

_CLOSED_CN_STATUSES = ("void", "closed")


def fetch_associated_credit_notes(db, invoices):
    """
    Given a list of invoice docs (each with `invoice_id` and `customer_id`),
    return a dict mapping each invoice's Zoho `invoice_id` -> list of its
    associated credit notes.

    "Associated" is the union of:
      * the customer's OPEN credit notes (status not void/closed) — these are
        the notes that make up the customer-level "Open Credit Note Amt.", so
        the two stay consistent; and
      * any credit note linked to that specific invoice via invoice_id (any
        status), for full per-invoice tracking.

    Notes are de-duplicated per invoice and sorted newest-first by date.
    """
    invoice_ids = [i.get("invoice_id") for i in invoices if i.get("invoice_id")]
    customer_ids = [i.get("customer_id") for i in invoices if i.get("customer_id")]
    if not invoice_ids and not customer_ids:
        return {}

    query = {
        "$or": [
            {
                "customer_id": {"$in": customer_ids},
                "status": {"$nin": list(_CLOSED_CN_STATUSES)},
            },
            {"invoice_id": {"$in": invoice_ids}},
        ]
    }

    open_by_customer = {}
    by_invoice = {}
    for cn in db.credit_notes.find(query, _ASSOCIATED_CREDIT_NOTE_PROJECTION):
        if cn.get("status") not in _CLOSED_CN_STATUSES:
            open_by_customer.setdefault(cn.get("customer_id"), []).append(cn)
        if cn.get("invoice_id"):
            by_invoice.setdefault(cn.get("invoice_id"), []).append(cn)

    result = {}
    for inv in invoices:
        inv_id = inv.get("invoice_id")
        cust_id = inv.get("customer_id")
        seen = {}
        for cn in open_by_customer.get(cust_id, []) + by_invoice.get(inv_id, []):
            seen[cn.get("creditnote_id")] = cn
        notes = list(seen.values())
        notes.sort(key=lambda n: n.get("date") or "", reverse=True)
        result[inv_id] = notes
    return result
