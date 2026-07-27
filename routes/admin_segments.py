"""
Customer segmentation for WhatsApp campaigns (mounted at /admin/segments).

A segment is a saved *rule*; resolving it returns the concrete list of recipients
(phone + name + context). Rules are built on the existing customer-analytics
machinery so they reuse the same tier / dormancy / billing definitions the rest
of the admin already trusts.

Two sources:
  - "b2b": resolved from `invoices` + `customers` via build_customer_analytics_pipeline
           (tier, dormancy, salesperson, brand, billing thresholds).
  - "b2c": resolved from the `chatbot_customers` registry (consumers who messaged us).

Phase 3 (campaigns) calls resolve_segment_rule() to get the recipient list.
"""
import datetime
import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Body, HTTPException, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId

from ..config.root import get_database, serialize_mongo_document
from .admin_customer_analytics import (
    build_customer_analytics_pipeline,
    _get_current_date_info,
    _build_match_and_filters,
)

router = APIRouter()

db = get_database()
segments_col = db["segments"]
customers_col = db["customers"]
products_col = db["products"]
invoices_col = db["invoices"]
chatbot_customers_col = db["chatbot_customers"]

DORMANCY_OPTIONS = {
    "all", "last_month", "last_45_days", "last_2_months", "last_3_months",
    "not_last_month", "not_last_45_days", "not_last_2_months", "not_last_3_months",
}

# dormancy value -> (analytics flag, expected value)
DORMANCY_FLAGS = {
    "last_month": ("hasBilledLastMonth", True),
    "last_45_days": ("hasBilledLast45Days", True),
    "last_2_months": ("hasBilledLast2Months", True),
    "last_3_months": ("hasBilledLast3Months", True),
    "not_last_month": ("hasBilledLastMonth", False),
    "not_last_45_days": ("hasBilledLast45Days", False),
    "not_last_2_months": ("hasBilledLast2Months", False),
    "not_last_3_months": ("hasBilledLast3Months", False),
}

BILLED_FLAGS = (
    "hasBilledLastMonth",
    "hasBilledLast45Days",
    "hasBilledLast2Months",
    "hasBilledLast3Months",
)


def _now():
    return datetime.datetime.now()


def _last10(phone) -> str:
    digits = "".join(ch for ch in str(phone or "") if ch.isdigit())
    return digits[-10:] if len(digits) >= 10 else digits


def _split_salespersons(value, known_codes=None) -> list:
    """`salesPerson` on an analytics row is a code, a list of codes, or a
    comma/slash separated string of codes ("SP8, SP22"). Normalise to a list.

    Some codes legitimately contain a comma ("Amazon,Flipkart And other ecom
    Platforms"), so a value that is already a known code is never split."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        parts = []
        for v in value:
            parts.extend(_split_salespersons(v, known_codes))
        return parts
    text = str(value).strip()
    if not text:
        return []
    if known_codes and text.lower() in known_codes:
        return [text]
    parts = [p.strip() for p in re.split(r"[,/|]", text) if p.strip()]
    if not known_codes:
        return parts
    # Re-join consecutive fragments that together form a known code.
    merged, i = [], 0
    while i < len(parts):
        for span in range(min(4, len(parts) - i), 0, -1):
            candidate = ",".join(parts[i : i + span])
            if span == 1 or candidate.lower() in known_codes:
                merged.append(candidate)
                i += span
                break
    return merged


def _billing_bound(value):
    """Normalise a billing bound. Billing is a rupee total, so anything at or below
    zero (including a negative from an older saved rule) means "no limit" — a
    literal `<= 0` upper bound would silently empty the audience."""
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _sp_display_name(user: dict) -> str:
    """Label for a salesperson. Some Zoho-imported users have a literal "-" name
    (non-person buckets like "Company customers") — fall back to the code."""
    for candidate in (user.get("name"), user.get("first_name")):
        candidate = (candidate or "").strip()
        if candidate and candidate != "-":
            return candidate
    return (user.get("code") or "").strip()


def _salesperson_maps():
    """Return (code -> name, alias -> code) for every salesperson user.

    Segment rules store salesperson *codes* (the value that actually appears on
    invoices); older rules stored first names, so names are accepted as aliases.
    """
    code_to_name = {}
    alias_to_code = {}
    for u in db.users.find(
        {"role": "sales_person"},
        {"first_name": 1, "name": 1, "code": 1},
    ):
        code = (u.get("code") or "").strip()
        name = _sp_display_name(u)
        if not code:
            continue
        code_to_name[code.lower()] = name or code
        for alias in (code, name, u.get("first_name")):
            alias = str(alias or "").strip()
            if alias and alias != "-":
                alias_to_code.setdefault(alias.lower(), code.lower())
    return code_to_name, alias_to_code


# ---------------------------------------------------------------------------
# Brand helper: contact_ids that purchased any of the given brands (since 2023-04)
# ---------------------------------------------------------------------------

def _contact_ids_for_brands(brands: list) -> set:
    if not brands:
        return set()
    # item_id -> brand map, then the set of item_ids belonging to the requested brands.
    wanted = {b.strip() for b in brands if b and b.strip()}
    item_ids = [
        p.get("item_id")
        for p in products_col.find(
            {"brand": {"$in": list(wanted)}}, {"item_id": 1, "_id": 0}
        )
        if p.get("item_id")
    ]
    if not item_ids:
        return set()
    pipeline = [
        {"$match": {"date": {"$gte": "2023-04-01"}, "status": {"$nin": ["void", "draft"]}}},
        {"$unwind": "$line_items"},
        {"$match": {"line_items.item_id": {"$in": item_ids}}},
        {"$group": {"_id": "$customer_id"}},
    ]
    return {r["_id"] for r in invoices_col.aggregate(pipeline, allowDiskUse=True) if r.get("_id")}


# ---------------------------------------------------------------------------
# Core resolver (also imported by the campaigns route in Phase 3)
# ---------------------------------------------------------------------------

def resolve_segment_rule(source: str, rule: dict) -> list:
    """Return a de-duplicated list of recipients for a segment rule.

    Each recipient: {phone, name, customerId?, companyName?, tier?, lastBillDate?,
    salesPerson?, billingCurrentFY?}. Recipients without a usable phone are dropped.
    """
    return resolve_segment_audience(source, rule)["recipients"]


def resolve_segment_audience(source: str, rule: dict) -> dict:
    """Resolve a rule into {recipients, matched, without_phone}.

    `matched` is every customer/contact the rule selects; `recipients` is the
    subset that has a usable phone number (what a campaign can actually send to).
    """
    source = (source or "b2b").lower()
    rule = rule or {}

    if source == "b2c":
        return _resolve_b2c(rule)
    return _resolve_b2b(rule)


def _resolve_b2c(rule: dict) -> dict:
    query = {}
    if rule.get("only_non_b2b"):
        query["is_b2b"] = False
    if rule.get("reviewed_only"):
        query["reviewed"] = True
    matched = []
    recipients = []
    seen = set()
    for c in chatbot_customers_col.find(query, {"phone": 1, "name": 1, "is_b2b": 1}):
        phone = c.get("phone")
        tail = _last10(phone)
        if tail and tail in seen:
            continue
        if tail:
            seen.add(tail)
        row = {
            "phone": phone,
            "name": c.get("name"),
            "source": "b2c",
            "is_b2b": c.get("is_b2b", False),
        }
        matched.append(row)
        if tail:
            recipients.append(row)
    return {
        "recipients": recipients,
        "matched": matched,
        "without_phone": len(matched) - len(recipients),
    }


def _resolve_b2b(rule: dict) -> dict:
    tier = rule.get("tier")
    dormancy = rule.get("dormancy") or "all"
    if dormancy not in DORMANCY_OPTIONS:
        raise HTTPException(status_code=400, detail=f"Invalid dormancy; must be one of {sorted(DORMANCY_OPTIONS)}")
    brands = rule.get("brands") or []
    min_billing = _billing_bound(rule.get("min_billing_current_fy"))
    max_billing = _billing_bound(rule.get("max_billing_current_fy"))

    code_to_name, alias_to_code = _salesperson_maps()
    wanted_sps = set()
    for s in rule.get("salespersons") or []:
        key = str(s or "").strip().lower()
        if key:
            wanted_sps.add(alias_to_code.get(key, key))

    match_stage, customer_status_match_stage, _, sales_person_logic = _build_match_and_filters(
        status="all", tier=tier, sort_by=False
    )
    # The analytics pipeline groups by customer *and shipping address*, so one
    # customer can produce several rows. Resolve with no recency filter and
    # collapse per customer below, otherwise a customer billed recently at one
    # address would also land in the "dormant" bucket via their other address
    # (and billing totals would be per-address rather than per-customer).
    pipeline = build_customer_analytics_pipeline(
        match_stage=match_stage,
        customer_status_match_stage=customer_status_match_stage,
        sales_person_logic=sales_person_logic,
        due_status="all",
        last_billed="all",
        current_date_info=_get_current_date_info(),
        include_all_invoices=False,
    )
    rows = list(invoices_col.aggregate(pipeline, allowDiskUse=True))

    # ---- collapse address-level rows into one entry per customer ----
    customers: dict = {}
    for r in rows:
        cid = r.get("customerId")
        if not cid:
            continue
        entry = customers.get(cid)
        if entry is None:
            entry = {
                "customerId": cid,
                "customerName": r.get("customerName"),
                "companyName": r.get("companyName"),
                "tier": r.get("tier"),
                "lastBillDate": r.get("lastBillDate"),
                "billing": 0.0,
                "salesPersonCodes": [],
                **{f: False for f in BILLED_FLAGS},
            }
            customers[cid] = entry
        entry["billing"] += r.get("billingTillDateCurrentYear") or 0
        for flag in BILLED_FLAGS:
            entry[flag] = entry[flag] or bool(r.get(flag))
        last = r.get("lastBillDate")
        if last and (not entry["lastBillDate"] or last > entry["lastBillDate"]):
            entry["lastBillDate"] = last
        for code in _split_salespersons(r.get("salesPerson"), code_to_name):
            if code not in entry["salesPersonCodes"]:
                entry["salesPersonCodes"].append(code)
        entry["tier"] = entry["tier"] or r.get("tier")
        entry["companyName"] = entry["companyName"] or r.get("companyName")

    brand_ids = _contact_ids_for_brands(brands) if brands else None
    dormancy_flag = DORMANCY_FLAGS.get(dormancy)

    contact_ids = list(customers.keys())
    phone_map = {}
    for cust in customers_col.find(
        {"contact_id": {"$in": contact_ids}},
        {"contact_id": 1, "phone": 1, "mobile": 1, "first_name": 1, "company_name": 1},
    ):
        phone_map[cust.get("contact_id")] = cust

    matched = []
    recipients = []
    seen = set()
    for cid, entry in customers.items():
        if brand_ids is not None and cid not in brand_ids:
            continue
        if dormancy_flag is not None:
            flag, expected = dormancy_flag
            if bool(entry[flag]) is not expected:
                continue
        if wanted_sps:
            codes = {c.lower() for c in entry["salesPersonCodes"]}
            if not (codes & wanted_sps):
                continue
        billing = entry["billing"]
        if min_billing is not None and billing < min_billing:
            continue
        if max_billing is not None and billing > max_billing:
            continue

        cust = phone_map.get(cid, {})
        phone = cust.get("phone") or cust.get("mobile")
        tail = _last10(phone)
        if tail and tail in seen:
            continue
        if tail:
            seen.add(tail)

        codes = entry["salesPersonCodes"]
        row = {
            "phone": phone,
            "name": entry["customerName"] or cust.get("first_name"),
            "customerId": cid,
            "companyName": entry["companyName"] or cust.get("company_name"),
            "tier": entry["tier"],
            "lastBillDate": entry["lastBillDate"],
            "salesPerson": ", ".join(codes),
            "salesPersonName": ", ".join(
                code_to_name.get(c.lower(), c) for c in codes
            ),
            "billingCurrentFY": round(billing, 2),
            "source": "b2b",
        }
        matched.append(row)
        if tail:
            recipients.append(row)

    matched.sort(key=lambda r: r["billingCurrentFY"], reverse=True)
    recipients.sort(key=lambda r: r["billingCurrentFY"], reverse=True)
    return {
        "recipients": recipients,
        "matched": matched,
        "without_phone": len(matched) - len(recipients),
    }


# ---------------------------------------------------------------------------
# Filter option helpers (for building rules in the UI)
# ---------------------------------------------------------------------------

@router.get("/options")
def get_filter_options():
    """Brands, tiers, salespeople and dormancy options the UI needs to build rules."""
    brands = sorted({
        b for b in products_col.distinct("brand") if b and str(b).strip()
    })
    # Rules match on the salesperson *code*, because that is what invoices carry
    # (e.g. "SP8"); the name is only a label for the UI.
    salespeople = sorted(
        (
            {
                "code": (u.get("code") or "").strip(),
                "name": _sp_display_name(u),
                "status": u.get("status", "active"),
            }
            for u in db.users.find(
                {"role": "sales_person"},
                {"first_name": 1, "name": 1, "code": 1, "status": 1},
            )
            if (u.get("code") or "").strip()
        ),
        key=lambda s: s["name"].lower(),
    )
    return {
        "brands": brands,
        "tiers": ["A", "B", "C"],
        "salespeople": salespeople,
        "dormancy_options": sorted(DORMANCY_OPTIONS),
        "sources": ["b2b", "b2c"],
    }


# ---------------------------------------------------------------------------
# Resolve (preview, no save) + resolve saved segment
# ---------------------------------------------------------------------------

@router.post("/resolve")
def resolve_preview(payload: dict = Body(...)):
    """Resolve an unsaved rule and return a count + a sample of recipients."""
    source = payload.get("source", "b2b")
    rule = payload.get("rule", {})
    sample_size = int(payload.get("sample_size", 100))
    audience = resolve_segment_audience(source, rule)
    return _audience_response(audience, sample_size)


@router.post("/{segment_id}/resolve")
def resolve_saved(segment_id: str, sample_size: int = Query(100, le=5000)):
    seg = segments_col.find_one({"_id": _oid(segment_id)})
    if not seg:
        raise HTTPException(status_code=404, detail="Segment not found")
    audience = resolve_segment_audience(seg.get("source", "b2b"), seg.get("rule", {}))
    # Cache the last resolved count for quick display in lists.
    segments_col.update_one(
        {"_id": seg["_id"]},
        {
            "$set": {
                "last_resolved_count": len(audience["recipients"]),
                "last_matched_count": len(audience["matched"]),
                "last_resolved_at": _now(),
            }
        },
    )
    return _audience_response(audience, sample_size)


def _audience_response(audience: dict, sample_size: int) -> dict:
    recipients = audience["recipients"]
    matched = audience["matched"]
    return {
        # `count` stays the reachable-recipient count (what campaigns send to).
        "count": len(recipients),
        "total_matched": len(matched),
        "without_phone": audience["without_phone"],
        "sample": recipients[:sample_size],
    }


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

B2B_COLUMNS = [
    ("Customer Name", "name"),
    ("Company Name", "companyName"),
    ("Phone", "phone"),
    ("Sales Person", "salesPersonName"),
    ("Sales Person Code", "salesPerson"),
    ("Tier", "tier"),
    ("Last Bill Date", "lastBillDate"),
    ("Billing Current FY", "billingCurrentFY"),
    ("Customer ID", "customerId"),
]

B2C_COLUMNS = [
    ("Name", "name"),
    ("Phone", "phone"),
    ("Matched to B2B", "is_b2b"),
]


def _build_audience_workbook(source: str, rule: dict, title: str) -> BytesIO:
    source = (source or "b2b").lower()
    audience = resolve_segment_audience(source, rule)
    rows = audience["matched"]
    columns = B2C_COLUMNS if source == "b2c" else B2B_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audience"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D92681", end_color="D92681", fill_type="solid")
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, rec in enumerate(rows, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            value = rec.get(key)
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")

    for col_idx, (label, key) in enumerate(columns, start=1):
        width = max(
            [len(label)] + [len(str(r.get(key) or "")) for r in rows[:500]]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 4, 12), 45)
    ws.freeze_panes = "A2"

    # Summary sheet so the numbers in the file match the preview.
    summary = wb.create_sheet("Summary")
    for i, (label, value) in enumerate(
        [
            ("Segment", title),
            ("Source", source.upper()),
            ("Total matched", len(rows)),
            ("Reachable (has phone)", len(audience["recipients"])),
            ("Missing phone", audience["without_phone"]),
            ("Generated at", _now().strftime("%Y-%m-%d %H:%M")),
        ],
        start=1,
    ):
        summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        summary.cell(row=i, column=2, value=value)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 32

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _xlsx_response(buffer: BytesIO, title: str) -> StreamingResponse:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", title or "audience").strip("_") or "audience"
    filename = f"{safe}_{_now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export")
def export_preview(payload: dict = Body(...)):
    """Download the full audience for an unsaved rule as XLSX."""
    source = payload.get("source", "b2b")
    rule = payload.get("rule", {})
    title = (payload.get("name") or "audience").strip() or "audience"
    return _xlsx_response(_build_audience_workbook(source, rule, title), title)


@router.get("/{segment_id}/export")
def export_saved(segment_id: str):
    """Download the full audience for a saved segment as XLSX."""
    seg = segments_col.find_one({"_id": _oid(segment_id)})
    if not seg:
        raise HTTPException(status_code=404, detail="Segment not found")
    title = seg.get("name") or "audience"
    buffer = _build_audience_workbook(seg.get("source", "b2b"), seg.get("rule", {}), title)
    return _xlsx_response(buffer, title)


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

@router.get("")
def list_segments(
    search: str = Query(None),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0),
):
    query: dict = {}
    if search:
        query["name"] = {"$regex": search.strip(), "$options": "i"}
    raw = list(segments_col.find(query).sort("updated_at", -1).skip(skip).limit(limit))
    total = segments_col.count_documents(query)
    return {"data": serialize_mongo_document(raw), "total": total, "limit": limit, "skip": skip}


@router.get("/{segment_id}")
def get_segment(segment_id: str):
    seg = segments_col.find_one({"_id": _oid(segment_id)})
    if not seg:
        raise HTTPException(status_code=404, detail="Segment not found")
    return {"data": serialize_mongo_document(seg)}


@router.post("")
def create_segment(payload: dict = Body(...)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Segment name is required")
    if segments_col.find_one({"name": name}):
        raise HTTPException(status_code=409, detail=f"A segment named '{name}' already exists")
    doc = {
        "name": name,
        "description": payload.get("description", ""),
        "source": (payload.get("source") or "b2b").lower(),
        "rule": payload.get("rule", {}),
        "created_at": _now(),
        "updated_at": _now(),
    }
    result = segments_col.insert_one(doc)
    doc["_id"] = result.inserted_id
    return {"data": serialize_mongo_document(doc)}


@router.put("/{segment_id}")
def update_segment(segment_id: str, payload: dict = Body(...)):
    seg = segments_col.find_one({"_id": _oid(segment_id)})
    if not seg:
        raise HTTPException(status_code=404, detail="Segment not found")
    set_fields = {"updated_at": _now()}
    for key in ("name", "description", "source", "rule"):
        if key in payload:
            set_fields[key] = payload[key]
    if "source" in set_fields:
        set_fields["source"] = str(set_fields["source"]).lower()
    segments_col.update_one({"_id": seg["_id"]}, {"$set": set_fields})
    return {"data": serialize_mongo_document(segments_col.find_one({"_id": seg["_id"]}))}


@router.delete("/{segment_id}")
def delete_segment(segment_id: str):
    res = segments_col.delete_one({"_id": _oid(segment_id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Segment not found")
    return {"status": "deleted"}


def _oid(segment_id: str) -> ObjectId:
    try:
        return ObjectId(segment_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid segment id")
