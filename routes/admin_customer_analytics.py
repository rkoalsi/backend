from fastapi import APIRouter, Query, HTTPException, Path
from fastapi.responses import JSONResponse, StreamingResponse
from ..config.root import get_database, serialize_mongo_document
from bson.objectid import ObjectId
from dotenv import load_dotenv
import boto3, logging, os, openpyxl, re
from datetime import datetime
from difflib import SequenceMatcher
from typing import Optional
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

load_dotenv()
router = APIRouter()
org_id = os.getenv("ORG_ID")
db = get_database()
products_collection = db["products"]
customers_collection = db["customers"]
orders_collection = db["orders"]
users_collection = db["users"]

logger = logging.getLogger(__name__)
logger.propagate = False

AWS_ACCESS_KEY_ID = os.getenv("S3_ACCESS_KEY")
AWS_SECRET_ACCESS_KEY = os.getenv("S3_SECRET_KEY")
AWS_S3_BUCKET_NAME = os.getenv("S3_BUCKET_NAME")
AWS_S3_REGION = os.getenv("S3_REGION", "ap-south-1")  # Default to ap-south-1
AWS_S3_URL = os.getenv("S3_URL")

s3_client = boto3.client(
    "s3",
    region_name=AWS_S3_REGION,
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
)


def build_customer_analytics_pipeline(
    match_stage,
    customer_status_match_stage,
    sales_person_logic,
    due_status,
    last_billed,
    current_date_info,
    include_all_invoices=False,
):
    """
    Builds the shared MongoDB aggregation pipeline for customer analytics.
    Both the GET endpoint and the /report endpoint call this.

    current_date_info is a dict with keys:
        current_year, current_month, current_fy_start_year,
        last_fy_start_year, previous_fy_start_year,
        completed_months_in_current_fy, current_date
    """
    current_year = current_date_info["current_year"]
    current_month = current_date_info["current_month"]
    current_fy_start_year = current_date_info["current_fy_start_year"]
    last_fy_start_year = current_date_info["last_fy_start_year"]
    previous_fy_start_year = current_date_info["previous_fy_start_year"]
    completed_months_in_current_fy = current_date_info["completed_months_in_current_fy"]
    current_date = current_date_info["current_date"]

    pipeline = [
        # Stage 1: Filter invoices
        {"$match": match_stage},
        # Stage 1b: Early $project — only carry forward fields needed downstream
        {
            "$project": {
                "date": 1,
                "customer_name": 1,
                "customer_id": 1,
                "invoice_number": 1,
                "invoice_id": 1,
                "due_date": 1,
                "status": 1,
                "total": 1,
                "balance": 1,
                "salesperson_name": 1,
                "cf_sales_person": 1,
                "shipping_address": 1,
                "billing_address": 1,
            }
        },
        # Stage 2: Parse date ONCE, then compute all date-derived fields
        {
            "$addFields": {
                "parsedDate": {"$dateFromString": {"dateString": "$date"}},
                "parsedYear": {"$year": {"$dateFromString": {"dateString": "$date"}}},
                "parsedMonth": {"$month": {"$dateFromString": {"dateString": "$date"}}},
            }
        },
        # Stage 3: Compute all boolean/derived fields using the pre-parsed date fields
        {
            "$addFields": {
                "yearMonth": {
                    "$concat": [
                        {"$toString": "$parsedYear"},
                        "-",
                        {
                            "$cond": {
                                "if": {"$lt": ["$parsedMonth", 10]},
                                "then": {
                                    "$concat": [
                                        "0",
                                        {"$toString": "$parsedMonth"},
                                    ]
                                },
                                "else": {"$toString": "$parsedMonth"},
                            }
                        },
                    ]
                },
                # Payment categorization
                "isDuePayment": {
                    "$not": {"$in": ["$status", ["void", "draft", "sent", "paid"]]}
                },
                "isNotDuePayment": {
                    "$not": {
                        "$in": ["$status", ["void", "overdue", "partially_paid"]]
                    }
                },
                # Current month check
                "isCurrentMonth": {
                    "$and": [
                        {"$eq": ["$parsedYear", current_year]},
                        {"$eq": ["$parsedMonth", current_month]},
                    ]
                },
                # Completed months in current FY
                "isCompletedMonth": {
                    "$or": (
                        [
                            {
                                "$and": [
                                    {"$eq": ["$parsedYear", current_fy_start_year]},
                                    {"$gte": ["$parsedMonth", 4]},
                                    {"$lt": ["$parsedMonth", current_month]},
                                ]
                            },
                            {
                                "$and": [
                                    {"$eq": ["$parsedYear", current_fy_start_year + 1]},
                                    {"$gte": ["$parsedMonth", 1]},
                                    {"$lt": ["$parsedMonth", current_month]},
                                    {"$lte": ["$parsedMonth", 3]},
                                ]
                            },
                        ]
                        if current_month > 4
                        else [
                            {
                                "$and": [
                                    {"$eq": ["$parsedYear", current_fy_start_year]},
                                    {"$gte": ["$parsedMonth", 4]},
                                    {"$lt": ["$parsedMonth", current_month]},
                                ]
                            }
                        ]
                    )
                },
                # Current financial year
                "isCurrentFY": {
                    "$or": [
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", current_fy_start_year]},
                                {"$gte": ["$parsedMonth", 4]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", current_fy_start_year + 1]},
                                {"$lte": ["$parsedMonth", 3]},
                            ]
                        },
                    ]
                },
                # Last financial year
                "isLastFY": {
                    "$or": [
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", last_fy_start_year]},
                                {"$gte": ["$parsedMonth", 4]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", last_fy_start_year + 1]},
                                {"$lte": ["$parsedMonth", 3]},
                            ]
                        },
                    ]
                },
                # Previous financial year
                "isPreviousFY": {
                    "$or": [
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", previous_fy_start_year]},
                                {"$gte": ["$parsedMonth", 4]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", previous_fy_start_year + 1]},
                                {"$lte": ["$parsedMonth", 3]},
                            ]
                        },
                    ]
                },
                # Normalize city
                "normalizedCity": {
                    "$switch": {
                        "branches": [
                            {
                                "case": {
                                    "$and": [
                                        {"$ne": ["$shipping_address.city", None]},
                                        {"$ne": ["$shipping_address.city", ""]},
                                        {
                                            "$regexMatch": {
                                                "input": {
                                                    "$ifNull": [
                                                        "$shipping_address.city",
                                                        "",
                                                    ]
                                                },
                                                "regex": "^(bangalore|bengaluru)$",
                                                "options": "i",
                                            }
                                        },
                                    ]
                                },
                                "then": "bengaluru",
                            },
                            {
                                "case": {
                                    "$and": [
                                        {"$ne": ["$shipping_address.city", None]},
                                        {"$ne": ["$shipping_address.city", ""]},
                                        {
                                            "$regexMatch": {
                                                "input": {
                                                    "$ifNull": [
                                                        "$shipping_address.city",
                                                        "",
                                                    ]
                                                },
                                                "regex": "^(mumbai|bombay)$",
                                                "options": "i",
                                            }
                                        },
                                    ]
                                },
                                "then": "mumbai",
                            },
                            {
                                "case": {
                                    "$and": [
                                        {"$ne": ["$shipping_address.city", None]},
                                        {"$ne": ["$shipping_address.city", ""]},
                                        {
                                            "$regexMatch": {
                                                "input": {
                                                    "$ifNull": [
                                                        "$shipping_address.city",
                                                        "",
                                                    ]
                                                },
                                                "regex": "^(delhi|new delhi)$",
                                                "options": "i",
                                            }
                                        },
                                    ]
                                },
                                "then": "delhi",
                            },
                        ],
                        "default": {
                            "$toLower": {
                                "$ifNull": [
                                    "$shipping_address.city",
                                    "unknown_city",
                                ]
                            }
                        },
                    }
                },
                # Normalized full street
                "normalizedFullStreet": {
                    "$trim": {
                        "input": {
                            "$replaceAll": {
                                "input": {
                                    "$replaceAll": {
                                        "input": {
                                            "$replaceAll": {
                                                "input": {
                                                    "$toLower": {
                                                        "$concat": [
                                                            {"$ifNull": ["$shipping_address.street", ""]},
                                                            " ",
                                                            {"$ifNull": ["$shipping_address.street2", ""]}
                                                        ]
                                                    }
                                                },
                                                "find": ",",
                                                "replacement": " "
                                            }
                                        },
                                        "find": ".",
                                        "replacement": " "
                                    }
                                },
                                "find": "  ",
                                "replacement": " "
                            }
                        }
                    }
                },
                # Shipping address checks
                "hasShippingAddress": {"$ne": ["$shipping_address", None]},
                "shippingAddressComplete": {
                    "$and": [
                        {"$ne": ["$shipping_address", None]},
                        {"$ne": ["$shipping_address.city", None]},
                        {"$ne": ["$shipping_address.city", ""]},
                        {"$ne": ["$shipping_address.state", None]},
                        {"$ne": ["$shipping_address.country", None]},
                    ]
                },
                # Billing period validations using parsedYear/parsedMonth
                "billedLastMonth": {
                    "$and": [
                        {
                            "$eq": [
                                "$parsedYear",
                                (
                                    current_year
                                    if current_month > 1
                                    else current_year - 1
                                ),
                            ]
                        },
                        {
                            "$eq": [
                                "$parsedMonth",
                                current_month - 1 if current_month > 1 else 12,
                            ]
                        },
                    ]
                },
                "billedLast45Days": {
                    "$gte": [
                        "$parsedDate",
                        {
                            "$dateFromString": {
                                "dateString": f"{current_date.year}-{current_date.month:02d}-{max(1, current_date.day - 45):02d}"
                            }
                        },
                    ]
                },
                "billedLast2Months": {
                    "$or": [
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", current_year]},
                                {"$eq": ["$parsedMonth", current_month]},
                            ]
                        },
                        {
                            "$and": [
                                {
                                    "$eq": [
                                        "$parsedYear",
                                        (
                                            current_year
                                            if current_month > 1
                                            else current_year - 1
                                        ),
                                    ]
                                },
                                {
                                    "$eq": [
                                        "$parsedMonth",
                                        (
                                            current_month - 1
                                            if current_month > 1
                                            else 12
                                        ),
                                    ]
                                },
                            ]
                        },
                    ]
                },
                "billedLast3Months": {
                    "$or": [
                        {
                            "$and": [
                                {"$eq": ["$parsedYear", current_year]},
                                {"$eq": ["$parsedMonth", current_month]},
                            ]
                        },
                        {
                            "$and": [
                                {
                                    "$eq": [
                                        "$parsedYear",
                                        (
                                            current_year
                                            if current_month > 1
                                            else current_year - 1
                                        ),
                                    ]
                                },
                                {
                                    "$eq": [
                                        "$parsedMonth",
                                        (
                                            current_month - 1
                                            if current_month > 1
                                            else 12
                                        ),
                                    ]
                                },
                            ]
                        },
                        {
                            "$and": [
                                {
                                    "$eq": [
                                        "$parsedYear",
                                        (
                                            current_year
                                            if current_month > 2
                                            else current_year - 1
                                        ),
                                    ]
                                },
                                {
                                    "$eq": [
                                        "$parsedMonth",
                                        (
                                            current_month - 2
                                            if current_month > 2
                                            else (12 + current_month - 2)
                                        ),
                                    ]
                                },
                            ]
                        },
                    ]
                },
            }
        },
    ]

    # Stage 4: Group by customer name AND address
    group_stage = {
        "$group": {
            "_id": {
                "customerNameLower": {"$toLower": {"$trim": {"input": "$customer_name"}}},
                "city": "$normalizedCity",
                "state": {"$toLower": {"$trim": {"input": {"$ifNull": ["$shipping_address.state", ""]}}}},
                "zip": {"$replaceAll": {"input": {"$ifNull": ["$shipping_address.zip", ""]}, "find": " ", "replacement": ""}},
                "fullStreet": "$normalizedFullStreet",
            },
            "customerId": {"$first": "$customer_id"},
            "customerName": {"$first": "$customer_name"},
            "shippingAddress": {"$first": "$shipping_address"},
            "billingAddress": {"$first": "$billing_address"},
            "salesPerson": {"$first": sales_person_logic},
            # Collect payment information
            "duePayments": {
                "$push": {
                    "$cond": [
                        "$isDuePayment",
                        {
                            "_id": "$_id",
                            "date": "$date",
                            "due_date": "$due_date",
                            "invoice_number": "$invoice_number",
                            "status": "$status",
                            "invoice_id": "$invoice_id",
                            "total": "$total",
                            "balance": "$balance",
                        },
                        "$$REMOVE",
                    ]
                }
            },
            "notDuePayments": {
                "$push": {
                    "$cond": [
                        "$isNotDuePayment",
                        {
                            "_id": "$_id",
                            "date": "$date",
                            "due_date": "$due_date",
                            "invoice_number": "$invoice_number",
                            "status": "$status",
                            "invoice_id": "$invoice_id",
                            "balance": "$balance",
                            "total": "$total",
                        },
                        "$$REMOVE",
                    ]
                }
            },
            "totalSalesCurrentMonth": {
                "$sum": {
                    "$cond": [{"$eq": ["$isCurrentMonth", True]}, "$total", 0]
                }
            },
            "lastBillDate": {"$max": "$parsedDate"},
            "currentFYOrders": {
                "$sum": {"$cond": [{"$eq": ["$isCurrentFY", True]}, 1, 0]}
            },
            "currentFYMonths": {
                "$addToSet": {
                    "$cond": [
                        {"$eq": ["$isCurrentFY", True]},
                        "$yearMonth",
                        None,
                    ]
                }
            },
            "completedMonthOrders": {
                "$sum": {"$cond": [{"$eq": ["$isCompletedMonth", True]}, 1, 0]}
            },
            "completedMonths": {
                "$addToSet": {
                    "$cond": [
                        {"$eq": ["$isCompletedMonth", True]},
                        "$yearMonth",
                        None,
                    ]
                }
            },
            "billingTillDateCurrentYear": {
                "$sum": {
                    "$cond": [{"$eq": ["$isCurrentFY", True]}, "$total", 0]
                }
            },
            "totalSalesLastFY": {
                "$sum": {"$cond": [{"$eq": ["$isLastFY", True]}, "$total", 0]}
            },
            "totalSalesPreviousFY": {
                "$sum": {
                    "$cond": [{"$eq": ["$isPreviousFY", True]}, "$total", 0]
                }
            },
            "hasBilledLastMonth": {
                "$sum": {"$cond": [{"$eq": ["$billedLastMonth", True]}, 1, 0]}
            },
            "hasBilledLast45Days": {
                "$sum": {"$cond": [{"$eq": ["$billedLast45Days", True]}, 1, 0]}
            },
            "hasBilledLast2Months": {
                "$sum": {"$cond": [{"$eq": ["$billedLast2Months", True]}, 1, 0]}
            },
            "hasBilledLast3Months": {
                "$sum": {"$cond": [{"$eq": ["$billedLast3Months", True]}, 1, 0]}
            },
            "totalInvoiceCount": {"$sum": 1},
        }
    }

    # Only include allInvoices and allInvoiceDates when requested (report endpoint)
    if include_all_invoices:
        group_stage["$group"]["allInvoices"] = {
            "$push": {
                "_id": "$_id",
                "invoice_number": "$invoice_number",
                "date": "$date",
                "due_date": "$due_date",
                "status": "$status",
                "total": "$total",
                "balance": "$balance",
                "customer_id": "$customer_id",
                "invoice_id": "$invoice_id",
                "yearMonth": "$yearMonth",
                "isCurrentMonth": "$isCurrentMonth",
                "isCurrentFY": "$isCurrentFY",
                "isLastFY": "$isLastFY",
                "isPreviousFY": "$isPreviousFY",
            }
        }
        group_stage["$group"]["allInvoiceDates"] = {"$push": "$parsedDate"}

    pipeline.append(group_stage)

    # Stage 5: Lookup customer details
    pipeline.append(
        {
            "$lookup": {
                "from": "customers",
                "localField": "customerId",
                "foreignField": "contact_id",
                "as": "customerDetails",
            }
        }
    )
    pipeline.append({"$match": customer_status_match_stage})

    # Stage 6: Calculate derived fields
    pipeline.append(
        {
            "$addFields": {
                "currentFYMonthsFiltered": {
                    "$filter": {
                        "input": "$currentFYMonths",
                        "cond": {"$ne": ["$$this", None]},
                    }
                },
                "completedMonthsFiltered": {
                    "$filter": {
                        "input": "$completedMonths",
                        "cond": {"$ne": ["$$this", None]},
                    }
                },
                "customerStatus": {
                    "$ifNull": [
                        {"$arrayElemAt": ["$customerDetails.status", 0]},
                        "unknown",
                    ]
                },
                "customerTier": {
                    "$ifNull": [
                        {"$arrayElemAt": ["$customerDetails.cf_tier", 0]},
                        "unknown",
                    ]
                },
                "salesPerson": {
                    "$ifNull": [
                        {"$arrayElemAt": ["$customerDetails.cf_sales_person", 0]},
                        "$salesPerson",
                    ]
                },
                "hasBilledLastMonth": {"$gt": ["$hasBilledLastMonth", 0]},
                "hasBilledLast45Days": {"$gt": ["$hasBilledLast45Days", 0]},
                "hasBilledLast2Months": {"$gt": ["$hasBilledLast2Months", 0]},
                "hasBilledLast3Months": {"$gt": ["$hasBilledLast3Months", 0]},
            }
        }
    )

    # Stage 7: Final metrics
    pipeline.append(
        {
            "$addFields": {
                "averageOrderFrequencyMonthly": {
                    "$cond": [
                        {"$gt": ["$currentFYOrders", 0]},
                        {
                            "$divide": [
                                "$currentFYOrders",
                                completed_months_in_current_fy,
                            ]
                        },
                        0,
                    ]
                },
                "shippingAddressFormatted": {
                    "$concat": [
                        {"$ifNull": ["$shippingAddress.street", ""]},
                        {
                            "$cond": [
                                {"$ne": ["$shippingAddress.street2", ""]},
                                {"$concat": [", ", "$shippingAddress.street2"]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$shippingAddress.city", ""]},
                                {"$concat": [", ", "$shippingAddress.city"]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$shippingAddress.state", ""]},
                                {"$concat": [", ", "$shippingAddress.state"]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$shippingAddress.zip", ""]},
                                {"$concat": [" - ", "$shippingAddress.zip"]},
                                "",
                            ]
                        },
                    ]
                },
                "billingAddressFormatted": {
                    "$concat": [
                        {"$ifNull": ["$billingAddress.street", ""]},
                        {
                            "$cond": [
                                {"$ne": ["$billingAddress.street2", ""]},
                                {"$concat": [", ", {"$ifNull": ["$billingAddress.street2", ""]}]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$billingAddress.city", ""]},
                                {"$concat": [", ", {"$ifNull": ["$billingAddress.city", ""]}]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$billingAddress.state", ""]},
                                {"$concat": [", ", {"$ifNull": ["$billingAddress.state", ""]}]},
                                "",
                            ]
                        },
                        {
                            "$cond": [
                                {"$ne": ["$billingAddress.zip", ""]},
                                {"$concat": [" - ", {"$ifNull": ["$billingAddress.zip", ""]}]},
                                "",
                            ]
                        },
                    ]
                },
            }
        }
    )

    # Stage 8: Due status filtering
    pipeline.append(
        {
            "$addFields": {
                "filteredDuePayments": {
                    "$cond": [
                        {
                            "$or": [
                                {"$eq": [due_status, "all"]},
                                {"$eq": [due_status, "due"]},
                            ]
                        },
                        "$duePayments",
                        [],
                    ]
                },
                "filteredNotDuePayments": {
                    "$cond": [
                        {
                            "$or": [
                                {"$eq": [due_status, "all"]},
                                {"$eq": [due_status, "not_due"]},
                            ]
                        },
                        "$notDuePayments",
                        [],
                    ]
                },
            }
        }
    )
    pipeline.append(
        {
            "$match": {
                "$expr": {
                    "$or": [
                        {"$eq": [due_status, "all"]},
                        {
                            "$and": [
                                {"$eq": [due_status, "due"]},
                                {"$gt": [{"$size": "$filteredDuePayments"}, 0]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [due_status, "not_due"]},
                                {"$gt": [{"$size": "$filteredNotDuePayments"}, 0]},
                            ]
                        },
                    ]
                }
            }
        }
    )

    # Stage 9: Last billed filtering
    pipeline.append(
        {
            "$match": {
                "$expr": {
                    "$or": [
                        {"$eq": [last_billed, "all"]},
                        {
                            "$and": [
                                {"$eq": [last_billed, "last_month"]},
                                {"$eq": ["$hasBilledLastMonth", True]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "last_45_days"]},
                                {"$eq": ["$hasBilledLast45Days", True]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "last_2_months"]},
                                {"$eq": ["$hasBilledLast2Months", True]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "last_3_months"]},
                                {"$eq": ["$hasBilledLast3Months", True]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "not_last_month"]},
                                {"$eq": ["$hasBilledLastMonth", False]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "not_last_45_days"]},
                                {"$eq": ["$hasBilledLast45Days", False]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "not_last_2_months"]},
                                {"$eq": ["$hasBilledLast2Months", False]},
                            ]
                        },
                        {
                            "$and": [
                                {"$eq": [last_billed, "not_last_3_months"]},
                                {"$eq": ["$hasBilledLast3Months", False]},
                            ]
                        },
                    ]
                }
            }
        }
    )

    # Stage 10: Final projection
    project_stage = {
        "$project": {
            "_id": 0,
            "customerId": 1,
            "customerName": 1,
            "shippingAddress": "$shippingAddressFormatted",
            "billingAddress": "$billingAddressFormatted",
            "status": "$customerStatus",
            "tier": "$customerTier",
            "totalSalesCurrentMonth": {
                "$round": ["$totalSalesCurrentMonth", 2]
            },
            "lastBillDate": {
                "$dateToString": {"format": "%Y-%m-%d", "date": "$lastBillDate"}
            },
            "averageOrderFrequencyMonthly": {
                "$round": ["$averageOrderFrequencyMonthly", 2]
            },
            "billingTillDateCurrentYear": {
                "$round": ["$billingTillDateCurrentYear", 2]
            },
            "totalSalesLastFY": {"$round": ["$totalSalesLastFY", 2]},
            "totalSalesPreviousFY": {"$round": ["$totalSalesPreviousFY", 2]},
            "salesPerson": 1,
            "hasBilledLastMonth": 1,
            "hasBilledLast45Days": 1,
            "hasBilledLast2Months": 1,
            "hasBilledLast3Months": 1,
            "duePayments": "$filteredDuePayments",
            "notDuePayments": "$filteredNotDuePayments",
            "totalInvoiceCount": 1,
            "currentFYInvoiceCount": "$currentFYOrders",
        }
    }

    if include_all_invoices:
        project_stage["$project"]["allInvoices"] = 1

    # Expose address key components so callers can build composite brand-lookup keys
    project_stage["$project"]["addressCity"] = "$_id.city"
    project_stage["$project"]["addressState"] = "$_id.state"
    project_stage["$project"]["addressZip"] = "$_id.zip"
    project_stage["$project"]["addressStreet"] = "$_id.fullStreet"

    pipeline.append(project_stage)

    # Stage 11: Sort by customer name
    pipeline.append({"$sort": {"customerName": 1}})

    return pipeline


def _get_current_date_info():
    """Compute all date-related info needed by the pipeline."""
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month

    if current_month >= 4:
        current_fy_start_year = current_year
    else:
        current_fy_start_year = current_year - 1

    last_fy_start_year = current_fy_start_year - 1
    previous_fy_start_year = current_fy_start_year - 2

    if current_month >= 4:
        completed_months_in_current_fy = current_month - 4
    else:
        completed_months_in_current_fy = (12 - 4) + current_month

    completed_months_in_current_fy = max(1, completed_months_in_current_fy)

    return {
        "current_date": current_date,
        "current_year": current_year,
        "current_month": current_month,
        "current_fy_start_year": current_fy_start_year,
        "last_fy_start_year": last_fy_start_year,
        "previous_fy_start_year": previous_fy_start_year,
        "completed_months_in_current_fy": completed_months_in_current_fy,
    }


def _build_match_and_filters(status, tier, sort_by):
    """Build common match_stage, customer_status_match_stage, sort_stage, and sales_person_logic."""
    match_stage = {
        "date": {"$gte": "2023-04-01"},
        "status": {"$nin": ["void", "draft"]},
        "$and": [
            {
                "customer_name": {
                    "$not": {
                        "$regex": r"\b(EC|NA|PUPEV|RS|MKT|SPUR|SSAM|OSAMP)\b",
                        "$options": "i",
                    }
                }
            },
            {
                "customer_name": {
                    "$not": {
                        "$regex": r"(amzb2b|amz2b2|Blinkit|Flipkart)",
                        "$options": "i",
                    }
                }
            },
        ],
    }
    customer_status_match_stage = {}
    sort_stage = {"$sort": {"totalSalesCurrentMonth": 1}}

    if status == "all":
        pass
    elif status:
        customer_status_match_stage["customerDetails.status"] = status

    if sort_by:
        sort_stage["$sort"] = [{"totalSalesCurrentMonth": -1}]

    sales_person_logic = {
        "$cond": [
            {"$ne": ["$salesperson_name", None]},
            "$salesperson_name",
            "$cf_sales_person",
        ]
    }

    if tier and len(tier) == 1:
        customer_status_match_stage["customerDetails.cf_tier"] = {
            "$regex": f"^{tier}$",
            "$options": "i",
        }

    return match_stage, customer_status_match_stage, sort_stage, sales_person_logic


@router.get("")
def get_admin_customer_analytics(
    status: Optional[str] = Query(None, description="Filter by invoice status"),
    tier: Optional[str] = Query(None, description="Filter by tiers (A,B,C)"),
    due_status: Optional[str] = Query(
        "all", description="Filter by Payments Due (all, due, not_due)"
    ),
    last_billed: Optional[str] = Query(
        "all",
        description="Filter by last billing activity (all, last_month, last_45_days, last_2_months, last_3_months, not_last_month, not_last_45_days, not_last_2_months, not_last_3_months)",
    ),
    sort_by: Optional[bool] = Query(True, description="Low to High or High to Low"),
):
    try:
        match_stage, customer_status_match_stage, sort_stage, sales_person_logic = (
            _build_match_and_filters(status, tier, sort_by)
        )
        current_date_info = _get_current_date_info()

        pipeline = build_customer_analytics_pipeline(
            match_stage=match_stage,
            customer_status_match_stage=customer_status_match_stage,
            sales_person_logic=sales_person_logic,
            due_status=due_status,
            last_billed=last_billed,
            current_date_info=current_date_info,
            include_all_invoices=False,
        )

        customers = list(db.invoices.aggregate(pipeline, allowDiskUse=True))
        return serialize_mongo_document(customers)
    except Exception as e:
        logger.error(f"Error in get_customer_analytics: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/report")
def download_customer_analytics_report(
    status: Optional[str] = Query(None, description="Filter by invoice status"),
    tier: Optional[str] = Query(None, description="Filter by tiers (A,B,C)"),
    due_status: Optional[str] = Query(
        "all", description="Filter by Payments Due (all, due, not_due)"
    ),
    last_billed: Optional[str] = Query(
        "all",
        description="Filter by last billing activity (all, last_month, last_45_days, last_2_months, last_3_months, not_last_month, not_last_45_days, not_last_2_months, not_last_3_months)",
    ),
    sort_by: Optional[bool] = Query(True, description="Low to High or High to Low"),
    include_brand_breakdown: Optional[bool] = Query(False, description="Include brand breakdown sheet"),
    brands: Optional[str] = Query(None, description="Comma-separated list of brands to include in brand breakdown (default: all)"),
):
    try:
        match_stage, customer_status_match_stage, sort_stage, sales_person_logic = (
            _build_match_and_filters(status, tier, sort_by)
        )
        current_date_info = _get_current_date_info()

        pipeline = build_customer_analytics_pipeline(
            match_stage=match_stage,
            customer_status_match_stage=customer_status_match_stage,
            sales_person_logic=sales_person_logic,
            due_status=due_status,
            last_billed=last_billed,
            current_date_info=current_date_info,
            include_all_invoices=True,
        )

        # Execute the aggregation
        customers = list(db.invoices.aggregate(pipeline, allowDiskUse=True))

        # Create Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Customer Analytics Report"

        # Define headers
        headers = [
            "Customer ID",
            "Customer Name",
            "Shipping Address",
            "Status",
            "Tier",
            "Sales Person",
            "Total Sales Current Month",
            "Last Bill Date",
            "Average Order Frequency (Monthly)",
            "Billing Till Date Current Year",
            "Total Sales Last FY",
            "Total Sales Previous FY",
            "Billed Last Month",
            "Billed Last 45 Days",
            "Billed Last 2 Months",
            "Billed Last 3 Months",
            "Due Payments Count",
            "Not Due Payments Count",
        ]

        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row, customer in enumerate(customers, 2):
            worksheet.cell(row=row, column=1, value=customer.get("customerId", ""))
            worksheet.cell(row=row, column=2, value=customer.get("customerName", ""))
            worksheet.cell(
                row=row, column=3, value=customer.get("shippingAddress", "")
            )
            worksheet.cell(row=row, column=4, value=customer.get("status", ""))
            worksheet.cell(row=row, column=5, value=customer.get("tier", ""))
            sales_person = customer.get("salesPerson", "")
            if isinstance(sales_person, list):
                sales_person = ", ".join(sales_person) if sales_person else ""
            worksheet.cell(row=row, column=6, value=sales_person)
            worksheet.cell(
                row=row, column=7, value=customer.get("totalSalesCurrentMonth", 0)
            )
            worksheet.cell(row=row, column=8, value=customer.get("lastBillDate", ""))
            worksheet.cell(
                row=row, column=9, value=customer.get("averageOrderFrequencyMonthly", 0)
            )
            worksheet.cell(
                row=row, column=10, value=customer.get("billingTillDateCurrentYear", 0)
            )
            worksheet.cell(
                row=row, column=11, value=customer.get("totalSalesLastFY", 0)
            )
            worksheet.cell(
                row=row, column=12, value=customer.get("totalSalesPreviousFY", 0)
            )
            worksheet.cell(
                row=row,
                column=13,
                value="Yes" if customer.get("hasBilledLastMonth", False) else "No",
            )
            worksheet.cell(
                row=row,
                column=14,
                value="Yes" if customer.get("hasBilledLast45Days", False) else "No",
            )
            worksheet.cell(
                row=row,
                column=15,
                value="Yes" if customer.get("hasBilledLast2Months", False) else "No",
            )
            worksheet.cell(
                row=row,
                column=16,
                value="Yes" if customer.get("hasBilledLast3Months", False) else "No",
            )
            worksheet.cell(
                row=row, column=17, value=len(customer.get("duePayments", []))
            )
            worksheet.cell(
                row=row, column=18, value=len(customer.get("notDuePayments", []))
            )

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add Customer Address Summary sheet — fuzzy-group same customer/similar address rows
        def _normalize_addr(addr):
            if not addr:
                return ""
            addr = addr.lower()
            addr = re.sub(r"[^\w\s]", " ", addr)
            addr = re.sub(r"\s+", " ", addr).strip()
            return addr

        def _addr_similarity(a, b):
            na, nb = _normalize_addr(a), _normalize_addr(b)
            if not na and not nb:
                return 1.0
            if not na or not nb:
                return 0.0
            return SequenceMatcher(None, na, nb).ratio()

        from collections import defaultdict

        name_groups = defaultdict(list)
        for c in customers:
            key = c.get("customerName", "").lower().strip()
            name_groups[key].append(c)

        def _addr_key(c):
            """Return the same (city, state, zip, street) tuple the brand pipeline uses."""
            return (
                c.get("addressCity", ""),
                c.get("addressState", ""),
                c.get("addressZip", ""),
                c.get("addressStreet", ""),
            )

        merged_rows = []
        for _name_key, group in name_groups.items():
            # Greedy address clustering: assign each item to first cluster it's similar to
            clusters = []
            for c in group:
                addr = c.get("shippingAddress", "")
                placed = False
                for cluster in clusters:
                    rep_addr = cluster[0].get("shippingAddress", "")
                    if _addr_similarity(addr, rep_addr) >= 0.75:
                        cluster.append(c)
                        placed = True
                        break
                if not placed:
                    clusters.append([c])

            for cluster in clusters:
                if len(cluster) == 1:
                    c = cluster[0]
                    sp = c.get("salesPerson", "")
                    if isinstance(sp, list):
                        sp = ", ".join(sp) if sp else ""
                    merged_rows.append({
                        "customerName": c.get("customerName", ""),
                        "customerIds": [c.get("customerId", "")] if c.get("customerId", "") else [],
                        # (cid, addr_key) pairs used for per-location brand lookup
                        "customerAddressKeys": [(c.get("customerId", ""), _addr_key(c))] if c.get("customerId", "") else [],
                        "allShippingAddresses": [c.get("shippingAddress", "")],
                        "billingAddress": c.get("billingAddress", ""),
                        "status": c.get("status", ""),
                        "tier": c.get("tier", ""),
                        "salesPerson": sp,
                        "totalSalesCurrentMonth": c.get("totalSalesCurrentMonth", 0),
                        "lastBillDate": c.get("lastBillDate", ""),
                        "averageOrderFrequencyMonthly": c.get("averageOrderFrequencyMonthly", 0),
                        "billingTillDateCurrentYear": c.get("billingTillDateCurrentYear", 0),
                        "totalSalesLastFY": c.get("totalSalesLastFY", 0),
                        "totalSalesPreviousFY": c.get("totalSalesPreviousFY", 0),
                        "totalInvoiceCount": c.get("totalInvoiceCount", 0),
                        "addressCount": 1,
                    })
                else:
                    sp = cluster[0].get("salesPerson", "")
                    if isinstance(sp, list):
                        sp = ", ".join(sp) if sp else ""
                    merged_rows.append({
                        "customerName": cluster[0].get("customerName", ""),
                        "customerIds": [c.get("customerId", "") for c in cluster if c.get("customerId", "")],
                        # (cid, addr_key) pairs used for per-location brand lookup
                        "customerAddressKeys": [
                            (c.get("customerId", ""), _addr_key(c))
                            for c in cluster if c.get("customerId", "")
                        ],
                        "allShippingAddresses": [c.get("shippingAddress", "") for c in cluster if c.get("shippingAddress", "")],
                        "billingAddress": cluster[0].get("billingAddress", ""),
                        "status": cluster[0].get("status", ""),
                        "tier": cluster[0].get("tier", ""),
                        "salesPerson": sp,
                        "totalSalesCurrentMonth": sum(c.get("totalSalesCurrentMonth", 0) for c in cluster),
                        "lastBillDate": max((c.get("lastBillDate", "") for c in cluster), default=""),
                        "averageOrderFrequencyMonthly": round(sum(c.get("averageOrderFrequencyMonthly", 0) for c in cluster), 2),
                        "billingTillDateCurrentYear": sum(c.get("billingTillDateCurrentYear", 0) for c in cluster),
                        "totalSalesLastFY": sum(c.get("totalSalesLastFY", 0) for c in cluster),
                        "totalSalesPreviousFY": sum(c.get("totalSalesPreviousFY", 0) for c in cluster),
                        "totalInvoiceCount": sum(c.get("totalInvoiceCount", 0) for c in cluster),
                        "addressCount": len(cluster),
                    })

        merged_rows.sort(key=lambda x: x["customerName"].lower())

        # Compute brand breakdown data if requested (before building the combined sheet)
        brand_totals = None
        all_brands = []
        current_fy_label = last_fy_label = previous_fy_label = ""
        if include_brand_breakdown:
            current_fy_start_year = current_date_info["current_fy_start_year"]
            last_fy_start_year = current_date_info["last_fy_start_year"]
            previous_fy_start_year = current_date_info["previous_fy_start_year"]

            current_fy_label = f"FY {current_fy_start_year}-{str(current_fy_start_year + 1)[2:]}"
            last_fy_label = f"FY {last_fy_start_year}-{str(last_fy_start_year + 1)[2:]}"
            previous_fy_label = f"FY {previous_fy_start_year}-{str(previous_fy_start_year + 1)[2:]}"

            # Pre-fetch product item_id -> brand map
            brand_map = {}
            for prod in products_collection.find({}, {"item_id": 1, "brand": 1, "_id": 0}):
                brand_map[prod.get("item_id", "")] = prod.get("brand", "Unknown") or "Unknown"

            customer_ids = [c.get("customerId", "") for c in customers if c.get("customerId", "")]

            brand_pipeline = [
                {
                    "$match": {
                        "customer_id": {"$in": customer_ids},
                        "date": {"$gte": "2023-04-01"},
                        "status": {"$nin": ["void", "draft"]},
                    }
                },
                {
                    "$addFields": {
                        "invoiceYear": {
                            "$year": {"$dateFromString": {"dateString": "$date"}}
                        },
                        "invoiceMonth": {
                            "$month": {"$dateFromString": {"dateString": "$date"}}
                        },
                        # Normalise address the same way the main pipeline does
                        "brandNormCity": {
                            "$switch": {
                                "branches": [
                                    {
                                        "case": {
                                            "$regexMatch": {
                                                "input": {"$toLower": {"$ifNull": ["$shipping_address.city", ""]}},
                                                "regex": "^(bangalore|bengaluru)$",
                                            }
                                        },
                                        "then": "bengaluru",
                                    },
                                    {
                                        "case": {
                                            "$regexMatch": {
                                                "input": {"$toLower": {"$ifNull": ["$shipping_address.city", ""]}},
                                                "regex": "^(mumbai|bombay)$",
                                            }
                                        },
                                        "then": "mumbai",
                                    },
                                    {
                                        "case": {
                                            "$regexMatch": {
                                                "input": {"$toLower": {"$ifNull": ["$shipping_address.city", ""]}},
                                                "regex": "^(delhi|new delhi)$",
                                            }
                                        },
                                        "then": "delhi",
                                    },
                                ],
                                "default": {"$toLower": {"$ifNull": ["$shipping_address.city", "unknown_city"]}},
                            }
                        },
                        "brandNormState": {
                            "$toLower": {"$trim": {"input": {"$ifNull": ["$shipping_address.state", ""]}}}
                        },
                        "brandNormZip": {
                            "$replaceAll": {
                                "input": {"$ifNull": ["$shipping_address.zip", ""]},
                                "find": " ",
                                "replacement": "",
                            }
                        },
                        "brandNormStreet": {
                            "$trim": {
                                "input": {
                                    "$replaceAll": {
                                        "input": {
                                            "$replaceAll": {
                                                "input": {
                                                    "$replaceAll": {
                                                        "input": {
                                                            "$toLower": {
                                                                "$concat": [
                                                                    {"$ifNull": ["$shipping_address.street", ""]},
                                                                    " ",
                                                                    {"$ifNull": ["$shipping_address.street2", ""]},
                                                                ]
                                                            }
                                                        },
                                                        "find": ",",
                                                        "replacement": " ",
                                                    }
                                                },
                                                "find": ".",
                                                "replacement": " ",
                                            }
                                        },
                                        "find": "  ",
                                        "replacement": " ",
                                    }
                                }
                            }
                        },
                    }
                },
                {
                    "$addFields": {
                        "isCurrentFY": {
                            "$or": [
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", current_fy_start_year]},
                                        {"$gte": ["$invoiceMonth", 4]},
                                    ]
                                },
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", current_fy_start_year + 1]},
                                        {"$lte": ["$invoiceMonth", 3]},
                                    ]
                                },
                            ]
                        },
                        "isLastFY": {
                            "$or": [
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", last_fy_start_year]},
                                        {"$gte": ["$invoiceMonth", 4]},
                                    ]
                                },
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", last_fy_start_year + 1]},
                                        {"$lte": ["$invoiceMonth", 3]},
                                    ]
                                },
                            ]
                        },
                        "isPreviousFY": {
                            "$or": [
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", previous_fy_start_year]},
                                        {"$gte": ["$invoiceMonth", 4]},
                                    ]
                                },
                                {
                                    "$and": [
                                        {"$eq": ["$invoiceYear", previous_fy_start_year + 1]},
                                        {"$lte": ["$invoiceMonth", 3]},
                                    ]
                                },
                            ]
                        },
                    }
                },
                {"$unwind": {"path": "$line_items", "preserveNullAndEmptyArrays": False}},
                {
                    "$group": {
                        "_id": {
                            "customer_id": "$customer_id",
                            # Include address components so totals are per-location
                            "city": "$brandNormCity",
                            "state": "$brandNormState",
                            "zip": "$brandNormZip",
                            "fullStreet": "$brandNormStreet",
                            "item_id": "$line_items.item_id",
                        },
                        "currentFY": {
                            "$sum": {
                                "$cond": [
                                    "$isCurrentFY",
                                    {"$ifNull": ["$line_items.item_total", 0]},
                                    0,
                                ]
                            }
                        },
                        "lastFY": {
                            "$sum": {
                                "$cond": [
                                    "$isLastFY",
                                    {"$ifNull": ["$line_items.item_total", 0]},
                                    0,
                                ]
                            }
                        },
                        "previousFY": {
                            "$sum": {
                                "$cond": [
                                    "$isPreviousFY",
                                    {"$ifNull": ["$line_items.item_total", 0]},
                                    0,
                                ]
                            }
                        },
                    }
                },
            ]

            # brand_totals[(cid, city, state, zip, street)][brand] = {currentFY, lastFY, previousFY}
            # Keying by (customer_id + address) ensures different locations of the same
            # customer (same customer_id) get separate totals.
            brand_totals = defaultdict(lambda: defaultdict(lambda: {"currentFY": 0, "lastFY": 0, "previousFY": 0}))
            for row in db.invoices.aggregate(brand_pipeline, allowDiskUse=True):
                cid = row["_id"]["customer_id"]
                addr_key = (
                    row["_id"].get("city", ""),
                    row["_id"].get("state", ""),
                    row["_id"].get("zip", ""),
                    row["_id"].get("fullStreet", ""),
                )
                item_id = row["_id"].get("item_id", "")
                brand = brand_map.get(item_id, "Unknown")
                composite_key = (cid, addr_key)
                brand_totals[composite_key][brand]["currentFY"] += row.get("currentFY", 0)
                brand_totals[composite_key][brand]["lastFY"] += row.get("lastFY", 0)
                brand_totals[composite_key][brand]["previousFY"] += row.get("previousFY", 0)

            all_brands_set = {brand for cid_brands in brand_totals.values() for brand in cid_brands}
            if brands:
                requested_brands = {b.strip() for b in brands.split(",") if b.strip()}
                all_brands = sorted(all_brands_set & requested_brands)
            else:
                all_brands = sorted(all_brands_set)

        # Build combined Address & Brand Breakdown sheet
        COLS_PER_BRAND = 5
        addr_base_headers = [
            "Customer ID",
            "Customer Name",
            "Primary Shipping Address",
            "All Shipping Addresses",
            "Billing Address",
            "Address Variants Merged",
            "Status",
            "Tier",
            "Sales Person",
            "Total Sales Current Month",
            "Last Bill Date",
            "Avg Order Frequency (Monthly)",
            "Billing Till Date Current Year",
            "Total Sales Last FY",
            "Total Sales Previous FY",
            "Total Invoice Count",
        ]
        brand_col_headers = []
        if include_brand_breakdown:
            for brand in all_brands:
                brand_col_headers.append(f"{brand} ({previous_fy_label})")
                brand_col_headers.append(f"{brand} ({last_fy_label})")
                brand_col_headers.append(f"{brand} YoY% ({previous_fy_label}→{last_fy_label})")
                brand_col_headers.append(f"{brand} ({current_fy_label})")
                brand_col_headers.append(f"{brand} YoY% ({last_fy_label}→{current_fy_label})")

        combined_headers = addr_base_headers + brand_col_headers

        addr_ws = workbook.create_sheet("Address & Brand Breakdown")
        for col, hdr in enumerate(combined_headers, 1):
            cell = addr_ws.cell(row=1, column=col, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        merged_highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        growth_positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        growth_negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        growth_positive_font = Font(color="276221")
        growth_negative_font = Font(color="9C0006")

        def _yoy_pct(current, previous):
            if previous == 0:
                return None
            return round((current - previous) / previous * 100, 1)

        def _write_growth(ws, row, col, pct):
            cell = ws.cell(row=row, column=col)
            if pct is None:
                cell.value = "N/A"
            else:
                cell.value = f"{'+' if pct >= 0 else ''}{pct}%"
                cell.fill = growth_positive_fill if pct >= 0 else growth_negative_fill
                cell.font = growth_positive_font if pct >= 0 else growth_negative_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, merged in enumerate(merged_rows, 2):
            all_addrs = merged.get("allShippingAddresses", [])
            primary_addr = all_addrs[0] if all_addrs else ""
            all_addrs_str = " | ".join(a for a in all_addrs if a)
            addr_ws.cell(row=row_idx, column=1, value=(merged.get("customerIds", [None])[0] or ""))
            addr_ws.cell(row=row_idx, column=2, value=merged.get("customerName", ""))
            addr_ws.cell(row=row_idx, column=3, value=primary_addr)
            addr_ws.cell(row=row_idx, column=4, value=all_addrs_str)
            addr_ws.cell(row=row_idx, column=5, value=merged.get("billingAddress", ""))
            addr_ws.cell(row=row_idx, column=6, value=merged.get("addressCount", 1))
            addr_ws.cell(row=row_idx, column=7, value=merged.get("status", ""))
            addr_ws.cell(row=row_idx, column=8, value=merged.get("tier", ""))
            addr_ws.cell(row=row_idx, column=9, value=merged.get("salesPerson", ""))
            addr_ws.cell(row=row_idx, column=10, value=round(merged.get("totalSalesCurrentMonth", 0), 2))
            addr_ws.cell(row=row_idx, column=11, value=merged.get("lastBillDate", ""))
            addr_ws.cell(row=row_idx, column=12, value=round(merged.get("averageOrderFrequencyMonthly", 0), 2))
            addr_ws.cell(row=row_idx, column=13, value=round(merged.get("billingTillDateCurrentYear", 0), 2))
            addr_ws.cell(row=row_idx, column=14, value=round(merged.get("totalSalesLastFY", 0), 2))
            addr_ws.cell(row=row_idx, column=15, value=round(merged.get("totalSalesPreviousFY", 0), 2))
            addr_ws.cell(row=row_idx, column=16, value=merged.get("totalInvoiceCount", 0))
            # Highlight rows where multiple address variants were merged
            if merged.get("addressCount", 1) > 1:
                for col in range(1, len(addr_base_headers) + 1):
                    addr_ws.cell(row=row_idx, column=col).fill = merged_highlight

            # Write brand breakdown columns for this merged row
            if include_brand_breakdown and brand_totals is not None:
                # Aggregate brand totals per (customer_id, address) key so that
                # multiple locations of the same customer_id are kept separate.
                row_brand_totals = defaultdict(lambda: {"currentFY": 0, "lastFY": 0, "previousFY": 0})
                for cid, addr_key in merged.get("customerAddressKeys", []):
                    composite_key = (cid, addr_key)
                    if composite_key in brand_totals:
                        for brand, totals in brand_totals[composite_key].items():
                            if brand in all_brands:
                                row_brand_totals[brand]["currentFY"] += totals["currentFY"]
                                row_brand_totals[brand]["lastFY"] += totals["lastFY"]
                                row_brand_totals[brand]["previousFY"] += totals["previousFY"]

                for brand_idx, brand in enumerate(all_brands):
                    base_col = len(addr_base_headers) + 1 + brand_idx * COLS_PER_BRAND
                    totals = row_brand_totals.get(brand, {"previousFY": 0, "lastFY": 0, "currentFY": 0})
                    prev_val = round(totals["previousFY"], 2)
                    last_val = round(totals["lastFY"], 2)
                    curr_val = round(totals["currentFY"], 2)
                    addr_ws.cell(row=row_idx, column=base_col, value=prev_val)
                    addr_ws.cell(row=row_idx, column=base_col + 1, value=last_val)
                    _write_growth(addr_ws, row_idx, base_col + 2, _yoy_pct(last_val, prev_val))
                    addr_ws.cell(row=row_idx, column=base_col + 3, value=curr_val)
                    _write_growth(addr_ws, row_idx, base_col + 4, _yoy_pct(curr_val, last_val))

        for column in addr_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            addr_ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate filename with current date
        filename = f"customers_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        # Return as streaming response
        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Error in download_customer_analytics_report: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/brand-breakdown")
def get_brand_breakdown(
    customer_id: str = Query(..., description="Customer ID to get brand breakdown for"),
):
    try:
        current_date = datetime.now()
        current_year = current_date.year
        current_month = current_date.month

        if current_month >= 4:
            current_fy_start_year = current_year
        else:
            current_fy_start_year = current_year - 1

        last_fy_start_year = current_fy_start_year - 1
        previous_fy_start_year = current_fy_start_year - 2

        pipeline = [
            {
                "$match": {
                    "customer_id": customer_id,
                    "date": {"$gte": "2023-04-01"},
                    "status": {"$nin": ["void", "draft"]},
                }
            },
            {
                "$addFields": {
                    "invoiceYear": {
                        "$year": {"$dateFromString": {"dateString": "$date"}}
                    },
                    "invoiceMonth": {
                        "$month": {"$dateFromString": {"dateString": "$date"}}
                    },
                }
            },
            {
                "$addFields": {
                    "fyLabel": {
                        "$cond": {
                            "if": {"$gte": ["$invoiceMonth", 4]},
                            "then": {
                                "$concat": [
                                    "FY ",
                                    {"$toString": "$invoiceYear"},
                                    "-",
                                    {
                                        "$substr": [
                                            {"$toString": {"$add": ["$invoiceYear", 1]}},
                                            2,
                                            2,
                                        ]
                                    },
                                ]
                            },
                            "else": {
                                "$concat": [
                                    "FY ",
                                    {
                                        "$toString": {
                                            "$subtract": ["$invoiceYear", 1]
                                        }
                                    },
                                    "-",
                                    {"$substr": [{"$toString": "$invoiceYear"}, 2, 2]},
                                ]
                            },
                        }
                    },
                    "isCurrentFY": {
                        "$or": [
                            {
                                "$and": [
                                    {"$eq": ["$invoiceYear", current_fy_start_year]},
                                    {"$gte": ["$invoiceMonth", 4]},
                                ]
                            },
                            {
                                "$and": [
                                    {
                                        "$eq": [
                                            "$invoiceYear",
                                            current_fy_start_year + 1,
                                        ]
                                    },
                                    {"$lte": ["$invoiceMonth", 3]},
                                ]
                            },
                        ]
                    },
                    "isLastFY": {
                        "$or": [
                            {
                                "$and": [
                                    {"$eq": ["$invoiceYear", last_fy_start_year]},
                                    {"$gte": ["$invoiceMonth", 4]},
                                ]
                            },
                            {
                                "$and": [
                                    {
                                        "$eq": [
                                            "$invoiceYear",
                                            last_fy_start_year + 1,
                                        ]
                                    },
                                    {"$lte": ["$invoiceMonth", 3]},
                                ]
                            },
                        ]
                    },
                    "isPreviousFY": {
                        "$or": [
                            {
                                "$and": [
                                    {"$eq": ["$invoiceYear", previous_fy_start_year]},
                                    {"$gte": ["$invoiceMonth", 4]},
                                ]
                            },
                            {
                                "$and": [
                                    {
                                        "$eq": [
                                            "$invoiceYear",
                                            previous_fy_start_year + 1,
                                        ]
                                    },
                                    {"$lte": ["$invoiceMonth", 3]},
                                ]
                            },
                        ]
                    },
                }
            },
            {"$unwind": {"path": "$line_items", "preserveNullAndEmptyArrays": False}},
            {
                "$lookup": {
                    "from": "products",
                    "localField": "line_items.item_id",
                    "foreignField": "item_id",
                    "as": "product_info",
                }
            },
            {
                "$addFields": {
                    "brand": {
                        "$ifNull": [
                            {"$arrayElemAt": ["$product_info.brand", 0]},
                            "Unknown",
                        ]
                    }
                }
            },
            {
                "$group": {
                    "_id": "$brand",
                    "currentFY": {
                        "$sum": {
                            "$cond": [
                                "$isCurrentFY",
                                {"$ifNull": ["$line_items.item_total", 0]},
                                0,
                            ]
                        }
                    },
                    "lastFY": {
                        "$sum": {
                            "$cond": [
                                "$isLastFY",
                                {"$ifNull": ["$line_items.item_total", 0]},
                                0,
                            ]
                        }
                    },
                    "previousFY": {
                        "$sum": {
                            "$cond": [
                                "$isPreviousFY",
                                {"$ifNull": ["$line_items.item_total", 0]},
                                0,
                            ]
                        }
                    },
                    # Quarterly: Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar
                    "currentFY_Q1": {"$sum": {"$cond": [{"$and": ["$isCurrentFY", {"$in": ["$invoiceMonth", [4, 5, 6]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "currentFY_Q2": {"$sum": {"$cond": [{"$and": ["$isCurrentFY", {"$in": ["$invoiceMonth", [7, 8, 9]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "currentFY_Q3": {"$sum": {"$cond": [{"$and": ["$isCurrentFY", {"$in": ["$invoiceMonth", [10, 11, 12]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "currentFY_Q4": {"$sum": {"$cond": [{"$and": ["$isCurrentFY", {"$in": ["$invoiceMonth", [1, 2, 3]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "lastFY_Q1": {"$sum": {"$cond": [{"$and": ["$isLastFY", {"$in": ["$invoiceMonth", [4, 5, 6]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "lastFY_Q2": {"$sum": {"$cond": [{"$and": ["$isLastFY", {"$in": ["$invoiceMonth", [7, 8, 9]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "lastFY_Q3": {"$sum": {"$cond": [{"$and": ["$isLastFY", {"$in": ["$invoiceMonth", [10, 11, 12]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "lastFY_Q4": {"$sum": {"$cond": [{"$and": ["$isLastFY", {"$in": ["$invoiceMonth", [1, 2, 3]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "previousFY_Q1": {"$sum": {"$cond": [{"$and": ["$isPreviousFY", {"$in": ["$invoiceMonth", [4, 5, 6]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "previousFY_Q2": {"$sum": {"$cond": [{"$and": ["$isPreviousFY", {"$in": ["$invoiceMonth", [7, 8, 9]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "previousFY_Q3": {"$sum": {"$cond": [{"$and": ["$isPreviousFY", {"$in": ["$invoiceMonth", [10, 11, 12]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                    "previousFY_Q4": {"$sum": {"$cond": [{"$and": ["$isPreviousFY", {"$in": ["$invoiceMonth", [1, 2, 3]]}]}, {"$ifNull": ["$line_items.item_total", 0]}, 0]}},
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "brand": "$_id",
                    "currentFY": {"$round": ["$currentFY", 2]},
                    "lastFY": {"$round": ["$lastFY", 2]},
                    "previousFY": {"$round": ["$previousFY", 2]},
                    "currentFY_Q1": {"$round": ["$currentFY_Q1", 2]},
                    "currentFY_Q2": {"$round": ["$currentFY_Q2", 2]},
                    "currentFY_Q3": {"$round": ["$currentFY_Q3", 2]},
                    "currentFY_Q4": {"$round": ["$currentFY_Q4", 2]},
                    "lastFY_Q1": {"$round": ["$lastFY_Q1", 2]},
                    "lastFY_Q2": {"$round": ["$lastFY_Q2", 2]},
                    "lastFY_Q3": {"$round": ["$lastFY_Q3", 2]},
                    "lastFY_Q4": {"$round": ["$lastFY_Q4", 2]},
                    "previousFY_Q1": {"$round": ["$previousFY_Q1", 2]},
                    "previousFY_Q2": {"$round": ["$previousFY_Q2", 2]},
                    "previousFY_Q3": {"$round": ["$previousFY_Q3", 2]},
                    "previousFY_Q4": {"$round": ["$previousFY_Q4", 2]},
                }
            },
            {"$sort": {"currentFY": -1}},
        ]

        results = list(db.invoices.aggregate(pipeline))

        def _calc_growth(current, previous):
            if previous == 0:
                return None
            return round((current - previous) / previous * 100, 1)

        for r in results:
            r["yoyGrowth"] = {
                "prevToLast": _calc_growth(r.get("lastFY", 0), r.get("previousFY", 0)),
                "lastToCurrent": _calc_growth(r.get("currentFY", 0), r.get("lastFY", 0)),
            }
            r["quarterlyGrowth"] = {
                q: {
                    "prevToLast": _calc_growth(r.get(f"lastFY_{q}", 0), r.get(f"previousFY_{q}", 0)),
                    "lastToCurrent": _calc_growth(r.get(f"currentFY_{q}", 0), r.get(f"lastFY_{q}", 0)),
                }
                for q in ["Q1", "Q2", "Q3", "Q4"]
            }

        # Build FY labels
        current_fy_label = f"FY {current_fy_start_year}-{str(current_fy_start_year + 1)[2:]}"
        last_fy_label = f"FY {last_fy_start_year}-{str(last_fy_start_year + 1)[2:]}"
        previous_fy_label = f"FY {previous_fy_start_year}-{str(previous_fy_start_year + 1)[2:]}"

        return JSONResponse(
            content={
                "brands": results,
                "fyLabels": {
                    "currentFY": current_fy_label,
                    "lastFY": last_fy_label,
                    "previousFY": previous_fy_label,
                },
            }
        )

    except Exception as e:
        logger.error(f"Error in get_brand_breakdown: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})
