"""
Distributor portal — the read-only MVP surface a distributor sees after login.

Scope discipline: a distributor must only ever see their own brand. Every query
here derives its filter from `_distributor_scope()`, which reads the registration
id off the JWT — never off a path or body parameter. Do not hand-roll a filter in
a handler; that is where a cross-tenant leak eventually gets introduced.

Router-level JWT is applied at include time in api.py; `_distributor_scope` adds
the role check on top, because a valid token alone is not authorisation.
"""

import io
import os
import time
from datetime import datetime, timedelta
from typing import Optional

from bson.objectid import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from pymongo import UpdateOne

from ..config.auth import get_current_user
from ..config.root import get_database, serialize_mongo_document

router = APIRouter()
db = get_database()


def _distributor_scope(current_user: dict = Depends(get_current_user)) -> dict:
    """Resolve the caller to exactly one distributor registration.

    Returns the registration document. Raises rather than returning None so a
    handler can never accidentally treat "no scope" as "all rows".
    """
    payload = current_user.get("data") or current_user

    if payload.get("role") != "distributor":
        raise HTTPException(status_code=403, detail="Distributor access only")

    registration_id = payload.get("distributor_id")
    if not registration_id or not ObjectId.is_valid(str(registration_id)):
        raise HTTPException(
            status_code=403, detail="This login is not linked to a distributor"
        )

    registration = db.distributor_registrations.find_one(
        {"_id": ObjectId(str(registration_id))}
    )
    if not registration:
        raise HTTPException(status_code=404, detail="Distributor record not found")

    return registration


def _brand_filter(registration: dict) -> dict:
    """Products belonging to this distributor's brand.

    Brand is stored as a free-text string on products, so match case-insensitively
    and anchored — a substring match would let "Zippy" pull in "Zippy Paws Pro".
    """
    brand = (registration.get("brand_name") or "").strip()
    if not brand:
        # No brand on the registration means nothing is theirs yet. Return a
        # filter that matches nothing rather than one that matches everything.
        return {"_id": {"$in": []}}
    return {"brand": {"$regex": f"^{brand}$", "$options": "i"}}


@router.get("/me")
def get_profile(registration: dict = Depends(_distributor_scope)):
    """Profile card for the dashboard header and account page."""
    return {
        "distributor": serialize_mongo_document(
            {
                "_id": registration["_id"],
                "company_name": registration.get("company_name", ""),
                "brand_name": registration.get("brand_name", ""),
                "contact_person_name": registration.get("contact_person_name", ""),
                "email": registration.get("email", ""),
                "phone": registration.get("phone", ""),
                "gst_number": registration.get("gst_number", ""),
                "pan_number": registration.get("pan_number", ""),
                "categories": registration.get("categories", []),
                "distribution_states": registration.get("distribution_states", []),
                "billing_address": registration.get("billing_address", {}),
                "ship_from_address": registration.get("ship_from_address", {}),
                # Proposed is what they typed on the application form; agreed is
                # what an admin later signs off. Kept separate on purpose — see
                # the admin route. `agreed_margin` is absent until then.
                "proposed_margin": registration.get("margin", ""),
                "agreed_margin": registration.get("agreed_margin"),
                "status": registration.get("status", ""),
                "created_at": registration.get("created_at"),
            }
        )
    }


@router.get("/dashboard")
def get_dashboard(registration: dict = Depends(_distributor_scope)):
    """Headline counts. Deliberately cheap — no per-order aggregation."""
    brand_filter = _brand_filter(registration)

    products_total = db.products.count_documents(brand_filter)
    products_active = db.products.count_documents({**brand_filter, "status": "active"})

    product_names = [
        p["name"]
        for p in db.products.find(brand_filter, {"name": 1})
        if p.get("name")
    ]

    thirty_days_ago = datetime.now() - timedelta(days=30)
    orders_30d = 0
    units_30d = 0
    if product_names:
        cursor = db.orders.find(
            {
                "products.name": {"$in": product_names},
                "created_at": {"$gte": thirty_days_ago},
            },
            {"products": 1},
        )
        name_set = set(product_names)
        for order in cursor:
            matched = [
                p
                for p in (order.get("products") or [])
                if p.get("name") in name_set and _as_int(p.get("quantity")) > 0
            ]
            if matched:
                orders_30d += 1
                units_30d += sum(_as_int(p.get("quantity")) for p in matched)

    return {
        "brand_name": registration.get("brand_name", ""),
        "products_total": products_total,
        "products_active": products_active,
        "states_served": len(registration.get("distribution_states") or []),
        "categories": registration.get("categories", []),
        "orders_30d": orders_30d,
        "units_30d": units_30d,
        "agreed_margin": registration.get("agreed_margin"),
        "status": registration.get("status", ""),
    }


def _as_int(value, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError, AttributeError):
        return default


@router.get("/products")
def list_products(
    registration: dict = Depends(_distributor_scope),
    page: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None, description="active | inactive"),
    stock: Optional[str] = Query(None, description="zero | gt_zero"),
):
    """This distributor's catalogue as we hold it. Read-only in the MVP.

    `status` and `stock` take the same values as the admin products filter
    (`active|inactive`, `zero|gt_zero`) so the two screens behave identically.
    """
    match = _brand_filter(registration)
    if search and search.strip():
        term = search.strip()
        match = {
            **match,
            "$or": [
                {"name": {"$regex": term, "$options": "i"}},
                {"cf_sku_code": {"$regex": term, "$options": "i"}},
            ],
        }
    if status in ("active", "inactive"):
        match["status"] = status
    if stock == "zero":
        match["stock"] = {"$lte": 0}
    elif stock == "gt_zero":
        match["stock"] = {"$gt": 0}

    total = db.products.count_documents(match)
    cursor = (
        db.products.find(
            match,
            {
                "name": 1,
                "cf_sku_code": 1,
                "category": 1,
                "sub_category": 1,
                "rate": 1,
                "stock": 1,
                "status": 1,
                "hsn_or_sac": 1,
                "image_url": 1,
                "images": 1,
                # Needed by the order-form preview, which renders the real
                # ProductCard and reads tax/pre-order/clearance off the product.
                "brand": 1,
                "item_tax_preferences": 1,
                "pre_order": 1,
                "upcoming_stock": 1,
                "clearance": 1,
                "clearance_margin": 1,
                "catalogue_order": 1,
            },
        )
        .sort("name", 1)
        .skip(page * limit)
        .limit(limit)
    )

    return {
        "products": [serialize_mongo_document(d) for d in cursor],
        "total_count": total,
        "page": page,
        "per_page": limit,
        "total_pages": max(1, (total + limit - 1) // limit),
    }


def _orders_match(
    registration: dict,
    search: Optional[str],
    status: Optional[str],
    state: Optional[str],
    estimate: Optional[str],
) -> tuple[dict, set]:
    """Build the orders filter once so the table and the PDF export can never
    disagree about what "all the data" means."""
    name_set = {
        p["name"]
        for p in db.products.find(_brand_filter(registration), {"name": 1})
        if p.get("name")
    }
    if not name_set:
        return {}, name_set

    match: dict = {"products.name": {"$in": list(name_set)}}
    if status:
        match["status"] = status
    if state:
        match["shipping_address.state"] = state
    if estimate == "created":
        match["estimate_number"] = {"$nin": [None, ""]}
    elif estimate == "not_created":
        match["$or"] = [
            {"estimate_number": {"$in": [None, ""]}},
            {"estimate_number": {"$exists": False}},
        ]
    if search and search.strip():
        term = search.strip()
        # $and-wrapped so it can't clobber the $or the estimate filter may set
        match["$and"] = [
            {
                "$or": [
                    {"estimate_number": {"$regex": term, "$options": "i"}},
                    {"customer_name": {"$regex": term, "$options": "i"}},
                ]
            }
        ]
    return match, name_set


def _shape_order(order: dict, name_set: set) -> Optional[dict]:
    """Reduce one order to this distributor's slice of it, or None if nothing
    of theirs survives the quantity filter."""
    mine = [
        p
        for p in (order.get("products") or [])
        if p.get("name") in name_set and _as_int(p.get("quantity")) > 0
    ]
    if not mine:
        return None
    return serialize_mongo_document(
        {
            "_id": order["_id"],
            "estimate_number": order.get("estimate_number") or "",
            "customer_name": order.get("customer_name", ""),
            "created_at": order.get("created_at"),
            "status": order.get("status", ""),
            "state": (order.get("shipping_address") or {}).get("state", ""),
            "line_count": len(mine),
            "units": sum(_as_int(p.get("quantity")) for p in mine),
            "products": [
                {
                    "name": p.get("name", ""),
                    "quantity": _as_int(p.get("quantity")),
                    "product_code": p.get("product_code", ""),
                }
                for p in mine
            ],
        }
    )


_ORDER_PROJECTION = {
    "estimate_number": 1,
    "customer_name": 1,
    "created_at": 1,
    "status": 1,
    "products": 1,
    "shipping_address.state": 1,
}


@router.get("/orders/filter-options")
def order_filter_options(registration: dict = Depends(_distributor_scope)):
    """Dropdown values for the orders filters, restricted to states this
    distributor actually sells into so the list stays short and relevant."""
    return {
        "statuses": ["draft", "sent", "accepted", "declined", "invoiced"],
        "states": sorted(registration.get("distribution_states") or []),
    }


@router.get("/orders")
def list_orders(
    registration: dict = Depends(_distributor_scope),
    page: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None, description="Estimate number or retailer"),
    status: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    estimate: Optional[str] = Query(None, description="created | not_created"),
):
    """Orders containing this distributor's products.

    Line items are filtered down to their SKUs before returning: a mixed-brand
    order must not leak another brand's products, quantities or the order total.
    """
    match, name_set = _orders_match(registration, search, status, state, estimate)
    if not name_set:
        return {"orders": [], "total_count": 0, "page": page,
                "per_page": limit, "total_pages": 1}

    total = db.orders.count_documents(match)
    cursor = (
        db.orders.find(match, _ORDER_PROJECTION)
        .sort("created_at", -1)
        .skip(page * limit)
        .limit(limit)
    )
    orders = [o for o in (_shape_order(d, name_set) for d in cursor) if o]

    return {
        "orders": orders,
        "total_count": total,
        "page": page,
        "per_page": limit,
        "total_pages": max(1, (total + limit - 1) // limit),
    }


# Hard ceiling on the spreadsheet export. A distributor with years of history
# would otherwise build a response big enough to time out the request.
EXPORT_MAX_ORDERS = 5000


@router.get("/orders/export")
def export_orders(
    registration: dict = Depends(_distributor_scope),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    state: Optional[str] = Query(None),
    estimate: Optional[str] = Query(None),
):
    """Every order matching the current filters, flattened for a spreadsheet.

    Unpaginated by design — the table on screen is paged, this is the whole
    filtered set. One row per order, with the item list collapsed into a cell.
    """
    match, name_set = _orders_match(registration, search, status, state, estimate)
    if not name_set:
        return {"orders": [], "total_count": 0, "truncated": False}

    matched_total = db.orders.count_documents(match)
    cursor = (
        db.orders.find(match, _ORDER_PROJECTION)
        .sort("created_at", -1)
        .limit(EXPORT_MAX_ORDERS)
    )

    rows = []
    for doc in cursor:
        shaped = _shape_order(doc, name_set)
        if not shaped:
            continue
        rows.append(
            {
                "reference": shaped["estimate_number"]
                or f"#{str(shaped['_id'])[-8:]}",
                "estimate_number": shaped["estimate_number"],
                "estimate_raised": "Yes" if shaped["estimate_number"] else "No",
                "order_id": str(shaped["_id"]),
                "customer_name": shaped["customer_name"],
                "state": shaped["state"],
                "status": shaped["status"],
                "created_at": shaped["created_at"],
                "line_count": shaped["line_count"],
                "units": shaped["units"],
                "items": "; ".join(
                    f"{p['quantity']} x {p['name']}" for p in shaped["products"]
                ),
            }
        )

    return {
        "orders": rows,
        "total_count": matched_total,
        "truncated": matched_total > EXPORT_MAX_ORDERS,
    }


@router.get("/orders/{order_id}/pdf")
def export_order_pdf(
    order_id: str, registration: dict = Depends(_distributor_scope)
):
    """One order as a branded PDF sheet, scoped to this distributor's items.

    The scope check is the `_shape_order` result: if none of the order's line
    items belong to this brand it 404s, so an id guessed from another
    distributor's order reveals nothing.
    """
    from ..config.pdf import build_order_pdf  # WeasyPrint is heavy — import late

    if not ObjectId.is_valid(order_id):
        raise HTTPException(status_code=400, detail="Invalid order id")

    name_set = {
        p["name"]
        for p in db.products.find(_brand_filter(registration), {"name": 1})
        if p.get("name")
    }
    doc = db.orders.find_one({"_id": ObjectId(order_id)}, _ORDER_PROJECTION)
    shaped = _shape_order(doc, name_set) if doc else None
    if not shaped:
        raise HTTPException(status_code=404, detail="Order not found")

    pdf_bytes = build_order_pdf(registration, shaped)
    ref = (shaped["estimate_number"] or str(shaped["_id"])[-8:]).replace("/", "-")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="order_{ref}.pdf"'},
    )


# ---------------------------------------------------------------------------
# Brand profile
#
# Kept in its own collection rather than on the registration doc: the
# registration is the application record (what they claimed at signup, plus the
# admin's follow-up state) and should stay an immutable-ish audit trail, while
# the brand profile is marketing content the distributor edits freely.
# ---------------------------------------------------------------------------

MAX_IMAGE_MB = 5
_IMAGE_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp", "image/svg+xml"}


def _brand_profile(registration: dict) -> dict:
    return db.distributor_brand_profiles.find_one(
        {"distributor_id": registration["_id"]}
    ) or {}


@router.get("/brand-profile")
def get_brand_profile(registration: dict = Depends(_distributor_scope)):
    profile = _brand_profile(registration)
    return {
        "brand_profile": {
            "brand_name": registration.get("brand_name", ""),
            "description": profile.get("description", ""),
            "tagline": profile.get("tagline", ""),
            "website": profile.get("website", ""),
            "logo_url": profile.get("logo_url", ""),
            "secondary_image_url": profile.get("secondary_image_url", ""),
            "catalogues": [
                serialize_mongo_document(c) for c in profile.get("catalogues", [])
            ],
            "updated_at": profile.get("updated_at"),
        }
    }


@router.put("/brand-profile")
def update_brand_profile(
    body: dict, registration: dict = Depends(_distributor_scope)
):
    """Text fields only — images go through the upload endpoint below."""
    allowed = {"description", "tagline", "website"}
    updates = {k: str(v or "").strip() for k, v in (body or {}).items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    if len(updates.get("description", "")) > 4000:
        raise HTTPException(status_code=400, detail="Description is too long")

    db.distributor_brand_profiles.update_one(
        {"distributor_id": registration["_id"]},
        {
            "$set": {**updates, "updated_at": datetime.now()},
            "$setOnInsert": {
                "distributor_id": registration["_id"],
                "created_at": datetime.now(),
            },
        },
        upsert=True,
    )
    return {"success": True}


@router.post("/brand-profile/image")
async def upload_brand_image(
    kind: str = Query(..., description="logo | secondary"),
    file: UploadFile = File(...),
    registration: dict = Depends(_distributor_scope),
):
    """Upload the brand logo or secondary image to S3.

    Keyed by distributor id so two brands can never collide, and prefixed
    `distributor_brands/` to keep it out of the product image namespace.
    """
    from .admin import s3_client, AWS_S3_BUCKET_NAME, AWS_S3_URL

    if kind not in ("logo", "secondary"):
        raise HTTPException(status_code=400, detail="kind must be logo or secondary")
    if file.content_type not in _IMAGE_TYPES:
        raise HTTPException(status_code=400, detail="Upload a PNG, JPG, WEBP or SVG")

    contents = await file.read()
    if len(contents) > MAX_IMAGE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=400, detail=f"Image must be under {MAX_IMAGE_MB} MB"
        )
    file.file.seek(0)

    ext = os.path.splitext(file.filename or "")[1] or ".png"
    key = (
        f"distributor_brands/{registration['_id']}_{kind}_"
        f"{int(time.time() * 1000)}{ext}"
    )
    s3_client.upload_fileobj(
        file.file,
        AWS_S3_BUCKET_NAME,
        key,
        ExtraArgs={"ACL": "public-read", "ContentType": file.content_type},
    )
    url = f"{AWS_S3_URL}/{key}"

    field = "logo_url" if kind == "logo" else "secondary_image_url"
    db.distributor_brand_profiles.update_one(
        {"distributor_id": registration["_id"]},
        {
            "$set": {field: url, "updated_at": datetime.now()},
            "$setOnInsert": {
                "distributor_id": registration["_id"],
                "created_at": datetime.now(),
            },
        },
        upsert=True,
    )
    return {"success": True, "url": url}


# ---------------------------------------------------------------------------
# Stock
#
# Distributor products are not in Zoho, so nothing syncs their stock — the brand
# maintains it. Every submission is written to `distributor_stock_entries` as a
# dated snapshot before `products` is touched, so there is always a record of
# what a brand claimed and when. `products.stock` holds the latest value.
# ---------------------------------------------------------------------------

STOCK_SOURCES = ("manual", "upload")


def _record_stock_entries(
    registration: dict, rows: list, source: str, filename: str = ""
) -> int:
    """Append-only audit trail. Written before products are updated so a failed
    product write still leaves evidence of the attempt."""
    if not rows:
        return 0
    now = datetime.now()
    db.distributor_stock_entries.insert_many(
        [
            {
                "distributor_id": registration["_id"],
                "brand_name": registration.get("brand_name", ""),
                "product_id": r["product_id"],
                "product_name": r.get("product_name", ""),
                "sku": r.get("sku", ""),
                "previous_stock": r.get("previous_stock"),
                "stock": r.get("stock"),
                "upcoming_stock": r.get("upcoming_stock"),
                "source": source,
                "filename": filename,
                "created_at": now,
            }
            for r in rows
        ]
    )
    return len(rows)


def _own_product(registration: dict, product_id: str) -> dict:
    """Fetch a product only if it belongs to this distributor's brand."""
    if not ObjectId.is_valid(product_id):
        raise HTTPException(status_code=400, detail="Invalid product id")
    product = db.products.find_one(
        {"_id": ObjectId(product_id), **_brand_filter(registration)}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/products/{product_id}/stock")
def set_product_stock(
    product_id: str, body: dict, registration: dict = Depends(_distributor_scope)
):
    """Set current and/or upcoming stock for a single product."""
    product = _own_product(registration, product_id)

    updates: dict = {}
    if "stock" in (body or {}):
        stock = _as_int(body.get("stock"), -1)
        if stock < 0:
            raise HTTPException(status_code=400, detail="Stock must be zero or more")
        updates["stock"] = stock
    if "upcoming_stock" in (body or {}):
        upcoming = _as_int(body.get("upcoming_stock"), -1)
        if upcoming < 0:
            raise HTTPException(status_code=400, detail="Upcoming stock must be zero or more")
        updates["upcoming_stock"] = upcoming

    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    _record_stock_entries(
        registration,
        [
            {
                "product_id": product["_id"],
                "product_name": product.get("name", ""),
                "sku": product.get("cf_sku_code", ""),
                "previous_stock": product.get("stock"),
                "stock": updates.get("stock", product.get("stock")),
                "upcoming_stock": updates.get("upcoming_stock"),
            }
        ],
        source="manual",
    )
    db.products.update_one(
        {"_id": product["_id"]},
        {"$set": {**updates, "stock_updated_at": datetime.now()}},
    )
    return {"success": True, **updates}


@router.get("/stock/template")
def stock_template(registration: dict = Depends(_distributor_scope)):
    """The distributor's SKUs pre-filled, so an upload can't invent products."""
    cursor = db.products.find(
        _brand_filter(registration),
        {"name": 1, "cf_sku_code": 1, "stock": 1, "upcoming_stock": 1},
    ).sort("name", 1)
    return {
        "rows": [
            {
                "sku": p.get("cf_sku_code", ""),
                "name": p.get("name", ""),
                "stock": p.get("stock", 0),
                "upcoming_stock": p.get("upcoming_stock", 0),
            }
            for p in cursor
        ]
    }


@router.post("/stock/upload")
async def upload_stock(
    file: UploadFile = File(...), registration: dict = Depends(_distributor_scope)
):
    """Bulk stock update from a spreadsheet.

    Matched strictly on SKU against this distributor's own products — an
    unknown SKU is reported back as skipped rather than creating anything, so
    the upload can never introduce a product or touch another brand's row.
    """
    from openpyxl import load_workbook

    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload a .xlsx file")

    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read that spreadsheet")

    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_sku = col("sku", "sku code", "cf_sku_code")
    i_stock = col("stock", "current stock", "quantity")
    i_upcoming = col("upcoming_stock", "upcoming stock", "incoming stock")

    if i_sku is None or (i_stock is None and i_upcoming is None):
        raise HTTPException(
            status_code=400,
            detail="The sheet needs an 'SKU' column and a 'Stock' or 'Upcoming Stock' column",
        )

    own = {
        (p.get("cf_sku_code") or "").strip().upper(): p
        for p in db.products.find(
            _brand_filter(registration), {"name": 1, "cf_sku_code": 1, "stock": 1}
        )
        if p.get("cf_sku_code")
    }

    entries, operations, skipped = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or i_sku >= len(row):
            continue
        sku = str(row[i_sku] or "").strip().upper()
        if not sku:
            continue
        product = own.get(sku)
        if not product:
            skipped.append(sku)
            continue

        updates: dict = {}
        if i_stock is not None and i_stock < len(row) and row[i_stock] is not None:
            updates["stock"] = max(0, _as_int(row[i_stock]))
        if i_upcoming is not None and i_upcoming < len(row) and row[i_upcoming] is not None:
            updates["upcoming_stock"] = max(0, _as_int(row[i_upcoming]))
        if not updates:
            continue

        entries.append(
            {
                "product_id": product["_id"],
                "product_name": product.get("name", ""),
                "sku": product.get("cf_sku_code", ""),
                "previous_stock": product.get("stock"),
                "stock": updates.get("stock", product.get("stock")),
                "upcoming_stock": updates.get("upcoming_stock"),
            }
        )
        operations.append(
            UpdateOne(
                {"_id": product["_id"]},
                {"$set": {**updates, "stock_updated_at": datetime.now()}},
            )
        )

    if not operations:
        raise HTTPException(
            status_code=400,
            detail="No matching SKUs found. Download the template to see your SKUs.",
        )

    _record_stock_entries(
        registration, entries, source="upload", filename=file.filename or ""
    )
    db.products.bulk_write(operations)

    return {
        "success": True,
        "updated": len(operations),
        "skipped": len(skipped),
        # Capped so one bad sheet can't return thousands of strings.
        "skipped_skus": skipped[:25],
    }


@router.get("/stock/history")
def stock_history(
    registration: dict = Depends(_distributor_scope),
    page: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
):
    """What this brand submitted, and when."""
    match = {"distributor_id": registration["_id"]}
    total = db.distributor_stock_entries.count_documents(match)
    cursor = (
        db.distributor_stock_entries.find(match)
        .sort("created_at", -1)
        .skip(page * limit)
        .limit(limit)
    )
    return {
        "entries": [serialize_mongo_document(d) for d in cursor],
        "total_count": total,
        "page": page,
        "per_page": limit,
        "total_pages": max(1, (total + limit - 1) // limit),
    }


# ---------------------------------------------------------------------------
# Brand catalogue
#
# Sales collateral (a PDF/deck the brand wants retailers and our sales team to
# see). Stored as a list on the brand profile rather than its own collection —
# it is brand content, and there will only ever be a handful per distributor.
# ---------------------------------------------------------------------------

MAX_CATALOGUE_MB = 25
_CATALOGUE_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "image/png",
    "image/jpeg",
}


@router.post("/brand-profile/catalogue")
async def upload_brand_catalogue(
    file: UploadFile = File(...),
    title: Optional[str] = Query(None),
    registration: dict = Depends(_distributor_scope),
):
    from .admin import s3_client, AWS_S3_BUCKET_NAME, AWS_S3_URL

    if file.content_type not in _CATALOGUE_TYPES:
        raise HTTPException(status_code=400, detail="Upload a PDF, PPTX, PNG or JPG")

    contents = await file.read()
    if len(contents) > MAX_CATALOGUE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=400, detail=f"File must be under {MAX_CATALOGUE_MB} MB"
        )
    file.file.seek(0)

    ext = os.path.splitext(file.filename or "")[1] or ".pdf"
    key = (
        f"distributor_catalogues/{registration['_id']}_"
        f"{int(time.time() * 1000)}{ext}"
    )
    s3_client.upload_fileobj(
        file.file,
        AWS_S3_BUCKET_NAME,
        key,
        ExtraArgs={"ACL": "public-read", "ContentType": file.content_type},
    )

    entry = {
        # Own id so a single catalogue can be removed without matching on URL.
        "_id": ObjectId(),
        "title": (title or os.path.splitext(file.filename or "Catalogue")[0]).strip(),
        "filename": file.filename or "",
        "url": f"{AWS_S3_URL}/{key}",
        "s3_key": key,
        "size_bytes": len(contents),
        "content_type": file.content_type,
        "uploaded_at": datetime.now(),
    }
    db.distributor_brand_profiles.update_one(
        {"distributor_id": registration["_id"]},
        {
            "$push": {"catalogues": entry},
            "$set": {"updated_at": datetime.now()},
            "$setOnInsert": {
                "distributor_id": registration["_id"],
                "created_at": datetime.now(),
            },
        },
        upsert=True,
    )
    return {"success": True, "catalogue": serialize_mongo_document(entry)}


@router.delete("/brand-profile/catalogue/{catalogue_id}")
def delete_brand_catalogue(
    catalogue_id: str, registration: dict = Depends(_distributor_scope)
):
    """Removes the entry from the profile. The S3 object is left in place —
    deleting it would break any link already shared with a retailer."""
    if not ObjectId.is_valid(catalogue_id):
        raise HTTPException(status_code=400, detail="Invalid catalogue id")

    result = db.distributor_brand_profiles.update_one(
        {"distributor_id": registration["_id"]},
        {"$pull": {"catalogues": {"_id": ObjectId(catalogue_id)}}},
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Catalogue not found")
    return {"success": True}
