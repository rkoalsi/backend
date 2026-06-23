from types import NoneType
from pymongo.collection import Collection
from datetime import datetime
from typing import List, Dict, Tuple
from .helpers import get_access_token
from fastapi import APIRouter, HTTPException, Request, BackgroundTasks, Query
from ..config.root import get_database, serialize_mongo_document
from bson.objectid import ObjectId
import time, os, httpx, requests, asyncio, ssl, socket, re, io
from dotenv import load_dotenv
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..config.constants import terms, STATE_CODES 
from ..config.whatsapp import send_whatsapp
from .notifications import create_notification, create_notifications_for_roles
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from pathlib import Path
from collections import defaultdict

load_dotenv()

org_id = os.getenv("ORG_ID")
ESTIMATE_URL = os.getenv("ESTIMATE_URL")
PDF_URL = os.getenv("PDF_URL")


# Connect to MongoDB
db = get_database()
orders_collection = db["orders"]
customers_collection = db["customers"]
users_collection = db["users"]

router = APIRouter()

timeout = httpx.Timeout(30.0, connect=10.0, read=30.0, write=30.0)
# Rate limiting to prevent API quota exhaustion
class APIRateLimiter:
    def __init__(self, max_calls_per_minute=100):
        self.max_calls = max_calls_per_minute
        self.calls = defaultdict(list)
    
    async def wait_if_needed(self, key="default"):
        """Wait if rate limit would be exceeded"""
        now = time.time()
        minute_ago = now - 60
        
        # Clean old calls
        self.calls[key] = [call_time for call_time in self.calls[key] if call_time > minute_ago]
        
        # Check if we need to wait
        if len(self.calls[key]) >= self.max_calls:
            sleep_time = 60 - (now - self.calls[key][0])
            if sleep_time > 0:
                print(f"⏱️ Rate limiting: waiting {sleep_time:.1f} seconds")
                await asyncio.sleep(sleep_time)
        
        # Record this call
        self.calls[key].append(now)

# Global rate limiter
rate_limiter = APIRateLimiter(max_calls_per_minute=100)

def create_fresh_sheets_service():
    """Create a new service instance to avoid threading/memory issues"""
    credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)
async def robust_api_call(operation_name: str, func, max_retries: int = 3):
    """
    Execute Google API calls with robust error handling for SSL and connection issues
    """
    # Apply rate limiting
    await rate_limiter.wait_if_needed()
    
    for attempt in range(max_retries):
        try:
            result = await asyncio.to_thread(func)
            print(f"✓ {operation_name} succeeded")
            return result, None
        except (ssl.SSLError, socket.error, ConnectionError) as e:
            if attempt < max_retries - 1:
                wait_time = min((attempt + 1) * 2, 10)  # Cap at 10 seconds
                print(f"🔄 {operation_name} SSL/Connection error (attempt {attempt + 1}), retrying in {wait_time}s: {type(e).__name__}")
                await asyncio.sleep(wait_time)
                continue
            else:
                error_msg = f"{operation_name} failed after {max_retries} attempts: {str(e)}"
                print(f"✗ {error_msg}")
                return None, error_msg
        except HttpError as e:
            if e.resp.status == 429:  # Rate limit
                wait_time = 60
                print(f"⏱️ {operation_name} rate limited, waiting {wait_time}s")
                await asyncio.sleep(wait_time)
                if attempt < max_retries - 1:
                    continue
            error_msg = f"{operation_name} HTTP error: {e.resp.status} - {str(e)}"
            print(f"✗ {error_msg}")
            return None, error_msg
        except Exception as e:
            error_msg = f"{operation_name} unexpected error: {str(e)}"
            print(f"✗ {error_msg}")
            return None, error_msg
    
    return None, f"{operation_name} failed after all retries"

async def safe_execute(operation_name: str, func, critical: bool = True):
    """
    Execute an operation with proper error handling
    critical=True: Will raise HTTPException if fails
    critical=False: Will log error but continue
    """
    result, error = await robust_api_call(operation_name, func)
    
    if error and critical:
        raise HTTPException(status_code=500, detail=error)
    
    return result, error
# Create a new order
def create_order(order: dict, collection: Collection) -> str:
    # Explicitly convert customer_id and product_ids to ObjectId
    customer_id = order.get("customer_id", "")
    products = order.get("products", [])
    if customer_id:
        order["customer_id"] = ObjectId(order.get("customer_id"))

    if len(products) > 0:
        order["products"] = [
            {"product_id": ObjectId(item["product_id"]), "quantity": item["quantity"]}
            for item in products
        ]
    order["created_by"] = ObjectId(order.get("created_by", ""))
    order["created_at"] = datetime.now()
    order["updated_at"] = datetime.now()

    # Insert the document into MongoDB
    result = collection.insert_one(order)
    return str(result.inserted_id)


def check_if_order_exists(
    created_by: str, orders_collection: Collection
) -> dict | bool:
    order = orders_collection.find_one(
        {"created_by": ObjectId(created_by), "status": "draft"}
    )
    if order:
        return order
    else:
        return False


# Get an order by ID and populate customer and product details
def get_order(
    order_id: str,
    orders_collection: Collection,
):
    result = orders_collection.find_one({"_id": ObjectId(order_id)})
    if result:
        order = result
        order["status"] = str(order["status"]).capitalize()
        return serialize_mongo_document(order)
    return None


def get_all_orders(
    role: str,
    created_by: str,
    status: str,
    collection: Collection,
    users_collection: Collection,
):
    query = {}

    # Salesperson-specific query
    if role == "salesperson":
        if not created_by:
            raise ValueError("Salesperson role requires 'created_by'")
        query["created_by"] = ObjectId(created_by)
        query["is_deleted"] = {"$exists": False}
        query["$or"] = [
            {"total_amount": {"$gte": 0}},
            {"spreadsheet_created": True}
        ]
    if status:
        query["status"] = status

    # Fetch orders
    orders = collection.find(query).sort({"created_at": -1})

    # For admin, populate created_by_info with user information
    orders_with_user_info = []
    if "admin" in role:
        for order in orders:
            user_info = users_collection.find_one({"_id": order["created_by"]})
            if user_info:
                order["created_by_info"] = {
                    "id": str(user_info["_id"]),
                    "name": user_info.get("name"),
                    "email": user_info.get("email"),
                }
            orders_with_user_info.append(serialize_mongo_document(order))
    else:
        # For salesperson, no need to populate created_by_info
        orders_with_user_info = [serialize_mongo_document(order) for order in orders]

    return orders_with_user_info


# Update an order


def update_order(
    order_id: str,
    order_update: dict,
    order_collection: Collection,
    customer_collection: Collection,
):
    order_update["updated_at"] = datetime.now()
    if "created_by" in order_update:
        order_update["created_by"] = ObjectId(order_update.get("created_by"))
    # Handle customer updates
    if "customer_id" in order_update:
        customer_id = order_update.get("customer_id")
        customer = customer_collection.find_one({"_id": ObjectId(customer_id)})

        if customer:
            order_update["customer_id"] = ObjectId(customer_id)
            order_update["customer_name"] = (
                customer.get("company_name")
                if customer.get("company_name") != ""
                else customer.get("contact_name")
            )
            order_update["gst_type"] = (
                customer.get("cf_in_ex")
                if type(customer.get("cf_in_ex")) is not NoneType
                else "Exclusive"
            )

    # Handle product updates (replace the entire product list)
    if "products" in order_update:
        updated_products = []
        for product in order_update.get("products", []):
            product_id = ObjectId(product["_id"])
            updated_products.append(
                {
                    "product_id": product_id,
                    "tax_percentage": (
                        product.get("item_tax_preferences", [{}])[0].get(
                            "tax_percentage", 0
                        )
                    ),
                    "brand": product.get("brand", ""),
                    "product_code": product.get("cf_sku_code", ""),
                    "quantity": int(product.get("quantity") or 0),
                    "pre_order_quantity": int(product.get("pre_order_quantity") or 0),
                    "name": product.get("item_name") or product.get("name", ""),
                    "image_url": product.get("image_url", ""),
                    "margin": product.get("margin", ""),
                    "price": product.get("rate", 0),
                    "added_by": product.get("added_by", ""),
                    "pre_order": product.get("pre_order", False),
                }
            )
        # Replace the product list in the update payload
        order_update["products"] = updated_products
    # Perform the update in MongoDB
    order_collection.update_one({"_id": ObjectId(order_id)}, {"$set": order_update})


# Delete an order
def delete_order(order_id: str, deleted_by: str, collection: Collection):
    order = collection.find_one({"_id": ObjectId(order_id)})
    if not order.get("estimate_created", False):
        collection.update_one(
            {"_id": order.get("_id")},
            {
                "$set": {
                    "status": "deleted",
                    "is_deleted": True,
                    "deleted_at": datetime.now(),
                    "deleted_by": ObjectId(deleted_by),
                }
            },
        )
    else:
        raise HTTPException(
            status_code=400,
            detail="Order With Estimate Created Cannot Be Marked As Deleted",
        )


def clear_empty_orders(user_id: str, collection: Collection):
    orders = collection.find({"created_by": ObjectId(user_id)})
    for order in orders:
        if not order.get("customer_id"):
            collection.delete_one(
                {"_id": order.get("_id")},
                {
                    "$set": {
                        "status": "deleted",
                        "is_deleted": True,
                        "deleted_at": datetime.now(),
                    }
                },
            )


async def email_estimate(
    status: str,
    order_id: str,
    estimate_id: str,
    estimate_number: str,
    estimate_url: str,
    headers: dict,
    timeout: any,
) -> str:
    async with httpx.AsyncClient(timeout=timeout) as client:
        if status in {"accepted", "declined"}:
            await client.post(
                url=f"https://books.zoho.com/api/v3/estimates/{estimate_id}/status/sent?organization_id={org_id}",
                headers=headers,
            )
            status_response = await client.post(
                url=f"https://books.zoho.com/api/v3/estimates/{estimate_id}/status/{status}?organization_id={org_id}",
                headers=headers,
            )
            status_response.raise_for_status()
            db.orders.update_one(
                {"_id": ObjectId(order_id)},
                {
                    "$set": {
                        "status": f"{status}",
                        "estimate_created": True,
                        "estimate_id": estimate_id,
                        "estimate_number": estimate_number,
                        "estimate_url": estimate_url,
                    }
                },
            )
            return status_response.json().get("message", "")
    return ""


def clear_cart(order_id: str, orders_collection: Collection):
    order = orders_collection.update_one(
        {"_id": ObjectId(order_id)}, {"$set": {"products": []}}
    )
    return order.did_upsert


def validate_order(order_id: str):
    order = db.orders.find_one({"_id": ObjectId(order_id)})

    if not order:
        raise HTTPException(status_code=400, detail="Order not found")
    # Check if shipping address is missing or invalid
    customer_id = order.get("customer_id", "")
    customer = customers_collection.find_one({"_id": ObjectId(customer_id)})
    if customer.get("status") == "inactive":
        raise HTTPException(
            status_code=400, detail="Cannot Proceed, Customer is Inactive"
        )
    shipping_address = order.get("shipping_address", {}).get("address")
    if not shipping_address:
        raise HTTPException(status_code=400, detail="Shipping address is missing")

    # Check if billing address is missing or invalid
    billing_address = order.get("billing_address", {}).get("address")
    if not billing_address:
        raise HTTPException(status_code=400, detail="Billing address is missing")

    # Check if place of supply is missing or invalid
    place_of_supply = order.get("shipping_address", {}).get("state_code")
    state_str = str(order.get("shipping_address", {}).get("state", ""))
    place_of_supply_backup = STATE_CODES.get(state_str.title())
    if not place_of_supply and not place_of_supply_backup:
        raise HTTPException(status_code=400, detail="Place of supply is missing")

    # Check if products are missing or invalid
    products = order.get("products", [])
    if not products:
        raise HTTPException(status_code=400, detail="Products are missing")
    for product in products:
        doc = dict(db.products.find_one({"_id": ObjectId(product.get("product_id"))}))
        if doc.get("status") == "inactive":
            raise HTTPException(
                status_code=400, detail=f"Cannot Proceed, {doc.get('name')} is inactive"
            )
    # Check if total amount is missing or invalid
    total_amount = order.get("total_amount")
    if total_amount is None:
        raise HTTPException(status_code=400, detail="Total amount is missing")

    return True


# API Endpoints


# Create a new order
@router.post("/")
def create_new_order(order: dict, request: Request, background_tasks: BackgroundTasks):
    """
    Create a new order with raw dictionary data.
    """
    try:
        order_id = create_order(order, orders_collection)
        order["_id"] = order_id  # Add the generated ID back to the response

        created_by = order.get("created_by")
        if created_by:
            from .customer_activity import log_order_activity_for_user, extract_client_info
            ip, ua = extract_client_info(request)
            background_tasks.add_task(
                log_order_activity_for_user,
                action="create_order",
                user_id=str(created_by),
                metadata={"order_id": str(order_id)},
                ip_address=ip,
                user_agent=ua,
            )

        return serialize_mongo_document(order)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/check")
def check_order_status(order: dict):
    """
    Create a new order with raw dictionary data.
    """
    try:
        created_by = order.get("created_by", "")
        if not created_by:
            raise HTTPException(status_code=400, detail="created_by is required")
        order = check_if_order_exists(created_by, orders_collection)
        if order:
            return {
                **serialize_mongo_document(order),
                "message": "Existing Draft Order Found",
                "can_create": False,
            }
        else:
            return {"message": "Existing Draft Order Not Found", "can_create": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def get_active_products(sort: dict):
    products = list(
        db.products.aggregate(
            [
                {"$match": {"status": "active", "stock": {"$gt": 0}}},
                {"$sort": sort},
            ]
        )
    )
    return products


import asyncio
import json
import traceback
from pathlib import Path
from typing import List, Dict, Any, Optional
import time

from fastapi import HTTPException
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2.service_account import Credentials
from bson import ObjectId


BASE_DIR = Path(__file__).resolve().parent.parent
SERVICE_ACCOUNT_FILE = BASE_DIR / "creds.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Global service instances
_sheets_service = None
_drive_service = None

def get_sheets_service():
    global _sheets_service
    if _sheets_service is None:
        credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        _sheets_service = build("sheets", "v4", credentials=credentials)
    return _sheets_service

def get_drive_service():
    global _drive_service
    if _drive_service is None:
        credentials = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, 
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        _drive_service = build("drive", "v3", credentials=credentials)
    return _drive_service

async def safe_execute(operation_name: str, func, critical: bool = True):
    """
    Execute an operation with proper error handling
    critical=True: Will raise HTTPException if fails
    critical=False: Will log error but continue
    """
    try:
        result = await asyncio.to_thread(func)
        print(f"✓ {operation_name} succeeded")
        return result, None
    except HttpError as e:
        error_msg = f"{operation_name} failed: HTTP {e.resp.status} - {str(e)}"
        print(f"✗ {error_msg}")
        if critical:
            raise HTTPException(status_code=500, detail=error_msg)
        return None, error_msg
    except Exception as e:
        error_msg = f"{operation_name} failed: {str(e)}"
        print(f"✗ {error_msg}")
        if critical:
            raise HTTPException(status_code=500, detail=error_msg)
        return None, error_msg

def prepare_brand_data(brands: Dict, customer: Dict, special_margins: Dict, cart_products: List[Dict] = None) -> Dict[str, List]:
    """Prepare all data upfront with cart quantities pre-filled"""
    brand_data = {}

    # Create a mapping of product_id to quantity from cart
    cart_quantities = {}
    if cart_products:
        for cart_item in cart_products:
            product_id = str(cart_item.get("product_id", ""))
            quantity = cart_item.get("quantity", 0)
            if product_id and quantity:
                cart_quantities[product_id] = quantity

    for brand_name, products in brands.items():
        rows = [[
            "Image", "Name", "Sub Category", "Series", "SKU",
            "Stock", "UPC/EAN Code", "Price", "Margin", "Selling Price",
            "Quantity", "Total"
        ]]

        for idx, product in enumerate(products, start=2):
            margin = special_margins.get(
                str(product.get("_id")),
                customer.get("cf_margin", "40%")
            )
            try:
                margin_value = int(margin.replace("%", "")) / 100
            except:
                margin_value = 0.4

            rate = product.get("rate", 0)

            # Get image URL - check image_url first, then images array
            image_url = product.get("image_url", "")
            if not image_url:
                # If image_url doesn't exist, check for images array
                images = product.get("images", [])
                if images and len(images) > 0:
                    # Use the first image from the images array
                    image_url = images[0] if isinstance(images[0], str) else images[0].get("url", "")

            # Get quantity from cart if product exists in cart
            product_id = str(product.get("_id", ""))
            quantity = cart_quantities.get(product_id, "")

            rows.append([
                f'=IMAGE("{image_url}", 1)',
                product.get("name", ""),
                product.get("sub_category", ""),
                product.get("series", ""),
                product.get("cf_sku_code", ""),
                product.get("stock", ""),
                product.get("upc_code", ""),
                rate,
                margin,
                rate * margin_value,
                quantity,
                f"=K{idx}*H{idx}"
            ])

        brand_data[brand_name] = rows

    return brand_data
def create_format_requests(sheet_id: int, rows_count: int) -> List[Dict]:
    """Create formatting requests for a sheet"""
    return [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9}
                    }
                },
                "fields": "userEnteredFormat(textFormat,horizontalAlignment,verticalAlignment,backgroundColor)",
            }
        },
        {
            "updateBorders": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": rows_count,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12,
                },
                "top": {"style": "SOLID"},
                "bottom": {"style": "SOLID"},
                "left": {"style": "SOLID"},
                "right": {"style": "SOLID"},
                "innerHorizontal": {"style": "SOLID"},
                "innerVertical": {"style": "SOLID"},
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": 1,
                },
                "properties": {"pixelSize": 150},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 1,
                    "endIndex": 2,
                },
                "properties": {"pixelSize": 250},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 1,
                    "endIndex": rows_count,
                },
                "properties": {"pixelSize": 80},
                "fields": "pixelSize",
            }
        },
    ]

async def try_set_permissions(drive_service, spreadsheet_id: str, user_email: Optional[str] = None) -> Dict[str, str]:
    """Try different permission methods and return status"""
    permission_status = {}
    
    # Method 1: Anyone with link (most important)
    result, error = await safe_execute(
        "Setting 'anyone with link' permissions",
        lambda: drive_service.permissions().create(
            fileId=spreadsheet_id,
            body={"type": "anyone", "role": "writer"},
            sendNotificationEmail=False
        ).execute(),
        critical=False
    )
    permission_status["anyone_with_link"] = "success" if result else f"failed: {error}"
    
    # Method 2: Specific user if provided
    if user_email:
        result, error = await safe_execute(
            f"Setting user permissions for {user_email}",
            lambda: drive_service.permissions().create(
                fileId=spreadsheet_id,
                body={"type": "user", "role": "writer", "emailAddress": user_email},
                sendNotificationEmail=False
            ).execute(),
            critical=False
        )
        permission_status["user_specific"] = "success" if result else f"failed: {error}"
    
    # Method 3: Try to make it publicly viewable as fallback
    if permission_status["anyone_with_link"].startswith("failed"):
        result, error = await safe_execute(
            "Setting public view permissions",
            lambda: drive_service.permissions().create(
                fileId=spreadsheet_id,
                body={"type": "anyone", "role": "reader"},
                sendNotificationEmail=False
            ).execute(),
            critical=False
        )
        permission_status["public_readonly"] = "success" if result else f"failed: {error}"
    
    return permission_status

@router.get("/download_order_form")
async def download_order_form(customer_id: str, order_id: str, sort: str = "default", user_email: Optional[str] = None):
    """
    Create order form spreadsheet with robust error handling
    user_email: Optional email to grant specific access to
    """
    
    if not customer_id or not order_id:
        raise HTTPException(status_code=400, detail="Customer and Order IDs are required")

    # Check existing order
    order = db.orders.find_one({"_id": ObjectId(order_id)})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("spreadsheet_created", False):
        return {"google_sheet_url": order.get("spreadsheet_url", "")}

    # Track operations for detailed response
    operations_status = {}
    sheet_url = None
    spreadsheet_id = None
    
    try:
        print(f"Starting order form creation for order {order_id}")
        
        # Step 1: Get all required data (CRITICAL)
        print("📊 Fetching data...")
        
        customer_task = asyncio.create_task(
            asyncio.to_thread(db.customers.find_one, {"_id": ObjectId(customer_id)})
        )
        
        sort_stage = {"brand": 1, "rate": 1, "name": 1}
        if sort == "price_asc":
            sort_stage = {"rate": 1}
        elif sort == "price_desc":
            sort_stage = {"rate": -1}
        elif sort == "catalogue":
            sort_stage = {"catalogue_order": 1}
            
        products_task = asyncio.create_task(
            asyncio.to_thread(get_active_products, sort_stage)
        )
        
        special_margins_task = asyncio.create_task(
            asyncio.to_thread(
                lambda: {
                    str(sm["product_id"]): sm["margin"]
                    for sm in db.special_margins.find({"customer_id": ObjectId(customer_id)})
                }
            )
        )
        
        customer, products, special_margins = await asyncio.gather(
            customer_task, products_task, special_margins_task
        )
        
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
            
        brands = {}
        for product in products:
            brand = product.get("brand", "Unknown")
            brands.setdefault(brand, []).append(product)
            
        if not brands:
            raise HTTPException(status_code=404, detail="No active products found")

        operations_status["data_fetch"] = f"✓ Loaded {len(products)} products in {len(brands)} brands"

        # Get cart products to pre-fill quantities
        cart_products = order.get("products", [])
        cart_product_count = len(cart_products)

        # Step 2: Create spreadsheet (CRITICAL)
        print("📝 Creating spreadsheet...")

        sheets_service = get_sheets_service()
        drive_service = get_drive_service()

        spreadsheet_body = {
            "properties": {
                "title": f"Order Form - {customer.get('display_name', 'Customer')} - {order_id[:8]}"
            }
        }

        spreadsheet, error = await safe_execute(
            "Creating spreadsheet",
            lambda: sheets_service.spreadsheets().create(body=spreadsheet_body).execute(),
            critical=True
        )

        spreadsheet_id = spreadsheet["spreadsheetId"]
        sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        operations_status["spreadsheet_creation"] = "✓ Spreadsheet created successfully"

        # Update DB immediately after successful creation
        db.orders.update_one(
            {"_id": ObjectId(order_id)},
            {"$set": {"spreadsheet_created": True, "spreadsheet_url": sheet_url}},
        )
        operations_status["database_update"] = "✓ Order updated in database"

        # Step 3: Create sheets and add data (SEMI-CRITICAL)
        print("📋 Adding sheets and data...")

        brand_data = prepare_brand_data(brands, customer, special_margins, cart_products)

        # Normalize brand names to avoid case-sensitive duplicates
        # Google Sheets treats sheet names as case-insensitive
        normalized_brand_data = {}
        brand_name_mapping = {}  # Maps normalized name to original name

        for brand_name, data in brand_data.items():
            # Use the first occurrence of each case-insensitive brand name
            normalized_name = brand_name
            normalized_lower = brand_name.lower()

            if normalized_lower not in [k.lower() for k in normalized_brand_data.keys()]:
                normalized_brand_data[normalized_name] = data
                brand_name_mapping[normalized_name] = brand_name
            else:
                # Find the existing key with same case-insensitive name
                existing_key = next(k for k in normalized_brand_data.keys() if k.lower() == normalized_lower)
                # Merge the products from this duplicate brand into the existing one
                normalized_brand_data[existing_key].extend(data)
                print(f"⚠️ Merged duplicate brand '{brand_name}' into '{existing_key}'")

        brand_data = normalized_brand_data
        # Exclude "New Arrivals" from spreadsheet generation (it's a virtual brand for frontend display only)
        sorted_brands = sorted([brand for brand in brand_data.keys() if brand != "New Arrivals"])

        # Get existing sheet names to avoid duplicates
        existing_sheets = {sheet["properties"]["title"]: sheet["properties"]["sheetId"]
                          for sheet in spreadsheet.get("sheets", [])}

        # Create case-insensitive lookup for existing sheets
        existing_sheets_lower = {name.lower(): (name, sheet_id)
                                for name, sheet_id in existing_sheets.items()}

        print(f"🔍 DEBUG: Existing sheets: {list(existing_sheets.keys())}")
        print(f"🔍 DEBUG: Brands needed: {sorted_brands}")

        # Create sheets only for brands that don't exist yet (case-insensitive check)
        sheet_requests = []
        brands_to_create = [brand for brand in sorted_brands
                           if brand.lower() not in existing_sheets_lower]

        print(f"🔍 DEBUG: Brands to create: {brands_to_create}")
        print(f"🔍 DEBUG: Brands already exist: {[b for b in sorted_brands if b in existing_sheets]}")

        for brand in brands_to_create:
            sheet_requests.append({"addSheet": {"properties": {"title": brand}}})

        # Only delete Sheet1 if it exists
        if "Sheet1" in existing_sheets:
            sheet_requests.append({"deleteSheet": {"sheetId": existing_sheets["Sheet1"]}})

        print(f"🔍 DEBUG: Total sheet requests: {len(sheet_requests)}")
        
        # Only execute batch update if there are requests
        if sheet_requests:
            batch_response, error = await safe_execute(
                "Creating brand sheets",
                lambda: sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={"requests": sheet_requests}
                ).execute(),
                critical=False
            )
        else:
            batch_response = None
            error = None

        # Map brand names to sheet IDs
        brand_sheets = {}

        if batch_response or not sheet_requests:
            operations_status["sheets_creation"] = f"✓ Using {len(sorted_brands)} brand sheets ({len(brands_to_create)} new, {len(sorted_brands) - len(brands_to_create)} existing)"

            # Add existing sheets to the mapping (case-insensitive)
            for brand in sorted_brands:
                if brand.lower() in existing_sheets_lower:
                    _, sheet_id = existing_sheets_lower[brand.lower()]
                    brand_sheets[brand] = sheet_id

            # Add newly created sheets to the mapping
            if batch_response and "replies" in batch_response:
                reply_index = 0
                for brand in brands_to_create:
                    # Find the corresponding reply (skip deleteSheet replies)
                    while reply_index < len(batch_response["replies"]):
                        reply = batch_response["replies"][reply_index]
                        if "addSheet" in reply:
                            brand_sheets[brand] = reply["addSheet"]["properties"]["sheetId"]
                            reply_index += 1
                            break
                        reply_index += 1
            
            # Add data to sheets
            data_updates = []
            all_format_requests = []
            
            for brand_name in sorted_brands:
                sheet_id = brand_sheets[brand_name]
                rows = brand_data[brand_name]
                
                data_updates.append({
                    "range": f"{brand_name}!A1:L{len(rows)}",
                    "values": rows
                })
                
                format_requests = create_format_requests(sheet_id, len(rows))
                all_format_requests.extend(format_requests)
            
            # Add data
            data_result, error = await safe_execute(
                "Adding data to sheets",
                lambda: sheets_service.spreadsheets().values().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "valueInputOption": "USER_ENTERED",
                        "data": data_updates,
                    },
                ).execute(),
                critical=False
            )
            
            if data_result:
                operations_status["data_population"] = "✓ Data added to all sheets"
            else:
                operations_status["data_population"] = f"✗ Data population failed: {error}"
            
            # Format sheets
            format_result, error = await safe_execute(
                "Formatting sheets",
                lambda: sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={"requests": all_format_requests},
                ).execute(),
                critical=False
            )
            
            if format_result:
                operations_status["formatting"] = "✓ Sheets formatted successfully"
            else:
                operations_status["formatting"] = f"✗ Formatting failed: {error}"
        else:
            operations_status["sheets_creation"] = f"✗ Failed to create sheets: {error}"
        
        # Step 4: Set permissions (NON-CRITICAL)
        print("🔐 Setting permissions...")
        
        permission_status = await try_set_permissions(drive_service, spreadsheet_id, user_email)
        operations_status["permissions"] = permission_status
        
        # Determine if we have any working access method
        has_access = any(status.startswith("success") for status in permission_status.values())
        
        # Build message based on whether cart had products
        if cart_product_count > 0:
            message = f"Order form created with {len(sorted_brands)} brand sheets. {cart_product_count} product quantities pre-filled from cart."
        else:
            message = f"Order form created with {len(sorted_brands)} brand sheets"

        response_data = {
            "google_sheet_url": sheet_url,
            "spreadsheet_id": spreadsheet_id,
            "status": "success",
            "operations": operations_status,
            "access_methods": permission_status,
            "message": message,
            "cart_products_added": cart_product_count
        }
        
        if not has_access:
            response_data["warning"] = "Sheet created but permission setting failed. You may need to manually share the sheet."
            response_data["manual_access_instructions"] = [
                f"1. Open {sheet_url}",
                "2. Click 'Share' button in top right",
                "3. Change 'Restricted' to 'Anyone with the link'",
                "4. Set permission to 'Editor'",
                "5. Click 'Done'"
            ]
        
        print("✅ Order form creation completed")
        return response_data
        
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except Exception as e:
        # Log the full error for debugging
        print(f"❌ Unexpected error: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # If we have a sheet URL, return it with error details
        if sheet_url:
            return {
                "google_sheet_url": sheet_url,
                "spreadsheet_id": spreadsheet_id,
                "status": "partial_success",
                "error": str(e),
                "operations": operations_status,
                "message": "Sheet was created but some operations failed. Check the sheet manually."
            }
        else:
            raise HTTPException(
                status_code=500, 
                detail=f"Failed to create order form: {str(e)}"
            )
@router.get("/download_order_xlsx")
async def download_order_xlsx(customer_id: str, order_id: str, sort: str = "default"):
    """Generate and return an XLSX version of the order form with embedded images"""
    if not customer_id or not order_id:
        raise HTTPException(status_code=400, detail="Customer and Order IDs are required")

    order = db.orders.find_one({"_id": ObjectId(order_id)})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    sort_stage = {"brand": 1, "rate": 1, "name": 1}
    if sort == "price_asc":
        sort_stage = {"rate": 1}
    elif sort == "price_desc":
        sort_stage = {"rate": -1}
    elif sort == "catalogue":
        sort_stage = {"catalogue_order": 1}

    customer, products, special_margins = await asyncio.gather(
        asyncio.to_thread(db.customers.find_one, {"_id": ObjectId(customer_id)}),
        asyncio.to_thread(get_active_products, sort_stage),
        asyncio.to_thread(lambda: {
            str(sm["product_id"]): sm["margin"]
            for sm in db.special_margins.find({"customer_id": ObjectId(customer_id)})
        }),
    )

    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    brands: Dict[str, list] = {}
    for product in products:
        brands.setdefault(product.get("brand", "Unknown"), []).append(product)
    if not brands:
        raise HTTPException(status_code=404, detail="No active products found")

    cart_products = order.get("products", [])
    brand_data = prepare_brand_data(brands, customer, special_margins, cart_products)
    sorted_brands = sorted([b for b in brand_data if b != "New Arrivals"])

    # Collect all image URLs in order
    image_urls: list[str] = []
    for brand_name in sorted_brands:
        for row in brand_data[brand_name][1:]:  # skip header
            formula = row[0]
            if isinstance(formula, str) and "IMAGE(" in formula.upper():
                m = re.search(r'=IMAGE\("([^"]+)"', formula, re.IGNORECASE)
                image_urls.append(m.group(1) if m else "")
            else:
                image_urls.append("")

    # Download all images in parallel using a single shared client
    semaphore = asyncio.Semaphore(30)

    async def fetch_image(client: httpx.AsyncClient, url: str):
        if not url:
            return None
        async with semaphore:
            try:
                r = await client.get(url)
                return r.content if r.status_code == 200 else None
            except Exception:
                return None

    async with httpx.AsyncClient(timeout=15.0, limits=httpx.Limits(max_connections=50, max_keepalive_connections=30)) as client:
        image_contents = await asyncio.gather(*[fetch_image(client, u) for u in image_urls])

    # Resize images to thumbnail size in parallel threads (avoids embedding huge originals)
    from PIL import Image as PILImage

    def resize_image(img_bytes):
        if not img_bytes:
            return None
        try:
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            pil_img.thumbnail((80, 80), PILImage.LANCZOS)
            out = io.BytesIO()
            pil_img.save(out, format="JPEG", quality=80, optimize=True)
            out.seek(0)
            return out.read()
        except Exception:
            return None

    image_contents = await asyncio.gather(*[
        asyncio.to_thread(resize_image, b) for b in image_contents
    ])

    # Build XLSX in a thread (CPU-bound, must not block the event loop)
    def build_xlsx():
        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        col_widths = [15, 35, 18, 18, 15, 10, 18, 10, 10, 14, 12, 12]

        img_idx = 0
        for brand_name in sorted_brands:
            rows = brand_data[brand_name]
            ws = wb.create_sheet(title=brand_name[:31])

            for c, header in enumerate(rows[0], start=1):
                cell = ws.cell(row=1, column=c, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[1].height = 20

            for c, w in enumerate(col_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

            for r_offset, row_data in enumerate(rows[1:], start=2):
                ws.row_dimensions[r_offset].height = 65

                img_bytes = image_contents[img_idx]
                img_idx += 1
                if img_bytes:
                    try:
                        xl_img = XLImage(io.BytesIO(img_bytes))
                        xl_img.width = 70
                        xl_img.height = 70
                        ws.add_image(xl_img, f"A{r_offset}")
                    except Exception:
                        ws.cell(row=r_offset, column=1, value="[img]").alignment = center

                for c_offset, value in enumerate(row_data[1:], start=2):
                    if c_offset == 12:
                        qty = row_data[10] if row_data[10] else 0
                        price = row_data[7] if row_data[7] else 0
                        value = qty * price if qty and price else ""
                    cell = ws.cell(row=r_offset, column=c_offset, value=value)
                    cell.alignment = center
                    cell.border = border

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    stream = await asyncio.to_thread(build_xlsx)

    customer_name = customer.get("display_name", "Customer").replace(" ", "_")
    filename = f"Order_Form_{customer_name}_{order_id[:8]}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def read_sheet_with_retry(service, spreadsheet_id: str, sheet_title: str, max_retries: int = 3,
                               initial_delay: float = 1.0, backoff_factor: float = 2.0) -> Tuple[Optional[dict], Optional[str]]:
    """
    Read a single sheet with exponential backoff retry logic
    """
    last_error = None
    delay = initial_delay
    
    for attempt in range(max_retries):
        try:
            print(f"Reading sheet '{sheet_title}' (attempt {attempt + 1}/{max_retries})")
            
            result = await safe_execute(
                f"Reading sheet {sheet_title}",
                lambda: service.spreadsheets().values().get(
                    spreadsheetId=spreadsheet_id, 
                    range=f"{sheet_title}!A1:L"
                ).execute(),
                critical=False
            )
            
            data, error = result
            if not error:
                print(f"✓ Successfully read sheet '{sheet_title}'")
                return data, None
            
            last_error = error
            print(f"✗ Attempt {attempt + 1} failed for sheet '{sheet_title}': {error}")
            
        except Exception as e:
            last_error = str(e)
            print(f"✗ Attempt {attempt + 1} failed for sheet '{sheet_title}': {e}")
        
        # Wait before retry (except on last attempt)
        if attempt < max_retries - 1:
            print(f"Waiting {delay:.1f}s before retry...")
            await asyncio.sleep(delay)
            delay *= backoff_factor
    
    return None, f"Failed after {max_retries} attempts: {last_error}"

async def read_sheets_batch(service, spreadsheet_id: str, sheet_titles: List[str], 
                           batch_size: int = 3, max_concurrent: int = 2) -> List[Tuple[str, Optional[dict], Optional[str]]]:
    """
    Read sheets in smaller batches to avoid overwhelming the API and network
    """
    results = []
    
    # Process sheets in batches
    for i in range(0, len(sheet_titles), batch_size):
        batch = sheet_titles[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1}: {batch}")
        
        # Limit concurrent requests within each batch
        semaphore = asyncio.Semaphore(max_concurrent)
        
        async def read_with_semaphore(title):
            async with semaphore:
                data, error = await read_sheet_with_retry(service, spreadsheet_id, title)
                return title, data, error
        
        batch_tasks = [read_with_semaphore(title) for title in batch]
        batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
        
        for result in batch_results:
            if isinstance(result, Exception):
                # Handle any unexpected exceptions
                results.append((f"unknown_sheet", None, str(result)))
            else:
                results.append(result)
        
        # Small delay between batches to be gentle on the API
        if i + batch_size < len(sheet_titles):
            await asyncio.sleep(0.5)
    
    return results
def process_sheet_data_fast(values: List[List], sheet_name: str, products_cache: Dict) -> tuple[List[Dict], List[str]]:
    """
    Optimized sheet data processing with pre-cached products
    """
    updated_products = []
    processing_errors = []
    
    if len(values) < 2:
        return updated_products, [f"Sheet {sheet_name}: No data rows found"]
        
    headers = values[0]
    
    # Find required columns
    try:
        sku_idx = headers.index("SKU")
        quantity_idx = headers.index("Quantity")
        name_idx = headers.index("Name") if "Name" in headers else -1
        margin_idx = headers.index("Margin") if "Margin" in headers else -1
    except ValueError as e:
        return updated_products, [f"Sheet {sheet_name}: Missing required columns - {str(e)}"]
    
    # Process rows in batches for better performance
    valid_rows = []
    for row_num, row in enumerate(values[1:], start=2):
        # Quick validation
        if (len(row) <= max(quantity_idx, sku_idx) or 
            not row[quantity_idx].strip() or 
            not row[sku_idx].strip()):
            continue
        
        try:
            quantity = int(float(row[quantity_idx].strip()))
            if quantity <= 0:
                continue
                
            valid_rows.append((row_num, row, quantity))
        except ValueError:
            processing_errors.append(f"Sheet {sheet_name} row {row_num}: Invalid quantity")
    
    # Process valid rows with cached products
    for row_num, row, quantity in valid_rows:
        try:
            product_sku = row[sku_idx].strip()
            
            # Use cached product lookup instead of database query
            product = products_cache.get(product_sku)
            if not product:
                processing_errors.append(f"Sheet {sheet_name} row {row_num}: Product not found for SKU '{product_sku}'")
                continue
            
            # Extract additional data safely
            product_name = row[name_idx] if name_idx >= 0 and len(row) > name_idx else ""
            margin = row[margin_idx] if margin_idx >= 0 and len(row) > margin_idx else ""
            
            # Create product entry
            updated_products.append({
                "product_id": ObjectId(product["_id"]),
                "tax_percentage": product.get("item_tax_preferences", [{}])[0].get("tax_percentage", 0),
                "brand": product.get("brand", "Unknown"),
                "product_code": product_sku,
                "quantity": quantity,
                "name": product_name or product.get("name", ""),
                "image_url": product.get("image_url"),
                "margin": margin,
                "price": float(product.get("rate", 0)),
                "added_by": "sales_person",
            })
            
        except Exception as e:
            processing_errors.append(f"Sheet {sheet_name} row {row_num}: Processing error - {str(e)}")
            continue
    
    return updated_products, processing_errors

async def fallback_individual_read(order_id: str, spreadsheet_id: str, sheet_titles: List[str], products_cache: Dict, service):
    """
    Fallback method: Read sheets individually if batch read fails
    """
    print("🔄 Using fallback individual read method")
    start_time = time.time()
    
    all_updated_products = []
    all_processing_errors = []
    successful_sheets = 0
    
    for sheet_title in sheet_titles:
        try:
            # Read individual sheet
            escaped_title = f"'{sheet_title}'" if " " in sheet_title or "'" in sheet_title else sheet_title
            sheet_data, error = await robust_api_call(
                f"Reading {sheet_title}",
                lambda: service.spreadsheets().values().get(
                    spreadsheetId=spreadsheet_id,
                    range=f"{escaped_title}!A1:L1000"
                ).execute()
            )
            
            if not sheet_data:
                all_processing_errors.append(f"Sheet {sheet_title}: {error}")
                continue
            
            values = sheet_data.get("values", [])
            if values:
                sheet_products, sheet_errors = process_sheet_data_fast(
                    values, sheet_title, products_cache
                )
                
                all_updated_products.extend(sheet_products)
                all_processing_errors.extend(sheet_errors)
                
                if sheet_products:
                    successful_sheets += 1
                    print(f"✅ {sheet_title}: {len(sheet_products)} products")
            
        except Exception as e:
            all_processing_errors.append(f"Sheet {sheet_title}: Exception - {str(e)}")
    
    if not all_updated_products:
        error_summary = "; ".join(all_processing_errors[:5])
        raise HTTPException(
            status_code=400, 
            detail=f"Fallback method failed - No valid products found. Errors: {error_summary}"
        )
    
    # Update database
    db.orders.update_one(
        {"_id": ObjectId(order_id)},
        {
            "$set": {
                "products": all_updated_products, 
                "updated_from_sheet": True,
                "last_sheet_update": datetime.now()
            }
        },
    )
    
    processing_time = round(time.time() - start_time, 2)
    
    return {
        "message": "Order updated successfully (fallback method)",
        "products_count": len(all_updated_products),
        "sheets_processed": len(sheet_titles),
        "successful_sheets": successful_sheets,
        "processing_time_seconds": processing_time,
        "method": "fallback_individual",
        "updated_products": serialize_mongo_document(all_updated_products),
        "warnings": all_processing_errors[:5] if all_processing_errors else []
    }

@router.get("/update_cart")
async def update_order_from_sheet(order_id: str):
    """
    Fast and reliable update order from spreadsheet using proper batch operations
    """
    
    # Validate order
    order = db.orders.find_one({"_id": ObjectId(order_id)})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    spreadsheet_url = order.get("spreadsheet_url", "")
    if not spreadsheet_url:
        raise HTTPException(status_code=400, detail="No spreadsheet associated with order")

    # Extract spreadsheet ID
    try:
        spreadsheet_id = spreadsheet_url.split("/d/")[1].split("/")[0]
    except IndexError:
        raise HTTPException(status_code=400, detail="Invalid spreadsheet URL")

    print(f"🚀 Starting fast cart update for order {order_id}")
    start_time = time.time()
    
    try:
        service = create_fresh_sheets_service()
        
        # Step 1: Get spreadsheet info and product cache in parallel
        spreadsheet_task = robust_api_call(
            "Reading spreadsheet info",
            lambda: service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        )
        
        # Pre-load all active products into memory
        products_task = asyncio.create_task(
            asyncio.to_thread(
                lambda: {
                    product["cf_sku_code"]: product 
                    for product in db.products.find({"status": "active"})
                    if product.get("cf_sku_code")
                }
            )
        )
        
        spreadsheet_result, products_cache = await asyncio.gather(
            spreadsheet_task, products_task
        )
        
        spreadsheet_info, error = spreadsheet_result
        if not spreadsheet_info:
            raise HTTPException(status_code=500, detail=f"Cannot access spreadsheet: {error}")
        
        sheet_titles = [sheet["properties"]["title"] for sheet in spreadsheet_info["sheets"]]
        print(f"📊 Found {len(sheet_titles)} sheets, {len(products_cache)} products cached")
        
        # Step 2: Create proper ranges for batch reading
        ranges = []
        for title in sheet_titles:
            # Properly escape sheet names with spaces or special characters
            escaped_title = f"'{title}'" if " " in title or "'" in title else title
            ranges.append(f"{escaped_title}!A1:L1000")
        
        print(f"📋 Reading ranges: {ranges}")
        
        # Step 3: Batch read all sheets
        batch_data, error = await robust_api_call(
            "Batch reading all sheets",
            lambda: service.spreadsheets().values().batchGet(
                spreadsheetId=spreadsheet_id,
                ranges=ranges,
                majorDimension="ROWS"
            ).execute()
        )
        
        if not batch_data:
            raise HTTPException(status_code=500, detail=f"Failed to read sheets: {error}")
        
        print(f"📊 Batch data keys: {list(batch_data.keys())}")
        
        # Debug: Check what we actually got
        value_ranges = batch_data.get("valueRanges", [])
        print(f"📋 Got {len(value_ranges)} value ranges")
        
        if not value_ranges:
            # Fallback: Try reading sheets individually if batch failed
            print("⚠️ Batch read returned no data, falling back to individual reads")
            return await fallback_individual_read(order_id, spreadsheet_id, sheet_titles, products_cache, service)
        
        # Step 4: Process all sheets data efficiently
        all_updated_products = []
        all_processing_errors = []
        successful_sheets = 0
        
        for i, sheet_data in enumerate(value_ranges):
            sheet_title = sheet_titles[i] if i < len(sheet_titles) else f"Sheet_{i+1}"
            values = sheet_data.get("values", [])
            
            print(f"📖 Processing {sheet_title}: {len(values)} rows")
            
            if not values:
                all_processing_errors.append(f"Sheet {sheet_title}: No data found")
                continue
            
            # Process this sheet's data with pre-cached products
            sheet_products, sheet_errors = process_sheet_data_fast(
                values, sheet_title, products_cache
            )
            
            all_updated_products.extend(sheet_products)
            all_processing_errors.extend(sheet_errors)
            
            if sheet_products:
                successful_sheets += 1
                print(f"✅ {sheet_title}: Found {len(sheet_products)} products")
            else:
                print(f"⚠️ {sheet_title}: No valid products found")
        
        # Step 5: Validate results
        if not all_updated_products:
            error_summary = "; ".join(all_processing_errors[:3])
            print(f"❌ No products found. Errors: {error_summary}")
            print(f"📊 Debug info - Sheets processed: {len(value_ranges)}, Products cache size: {len(products_cache)}")
            
            # Additional debugging
            debug_info = {
                "sheets_found": len(sheet_titles),
                "value_ranges_returned": len(value_ranges),
                "products_cache_size": len(products_cache),
                "sample_sheet_data": {}
            }
            
            # Show sample data from first sheet for debugging
            if value_ranges and len(value_ranges) > 0:
                sample_values = value_ranges[0].get("values", [])
                debug_info["sample_sheet_data"] = {
                    "first_sheet_rows": len(sample_values),
                    "first_few_rows": sample_values[:3] if sample_values else []
                }
            
            raise HTTPException(
                status_code=400, 
                detail=f"No valid products found. Errors: {error_summary}. Debug: {debug_info}"
            )
        
        processing_time = round(time.time() - start_time, 2)
        print(f"⚡ Processed {len(all_updated_products)} products in {processing_time}s")
        
        # Step 6: Update database
        db.orders.update_one(
            {"_id": ObjectId(order_id)},
            {
                "$set": {
                    "products": all_updated_products, 
                    "updated_from_sheet": True,
                    "last_sheet_update": datetime.now()
                }
            },
        )
        
        # Step 7: Prepare response
        response = {
            "message": "Order updated successfully",
            "products_count": len(all_updated_products),
            "sheets_processed": len(sheet_titles),
            "successful_sheets": successful_sheets,
            "processing_time_seconds": processing_time,
            "updated_products": serialize_mongo_document(all_updated_products),
        }
        
        if all_processing_errors:
            response["warnings"] = all_processing_errors[:5]  # Limit warnings
            response["total_warnings"] = len(all_processing_errors)
        
        print(f"🎉 Cart update completed in {processing_time}s")
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"💥 Error in update_order_from_sheet: {e}")
        raise HTTPException(status_code=500, detail=f"Error reading spreadsheet: {str(e)}")

# Get invoices linked to an order via its estimate's invoice_ids
@router.get("/{order_id}/invoices")
def get_order_invoices(order_id: str):
    """
    Find the estimate for this order by estimate_number, then return the
    invoices listed in the estimate's invoice_ids array.
    """
    order = get_order(order_id, orders_collection)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    estimate_number = order.get("estimate_number", "")
    if not estimate_number:
        return []

    estimate = db.estimates.find_one({"estimate_number": estimate_number})
    if not estimate:
        return []

    invoice_ids = estimate.get("invoice_ids", [])
    if not invoice_ids:
        return []

    # Build a lookup map from DB for any invoices that have been synced
    db_invoices = {
        doc["invoice_id"]: doc
        for doc in db.invoices.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"invoice_id": 1, "invoice_number": 1, "invoice_url": 1, "total": 1, "status": 1},
        )
    }

    # Always return one entry per invoice_id from the estimate; enrich with DB metadata when available
    result = []
    for zoho_id in invoice_ids:
        db_doc = db_invoices.get(zoho_id, {})
        result.append({
            "zoho_invoice_id": zoho_id,
            "invoice_number": db_doc.get("invoice_number", zoho_id),
            "invoice_url": db_doc.get("invoice_url", ""),
            "total": db_doc.get("total", None),
            "status": db_doc.get("status", ""),
        })
    return result


# Get an order by ID
@router.get("/my-performance")
def get_my_performance(user_id: str):
    """
    Return order performance stats for the given user for this month vs. last month.
    Includes order count, total value, and breakdown by status.
    """
    import datetime as dt
    from dateutil.relativedelta import relativedelta

    try:
        user_oid = ObjectId(user_id)
    except Exception:
        user_oid = user_id

    now = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None)
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = this_month_start - relativedelta(months=1)
    last_month_end = this_month_start

    def bucket_stats(start, end):
        pipeline = [
            {"$match": {
                "created_by": user_oid,
                "status": {"$ne": "deleted"},
                "created_at": {"$gte": start, "$lt": end},
            }},
            {"$group": {
                "_id": "$status",
                "count": {"$sum": 1},
                "total_value": {"$sum": {"$toDouble": {"$ifNull": ["$total_amount", 0]}}},
            }},
        ]
        rows = list(orders_collection.aggregate(pipeline))
        total_count = sum(r["count"] for r in rows)
        total_value = sum(r["total_value"] for r in rows)
        by_status = {r["_id"]: {"count": r["count"], "value": r["total_value"]} for r in rows}
        return {"total_count": total_count, "total_value": total_value, "by_status": by_status}

    this_month = bucket_stats(this_month_start, now)
    last_month = bucket_stats(last_month_start, last_month_end)

    def pct_change(current, previous):
        if previous == 0:
            return None
        return round((current - previous) / previous * 100, 1)

    return {
        "this_month": this_month,
        "last_month": last_month,
        "count_change_pct": pct_change(this_month["total_count"], last_month["total_count"]),
        "value_change_pct": pct_change(this_month["total_value"], last_month["total_value"]),
        "period": {
            "this_month_label": this_month_start.strftime("%B %Y"),
            "last_month_label": last_month_start.strftime("%B %Y"),
        },
    }


@router.get("/{order_id}")
def read_order(order_id: str):
    """
    Retrieve an order by its ID.

    Embeds the customer's current margin (`customer_margin`), GST type
    (`gst_type`) and special margins (`special_margins`: product_id -> margin)
    so unauthenticated shared-link visitors price products with exactly the
    same margins as the order creator.
    """
    order = get_order(order_id, orders_collection)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("customer_id"):
        try:
            customer = customers_collection.find_one(
                {"_id": ObjectId(order["customer_id"])},
                {"cf_margin": 1, "cf_in_ex": 1},
            )
            if customer:
                order["customer_margin"] = customer.get("cf_margin") or "40%"
                if not order.get("gst_type"):
                    order["gst_type"] = customer.get("cf_in_ex") or "Exclusive"
            order["special_margins"] = {
                str(doc["product_id"]): doc["margin"]
                for doc in db["special_margins"].find(
                    {"customer_id": ObjectId(order["customer_id"])},
                    {"product_id": 1, "margin": 1},
                )
                if doc.get("product_id") and doc.get("margin")
            }
        except Exception as e:
            # Pricing context is an enhancement — never fail the order fetch over it
            print(f"Error embedding pricing context on order {order_id}: {e}")

    # Embed live estimate statuses from the estimates collection (kept current by webhook)
    est_num = order.get("estimate_number")
    po_est_num = order.get("pre_order_estimate_number")
    if est_num or po_est_num:
        nums_to_fetch = [n for n in [est_num, po_est_num] if n]
        est_docs = {
            doc["estimate_number"]: doc.get("status", "")
            for doc in db.estimates.find(
                {"estimate_number": {"$in": nums_to_fetch}},
                {"estimate_number": 1, "status": 1},
            )
        }
        if est_num:
            order["estimate_status"] = est_docs.get(est_num, "")
        if po_est_num:
            order["pre_order_estimate_status"] = est_docs.get(po_est_num, "")

    return order


# Get all orders
@router.get("")
def read_all_orders(role: str = "salesperson", created_by: str = "", status: str = ""):
    """
    Retrieve all orders.
    If role is 'admin', return all orders.
    If role is 'salesperson', return only orders created by the specified user.
    """
    orders = get_all_orders(
        role, created_by, status, orders_collection, users_collection
    )
    return orders


# Update an order
@router.put("/{order_id}")
def update_existing_order(order_id: str, order_update: dict):
    """
    Update an existing order with raw dictionary data.
    """
    update_order(order_id, order_update, orders_collection, customers_collection)
    updated_order = get_order(order_id, orders_collection)
    if not updated_order:
        raise HTTPException(status_code=404, detail="Order not found")
    return updated_order


# Delete an order
@router.delete("/clear/{user_id}")
def clear_existing_order(user_id: str):
    """
    Deletes all orders by a given user who has created it if there is no customer information
    """
    clear_empty_orders(user_id, orders_collection)
    return {"detail": "Orders deleted successfully"}


@router.delete("/{order_id}")
def delete_existing_order(order_id: str, deleted_by: str):
    """
    Deletes all orders by a given user who has created it if there is no customer information
    """
    try:
        delete_order(order_id, deleted_by, orders_collection)
        return {"detail": "Orders deleted successfully"}
    except Exception as e:
        raise e


# Update an order
@router.put("/clear/{order_id}")
def clear_order_cart(order_id: str):
    """
    Update an existing order with raw dictionary data.
    """
    updated_order = clear_cart(order_id, orders_collection)
    return updated_order


# Finalise an order (Create Estimate)
@router.post("/finalise")
async def finalise(order_dict: dict, request: Request, background_tasks: BackgroundTasks):
    """
    finalise an existing order
    """
    order_id = order_dict.get("order_id")
    status = str(order_dict.get("status")).lower()
    create_stock = order_dict.get("create_stock", True)
    create_pre_order = order_dict.get("create_pre_order", True)
    try:
        # Perform order validation
        validate_order(order_id)
    except HTTPException as e:
        # Return validation error message if validation fails
        return {"status": "error", "message": e.detail}
    order = db.orders.find_one({"_id": ObjectId(order_id)})
    estimate_created = order.get("estimate_created", False)
    estimate_id = order.get("estimate_id", "")
    shipping_address_id = order.get("shipping_address", {}).get("address_id", "")
    billing_address_id = order.get("billing_address", {}).get("address_id", "")
    customer = db.customers.find_one({"_id": ObjectId(order.get("customer_id"))})
    state_str = str(order.get("shipping_address", {}).get("state", ""))
    place_of_supply = STATE_CODES.get(state_str.title())
    gst_type = order.get("gst_type", "")
    products = order.get("products", [])
    total_amount = order.get("total_amount")
    created_by = order.get("created_by")
    user = users_collection.find_one({"_id": ObjectId(created_by)})
    reference_number = order.get("reference_number", "")

    # Log finalize activity for customer accounts
    if user and user.get("customer_id"):
        from .customer_activity import log_activity, extract_client_info
        ip, ua = extract_client_info(request)
        customer_name = (
            user.get("contact_name")
            or f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
        )
        background_tasks.add_task(
            log_activity,
            action="finalize_order",
            category="orders",
            user_id=str(user["_id"]),
            customer_id=user.get("customer_id"),
            customer_name=customer_name,
            email=user.get("email"),
            metadata={"order_id": order_id, "status": status},
            ip_address=ip,
            user_agent=ua,
        )
    # Fetch SPecial Margins
    customer_id = order.get("customer_id")
    special_margins_cursor = db.special_margins.find(
        {"customer_id": ObjectId(customer_id)}
    )
    special_margin_dict = {
        str(sm["product_id"]): sm["margin"] for sm in special_margins_cursor
    }

    # Look up authoritative pre_order flag from products collection (not the saved order field,
    # which may be missing on orders saved before this field was persisted).
    all_product_ids = [ObjectId(p["product_id"]) for p in products if p.get("product_id")]
    product_docs_map = {
        str(doc["_id"]): doc
        for doc in db.products.find({"_id": {"$in": all_product_ids}})
    }

    # Split products into in-stock and pre-order.
    # Split products (pre_order=True in DB but also have physical stock) can carry
    # two quantities: `quantity` (stock portion) and `pre_order_quantity` (pre-order portion).
    in_stock_products = []
    pre_order_products_list = []
    for p in products:
        doc = product_docs_map.get(str(p.get("product_id")), {})
        qty = int(p.get("quantity") or 0)
        pre_order_qty = int(p.get("pre_order_quantity") or 0)
        db_is_pre_order = doc.get("pre_order", False)
        is_split = db_is_pre_order and (doc.get("stock") or 0) > 0

        # Mirror the frontend split logic exactly (Review.tsx).
        if is_split:
            # Split product: stock portion (`quantity`) -> in-stock estimate,
            # pre-order portion (`pre_order_quantity`) -> pre-order estimate.
            # These are two independent quantities on the same product.
            if qty > 0:
                in_stock_products.append(p)
            if pre_order_qty > 0:
                pre_order_products_list.append({**p, "quantity": pre_order_qty})
        elif db_is_pre_order:
            # Pure pre-order product (no live stock): the single `quantity`
            # field is the source of truth and goes entirely to the pre-order
            # estimate. `pre_order_quantity` is ignored here — a non-split
            # product can carry a STALE pre_order_quantity (e.g. it used to be
            # split before its stock dropped to <=0), and adding it would
            # double-count the line in the pre-order estimate.
            if qty > 0:
                pre_order_products_list.append(p)
        else:
            # In-stock product.
            if qty > 0:
                in_stock_products.append(p)

    pre_order_estimate_created = order.get("pre_order_estimate_created", False)
    pre_order_estimate_id = order.get("pre_order_estimate_id", "")

    def _make_line_items(prod_list, is_pre_order=False):
        items = []
        for idx, product in enumerate(prod_list):
            item = product_docs_map.get(str(product.get("product_id")))
            if item is None:
                continue
            product_id_str = str(product.get("product_id"))
            special_margin = special_margin_dict.get(
                product_id_str, customer.get("cf_margin", "40%")
            )
            discount_value = special_margin
            if not discount_value.endswith("%"):
                discount_value = f"{discount_value}%"
            tax_prefs = item.get("item_tax_preferences") or []
            description = (
                f"PRE-ORDER | SOH:{item.get('stock')}"
                if is_pre_order
                else f"SOH:{item.get('stock')}"
            )
            obj = {
                "item_order": idx + 1,
                "item_id": item.get("item_id"),
                "rate": item.get("rate"),
                "name": item.get("name"),
                "description": description,
                "quantity": product.get("quantity"),
                "discount": discount_value,
                "tax_id": (
                    tax_prefs[1].get("tax_id", 0)
                    if (place_of_supply == "MH" or place_of_supply == "") and len(tax_prefs) > 1
                    else tax_prefs[0].get("tax_id", 0) if tax_prefs else 0
                ),
                "tags": [],
                "tax_exemption_code": "",
                "item_custom_fields": [
                    {"label": "Manufacturer Code", "value": item.get("cf_item_code")},
                    {"label": "SKU Code", "value": item.get("cf_sku_code")},
                ],
                "hsn_or_sac": item.get("hsn_or_sac"),
                "gst_treatment_code": "",
                "unit": "pcs",
                "unit_conversion_id": "",
            }
            items.append(obj)
        return items

    stock_line_items = _make_line_items(in_stock_products)
    pre_order_line_items = _make_line_items(pre_order_products_list, is_pre_order=True)

    headers = {"Authorization": f"Zoho-oauthtoken {get_access_token('books')}"}
    message = ""
    stock_estimate_data: dict = {}
    preorder_estimate_data: dict = {}

    need_new_stock = create_stock and not estimate_created and bool(stock_line_items)
    need_new_preorder = create_pre_order and not pre_order_estimate_created and bool(pre_order_line_items)
    prefix = ""; num_width = 4; fy_str = ""; counter_id = ""
    errors: list = []

    async with httpx.AsyncClient(timeout=timeout) as client:
        # Fetch Zoho list ONCE if any new estimate needs creating
        if need_new_stock or need_new_preorder:
            now = datetime.now()
            fy_start = now.year if now.month >= 4 else now.year - 1
            fy_str = f"{str(fy_start)[-2:]}-{str(fy_start + 1)[-2:]}"
            y = await client.get(
                url=ESTIMATE_URL.format(org_id=org_id)
                + "&filter_by=Status.All&per_page=200&sort_column=estimate_number&sort_order=D",
                headers=headers,
            )
            if y.status_code != 200:
                return {"status": "error", "message": f"{y.json().get('message','')}"}
            all_estimates = y.json().get("estimates", [])
            fy_estimates = [
                e for e in all_estimates if f"/{fy_str}/" in e.get("estimate_number", "")
            ]
            if fy_estimates:
                last_parts = str(fy_estimates[0]["estimate_number"]).split("/")
                last_num = int(last_parts[-1]); num_width = len(last_parts[-1]); prefix = last_parts[0]
            else:
                last_parts = str(all_estimates[0]["estimate_number"]).split("/") if all_estimates else ["EST"]
                last_num = 0; num_width = len(last_parts[-1]) if len(last_parts) > 1 else 4; prefix = last_parts[0]
            counter_id = f"estimate_counter_{fy_str}"
            db.counters.update_one({"_id": counter_id}, {"$max": {"seq": last_num}}, upsert=True)

        def _base_payload(line_items_list: list, notes_text: str) -> dict:
            return {
                "location_id": "3220178000143298047",
                "contact_persons": [],
                "customer_id": customer.get("contact_id"),
                "date": datetime.now().strftime("%Y-%m-%d"),
                "expiry_date": "",
                "notes": notes_text,
                "terms": terms,
                "line_items": line_items_list,
                "custom_fields": [],
                "is_inclusive_tax": False if gst_type == "Exclusive" else True,
                "is_discount_before_tax": "",
                "discount": 0,
                "discount_type": "item_level",
                "adjustment": "",
                "adjustment_description": "Adjustment",
                "tax_exemption_code": "",
                "tax_authority_name": "",
                "pricebook_id": "",
                "salesperson_id": user.get("salesperson_id", ""),
                "payment_options": {"payment_gateways": []},
                "documents": [],
                "mail_attachments": [],
                "billing_address_id": billing_address_id,
                "shipping_address_id": shipping_address_id,
                "dispatch_from_address_id": "3220178000177830244",
                "project_id": "",
                "gst_treatment": customer.get("gst_treatment"),
                "gst_no": customer.get("gst_no", ""),
                "place_of_supply": place_of_supply,
                "is_tcs_amount_in_percent": True,
                "client_computation": {"total": total_amount or 0},
                "reference_number": reference_number,
            }

        # ── In-stock estimate ──
        if stock_line_items and create_stock:
            if not estimate_created:
                counter = db.counters.find_one_and_update(
                    {"_id": counter_id}, {"$inc": {"seq": 1}}, return_document=True
                )
                new_est_num = f"{prefix}/{fy_str}/{str(counter['seq']).zfill(num_width)}"
                payload = {"estimate_number": new_est_num, **_base_payload(stock_line_items, "Looking forward for your business.")}
                resp = await client.post(
                    url=ESTIMATE_URL.format(org_id=org_id) + "&ignore_auto_number_generation=true",
                    headers=headers,
                    json=payload,
                )
                print(resp.json())
                rj = resp.json()
                if resp.status_code != 201 or rj.get("code", 0) != 0:
                    errors.append(f"In-stock estimate: {rj.get('message', 'Unknown error')}")
                else:
                    stock_estimate_data = rj["estimate"]
                    db.estimates.insert_one({**stock_estimate_data, "order_id": ObjectId(order_id)})
                    db.orders.update_one({"_id": ObjectId(order_id)}, {"$set": {
                        "status": status, "estimate_created": True,
                        "estimate_id": stock_estimate_data["estimate_id"],
                        "estimate_number": stock_estimate_data["estimate_number"],
                        "estimate_url": stock_estimate_data.get("estimate_url", ""),
                    }})
                    message += f"Estimate created: {stock_estimate_data['estimate_number']}\n"
            else:
                if status in {"accepted", "declined"}:
                    # Don't PUT-update a closed estimate — Zoho rejects it.
                    # Just carry forward the stored IDs so email_estimate can
                    # push the status change.
                    stock_estimate_data = {
                        "estimate_id": estimate_id,
                        "estimate_number": order.get("estimate_number", ""),
                        "estimate_url": order.get("estimate_url", ""),
                    }
                    message += f"Estimate {order.get('estimate_number', '')} {status}\n"
                else:
                    payload = _base_payload(stock_line_items, "Looking forward for your business.")
                    resp = await client.put(
                        url=f"https://books.zoho.com/api/v3/estimates/{estimate_id}?organization_id={org_id}",
                        headers=headers,
                        json=payload,
                    )
                    rj = resp.json()
                    if resp.status_code != 200 or rj.get("code", 0) != 0:
                        errors.append(f"In-stock estimate: {rj.get('message', 'Unknown error')}")
                    else:
                        stock_estimate_data = rj["estimate"]
                        db.orders.update_one({"_id": ObjectId(order_id)}, {"$set": {
                            "status": status, "estimate_created": True,
                            "estimate_id": stock_estimate_data["estimate_id"],
                            "estimate_number": stock_estimate_data["estimate_number"],
                            "estimate_url": stock_estimate_data.get("estimate_url", ""),
                        }})
                        message += f"Estimate updated: {stock_estimate_data['estimate_number']}\n"

        # ── Pre-order estimate ──
        if pre_order_line_items and create_pre_order:
            if not pre_order_estimate_created:
                counter = db.counters.find_one_and_update(
                    {"_id": counter_id}, {"$inc": {"seq": 1}}, return_document=True
                )
                new_po_num = f"{prefix}/{fy_str}/{str(counter['seq']).zfill(num_width)}"
                payload = {"estimate_number": new_po_num, **_base_payload(pre_order_line_items, "PRE-ORDER — Items will be fulfilled when stock arrives.")}
                resp = await client.post(
                    url=ESTIMATE_URL.format(org_id=org_id) + "&ignore_auto_number_generation=true",
                    headers=headers,
                    json=payload,
                )
                print(resp.json())
                rj = resp.json()
                if resp.status_code != 201 or rj.get("code", 0) != 0:
                    errors.append(f"Pre-order estimate: {rj.get('message', 'Unknown error')}")
                else:
                    preorder_estimate_data = rj["estimate"]
                    db.estimates.insert_one({**preorder_estimate_data, "order_id": ObjectId(order_id), "is_pre_order": True})
                    db.orders.update_one({"_id": ObjectId(order_id)}, {"$set": {
                        "status": status,
                        "pre_order_estimate_created": True,
                        "pre_order_estimate_id": preorder_estimate_data["estimate_id"],
                        "pre_order_estimate_number": preorder_estimate_data["estimate_number"],
                        "pre_order_estimate_url": preorder_estimate_data.get("estimate_url", ""),
                    }})
                    message += f"Pre-order estimate created: {preorder_estimate_data['estimate_number']}\n"
            else:
                if status in {"accepted", "declined"}:
                    preorder_estimate_data = {
                        "estimate_id": pre_order_estimate_id,
                        "estimate_number": order.get("pre_order_estimate_number", ""),
                        "estimate_url": order.get("pre_order_estimate_url", ""),
                    }
                    message += f"Pre-order estimate {order.get('pre_order_estimate_number', '')} {status}\n"
                else:
                    payload = _base_payload(pre_order_line_items, "PRE-ORDER — Items will be fulfilled when stock arrives.")
                    resp = await client.put(
                        url=f"https://books.zoho.com/api/v3/estimates/{pre_order_estimate_id}?organization_id={org_id}",
                        headers=headers,
                        json=payload,
                    )
                    rj = resp.json()
                    if resp.status_code != 200 or rj.get("code", 0) != 0:
                        errors.append(f"Pre-order estimate: {rj.get('message', 'Unknown error')}")
                    else:
                        preorder_estimate_data = rj["estimate"]
                        db.orders.update_one({"_id": ObjectId(order_id)}, {"$set": {
                            "status": status,
                            "pre_order_estimate_created": True,
                            "pre_order_estimate_id": preorder_estimate_data["estimate_id"],
                            "pre_order_estimate_number": preorder_estimate_data["estimate_number"],
                            "pre_order_estimate_url": preorder_estimate_data.get("estimate_url", ""),
                        }})
                        message += f"Pre-order estimate updated: {preorder_estimate_data['estimate_number']}\n"

    # Always persist the status change (estimates may already exist)
    db.orders.update_one({"_id": ObjectId(order_id)}, {"$set": {"status": status}})

    # Email primary estimate (in-stock first, fall back to pre-order)
    primary_estimate = stock_estimate_data if stock_estimate_data else preorder_estimate_data
    if primary_estimate:
        msg_fragment = await email_estimate(
            status,
            order_id,
            primary_estimate["estimate_id"],
            primary_estimate["estimate_number"],
            primary_estimate.get("estimate_url", ""),
            headers,
            timeout,
        )
        if msg_fragment:
            message += msg_fragment + "\n"
        # Sync status to db.estimates immediately — don't wait for the Zoho webhook
        if status in {"accepted", "declined"} and primary_estimate.get("estimate_number"):
            db.estimates.update_one(
                {"estimate_number": primary_estimate["estimate_number"]},
                {"$set": {"status": status}},
            )

    # For accept/decline: also push the status to the pre-order estimate in Zoho
    # (email_estimate only handles the primary/in-stock estimate)
    if status in {"accepted", "declined"} and pre_order_estimate_id:
        _po_est_id = preorder_estimate_data.get("estimate_id") or pre_order_estimate_id
        async with httpx.AsyncClient(timeout=timeout) as _client:
            await _client.post(
                url=f"https://books.zoho.com/api/v3/estimates/{_po_est_id}/status/sent?organization_id={org_id}",
                headers=headers,
            )
            await _client.post(
                url=f"https://books.zoho.com/api/v3/estimates/{_po_est_id}/status/{status}?organization_id={org_id}",
                headers=headers,
            )
        db.orders.update_one(
            {"_id": ObjectId(order_id)},
            {"$set": {"pre_order_estimate_created": True, "pre_order_estimate_id": _po_est_id}},
        )
        # Sync pre-order status to db.estimates immediately — don't wait for the Zoho webhook
        _po_est_num = preorder_estimate_data.get("estimate_number") or order.get("pre_order_estimate_number", "")
        if _po_est_num:
            db.estimates.update_one(
                {"estimate_number": _po_est_num},
                {"$set": {"status": status}},
            )

    # In-app notification: order placed/updated → salesperson + customer
    estimate_data = stock_estimate_data or preorder_estimate_data
    try:
        est_number = estimate_data.get("estimate_number", order_id[-6:])
        customer_name = order.get("customer_name", "")
        sp_link = f"/orders/past/{order_id}"
        is_first_create = not estimate_created  # True only on the very first submit

        # Notify salesperson — skip if the order was placed by the customer themselves
        if created_by and user and user.get("role") != "customer":
            if is_first_create:
                create_notification(
                    db,
                    str(created_by),
                    "order_placed",
                    f"Estimate {est_number} created",
                    f"Order for {customer_name} has been finalised.",
                    sp_link,
                )
            else:
                create_notification(
                    db,
                    str(created_by),
                    "order_edited",
                    f"Order {est_number} updated",
                    f"Order for {customer_name} has been updated.",
                    sp_link,
                )

        # Notify the customer only on the first submit, not on every update
        if is_first_create:
            customer_doc = db.customers.find_one({"_id": ObjectId(order.get("customer_id", ""))}) if order.get("customer_id") else None
            if customer_doc:
                contact_id = str(customer_doc.get("contact_id", ""))
                customer_user = db.users.find_one({"role": "customer", "customer_id": contact_id})
                if customer_user:
                    create_notification(
                        db,
                        str(customer_user["_id"]),
                        "order_placed",
                        f"Your order {est_number} is confirmed",
                        f"Your estimate {est_number} has been created successfully.",
                        f"/orders/new/{order_id}",
                    )

        # Notify Invoicee users whenever a customer saves or updates an order
        if user and user.get("role") == "customer":
            notif_type = "order_placed" if is_first_create else "order_edited"
            notif_title = (
                f"New order {est_number} from {customer_name}"
                if is_first_create
                else f"Order {est_number} updated by {customer_name}"
            )
            notif_body = (
                f"Customer {customer_name} has placed order {est_number}."
                if is_first_create
                else f"Customer {customer_name} has updated order {est_number}."
            )
            for invoicee in db.users.find({"designation": "Invoicee", "status": "active"}, {"_id": 1}):
                create_notification(db, str(invoicee["_id"]), notif_type, notif_title, notif_body, sp_link)
    except Exception as _e:
        print(f"[notifications] order_placed error: {_e}")

    if errors and not message:
        return {"status": "error", "message": "; ".join(errors)}
    if errors:
        message += "\nWarnings: " + "; ".join(errors)
    return {"status": "success", "message": message}


@router.get("/download_pdf/{order_id}")
async def download_pdf(order_id: str = "", estimate_type: str = Query("stock", alias="type")):
    try:
        # Check if the order exists in the database
        if estimate_type == "pre_order":
            order = db.orders.find_one(
                {"_id": ObjectId(order_id), "pre_order_estimate_created": True}
            )
            if order is None:
                return {"status": "error", "message": "Pre-Order Estimate Not Created"}
            estimate_id = order.get("pre_order_estimate_id", "")
        else:
            order = db.orders.find_one(
                {"_id": ObjectId(order_id), "estimate_created": True}
            )
            if order is None:
                return {"status": "error", "message": "Draft Estimate Not Created"}
            estimate_id = order.get("estimate_id", "")
        headers = {"Authorization": f"Zoho-oauthtoken {get_access_token('books')}"}
        response = requests.get(
            url=PDF_URL.format(org_id=org_id, estimate_id=estimate_id),
            headers=headers,
            allow_redirects=False,  # Prevent automatic redirects
        )

        # Check if the response from Zoho is successful (200)
        if response.status_code == 200:
            # Return the PDF content
            if estimate_type == "pre_order":
                pdf_filename = order.get("pre_order_estimate_number", f"preorder_{order_id}") + ".pdf"
            else:
                pdf_filename = order.get("estimate_number", f"order_{order_id}") + ".pdf"
            return Response(
                content=response.content,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'attachment; filename="{pdf_filename}"'
                },
            )
        elif response.status_code == 307:
            raise HTTPException(
                status_code=307,
                detail="Redirect encountered. Check Zoho endpoint or token.",
            )
        else:
            # Raise an exception if Zoho's API returns an error
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Failed to fetch PDF: {response.text}",
            )

    except HTTPException as e:
        print(f"HTTP Exception: {e.detail}")
        raise e
    except Exception as e:
        print(f"Unexpected error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/notify")
async def notify(order_dict: dict):
    try:
        order_id = order_dict.get("order_id", "")
        if not order_id:
            raise HTTPException(status_code=404, detail="Order Id is neccesary")
        order = db.orders.find_one({"_id": ObjectId(order_id)})
        customer_name = order.get("customer_name")
        estimate_created = order.get("estimate_created", False)
        estimate_number = order.get("estimate_number", False)
        created_by = order.get("created_by", "")
        sales_person = db.users.find_one({"_id": ObjectId(created_by)})
        sales_person_phone = sales_person.get("phone")
        salesperson_name = sales_person.get("name")
        template = db.templates.find_one({"name": "customer_order_edit"})
        template_doc = {**template}
        params = {
            "salesperson_name": salesperson_name,
            "customer_name": customer_name,
            "estimate_number": estimate_number if estimate_created else order_id[-6:],
            "button_url": f"{order_id}",
        }
        # for item in [
        #     {"name": salesperson_name, "phone": sales_person_phone},
        #     {
        #         "name": os.getenv("NOTIFY_NUMBER_TO_CC4_NAME"),
        #         "phone": os.getenv("NOTIFY_NUMBER_TO_CC4"),
        #     },
        #     {
        #         "name": os.getenv("NOTIFY_NUMBER_TO_CC5_NAME"),
        #         "phone": os.getenv("NOTIFY_NUMBER_TO_CC5"),
        #     },
        # ]:
        #     params["salesperson_name"] = item["name"]
        #     send_whatsapp(to=item["phone"], template_doc=template_doc, params=params)

        # In-app notification for order edited → salesperson only
        ref = estimate_number if estimate_created else order_id[-6:]
        create_notification(
            db,
            str(created_by),
            "order_edited",
            f"Order {ref} edited",
            f"Order for {customer_name} has been updated.",
            f"/orders/past/{order_id}",
        )
        return
    except Exception as e:
        raise e


@router.post("/duplicate_order")
async def duplicate_order(order_dict: dict):
    try:
        order_id = order_dict.get("order_id", "")
        if not order_id:
            raise HTTPException(status_code=404, detail="Order Id is neccesary")
        order = db.orders.find_one({"_id": ObjectId(order_id)})
        order["created_at"] = datetime.now()
        order["updated_at"] = datetime.now()
        order["status"] = "draft"
        order.pop("_id")
        if "estimate_created" in order.keys():
            order.pop("estimate_created")
            order.pop("estimate_number")
            order.pop("estimate_id")
            order.pop("estimate_url")
        result = db.orders.insert_one(order)
        return str(result.inserted_id)
    except Exception as e:
        raise e


@router.delete("/clear_sheet/{order_id}")
async def clear_sheet(order_id: str):
    """
    Clear the Google Sheet URL and mark spreadsheet_created as false for an order
    """
    try:
        # Validate order exists
        order = db.orders.find_one({"_id": ObjectId(order_id)})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        # Update the order to remove sheet data
        update_result = db.orders.update_one(
            {"_id": ObjectId(order_id)},
            {
                "$set": {
                    "spreadsheet_created": False,
                    "updated_at": datetime.now()
                },
                "$unset": {
                    "spreadsheet_url": "",
                    "last_sheet_update": "",
                    "updated_from_sheet": ""
                }
            }
        )

        if update_result.modified_count == 0:
            return {
                "status": "warning",
                "message": "No changes made - sheet may already be cleared"
            }

        return {
            "status": "success",
            "message": "Sheet cleared successfully. You can now create a new sheet."
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error clearing sheet: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to clear sheet: {str(e)}")
