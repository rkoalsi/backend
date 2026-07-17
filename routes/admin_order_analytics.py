"""
Admin Order Analytics — how well the order form is actually being used.

Gives a month-wise / day-wise breakdown of orders *placed on the order form*
(the `orders` collection), plus a split of who is driving them:

  * "abandoned" orders — a user clicked "Create Order" which inserts a blank
    draft, then wandered off without adding any products or value. These are
    excluded from every real metric (and counted separately) so they don't
    inflate the platform-usage numbers.
  * created vs finalised — an order is "finalised" once it has been turned
    into a Zoho estimate (`estimate_created` / `pre_order_estimate_created`).
  * customer-driven vs salesperson-driven — an order is "customer created"
    when its `created_by` user has role "customer" (self-service order form),
    matching the `placed_by_customer` flag used elsewhere in orders.py.
  * products added by the customer vs the salesperson — the per-line-item
    `added_by` field on `products[]`.

Mounted under /api/admin/order_analytics (see admin.py).
"""

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, StreamingResponse
from ..config.root import get_database
from bson import ObjectId
from datetime import datetime
from typing import Optional
from io import BytesIO
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()
db = get_database()
orders_collection = db["orders"]
users_collection = db["users"]

logger = logging.getLogger(__name__)
logger.propagate = False


def _safe_num(field):
    """Coerce a possibly-string / missing numeric field to a double (0 on error)."""
    return {"$convert": {"input": field, "to": "double", "onError": 0, "onNull": 0}}


# Per-line-item stats. `quantity` / `price` have mixed types in the data
# (some empty strings) — $multiply / $toInt throw on strings, so guard every
# numeric access with $isNumber. See project-orders-products-analytics memory.
_PROD_STATS = {
    "$map": {
        "input": {"$ifNull": ["$products", []]},
        "as": "p",
        "in": {
            "added_by": {
                "$let": {
                    "vars": {"ab": {"$ifNull": ["$$p.added_by", ""]}},
                    # Keep the real value (customer / sales_person / admin / hr);
                    # only blank / missing collapses to "unknown".
                    "in": {"$cond": [{"$eq": ["$$ab", ""]}, "unknown", "$$ab"]},
                }
            },
            "qty": {"$cond": [{"$isNumber": "$$p.quantity"}, "$$p.quantity", 0]},
            "value": {
                "$multiply": [
                    {"$cond": [{"$isNumber": "$$p.price"}, "$$p.price", 0]},
                    {"$cond": [{"$isNumber": "$$p.quantity"}, "$$p.quantity", 0]},
                ]
            },
        },
    }
}


def _added_agg(added_by_val, metric):
    """Aggregate a metric ('items' | 'qty' | 'value') over line items added by
    a given actor, computed from the per-doc `_prodStats` array."""
    filt = {
        "$filter": {
            "input": "$_prodStats",
            "as": "s",
            "cond": {"$eq": ["$$s.added_by", added_by_val]},
        }
    }
    if metric == "items":
        return {"$size": filt}
    return {"$sum": {"$map": {"input": filt, "as": "s", "in": f"$$s.{metric}"}}}


def _metrics_group(id_expr):
    """Group spec producing the shared per-period / overall metric bundle."""
    return {
        "_id": id_expr,
        "totalOrders": {"$sum": 1},
        "finalisedOrders": {"$sum": {"$cond": ["$finalised", 1, 0]}},
        "customerCreatedOrders": {"$sum": {"$cond": ["$isCustomerCreated", 1, 0]}},
        "customerFinalisedOrders": {
            "$sum": {"$cond": [{"$and": ["$isCustomerCreated", "$finalised"]}, 1, 0]}
        },
        "salespersonCreatedOrders": {"$sum": {"$cond": ["$isCustomerCreated", 0, 1]}},
        "estimateOrders": {"$sum": {"$cond": ["$estimateCreated", 1, 0]}},
        "paidOrders": {"$sum": {"$cond": ["$paid", 1, 0]}},
        "customerPaidOrders": {
            "$sum": {"$cond": [{"$and": ["$isCustomerCreated", "$paid"]}, 1, 0]}
        },
        "totalValue": {"$sum": "$totalAmountNum"},
        "estimateValue": {"$sum": {"$cond": ["$estimateCreated", "$totalAmountNum", 0]}},
        "finalisedValue": {"$sum": {"$cond": ["$finalised", "$totalAmountNum", 0]}},
        "paidValue": {"$sum": {"$cond": ["$paid", "$totalAmountNum", 0]}},
        "customerAddedItems": {"$sum": "$customerAddedItems"},
        "customerAddedQty": {"$sum": "$customerAddedQty"},
        "customerAddedValue": {"$sum": "$customerAddedValue"},
        "salespersonAddedItems": {"$sum": "$salespersonAddedItems"},
        "salespersonAddedQty": {"$sum": "$salespersonAddedQty"},
        "salespersonAddedValue": {"$sum": "$salespersonAddedValue"},
    }


# Numeric fields to round when serialising the metric bundle.
_VALUE_KEYS = (
    "totalValue",
    "estimateValue",
    "finalisedValue",
    "paidValue",
    "customerAddedValue",
    "salespersonAddedValue",
)
_INT_KEYS = (
    "totalOrders",
    "estimateOrders",
    "finalisedOrders",
    "customerCreatedOrders",
    "customerFinalisedOrders",
    "salespersonCreatedOrders",
    "paidOrders",
    "customerPaidOrders",
    "customerAddedItems",
    "customerAddedQty",
    "salespersonAddedItems",
    "salespersonAddedQty",
)


def _build_paid_order_ids():
    """Return the set of order `_id`s (ObjectId) that resulted in a *paid*
    invoice, following the chain starting from the orders collection:

        orders.estimate_id (int) -> str
          -> estimates.estimate_id -> estimates.invoice_ids[] (str)
            -> invoices.invoice_id (str), status == 'paid' (void ignored)

    Only estimates / invoices actually referenced by platform orders are
    touched, so this stays cheap regardless of the full collection sizes.
    """
    # 1. Every platform order that carries an estimate id.
    order_to_est = []  # (order_id, estimate_id_str)
    est_ids = set()
    for o in orders_collection.find(
        {"estimate_id": {"$exists": True, "$ne": None}}, {"estimate_id": 1}
    ):
        es = str(o.get("estimate_id"))
        if es and es != "None":
            order_to_est.append((o["_id"], es))
            est_ids.add(es)
    if not est_ids:
        return set()

    # 2. Those estimates -> their invoice ids.
    est_to_inv = {}
    all_inv_ids = set()
    for e in db.estimates.find(
        {"estimate_id": {"$in": list(est_ids)}},
        {"estimate_id": 1, "invoice_ids": 1, "_id": 0},
    ):
        inv_ids = []
        for iv in e.get("invoice_ids") or []:
            iv_s = iv.get("invoice_id") if isinstance(iv, dict) else iv
            if iv_s is not None:
                inv_ids.append(str(iv_s))
                all_inv_ids.add(str(iv_s))
        est_to_inv[str(e.get("estimate_id"))] = inv_ids
    if not all_inv_ids:
        return set()

    # 3. Those invoices (excluding void) that are paid.
    paid_inv_ids = {
        str(inv.get("invoice_id"))
        for inv in db.invoices.find(
            {
                "invoice_id": {"$in": list(all_inv_ids)},
                "status": {"$ne": "void"},
            },
            {"invoice_id": 1, "status": 1, "_id": 0},
        )
        if inv.get("status") == "paid"
    }
    if not paid_inv_ids:
        return set()

    # 4. An order is "paid" if any of its estimate's invoices is paid.
    paid_orders = set()
    for oid, es in order_to_est:
        for iv in est_to_inv.get(es, []):
            if iv in paid_inv_ids:
                paid_orders.add(oid)
                break
    return paid_orders


# Internal/test accounts excluded from activity stats — kept in sync with the
# EXCLUDED_EMAILS set on the /admin/customer_activity frontend so counts match.
_EXCLUDED_ACTIVITY_EMAILS = [
    "rkoalsi2000@gmail.com",
    "rkoalsi2175@gmail.com",
    "rkoalsi2@gmail.com",
]


def _activity_order_ids(action):
    """Order `_id`s (ObjectId) that a *customer* account performed `action` on,
    from `customer_activity_logs` — the same source (and test-account exclusion)
    the /admin/customer_activity page uses. `create_order` = customer started it,
    `finalize_order` = ended it."""
    oids = []
    for x in db["customer_activity_logs"].distinct(
        "metadata.order_id",
        {"action": action, "email": {"$nin": _EXCLUDED_ACTIVITY_EMAILS}},
    ):
        if x and ObjectId.is_valid(str(x)):
            oids.append(ObjectId(str(x)))
    return oids


def _activity_customer_count(action):
    """Distinct customers who performed `action` (test accounts excluded)."""
    ids = db["customer_activity_logs"].distinct(
        "customer_id",
        {"action": action, "email": {"$nin": _EXCLUDED_ACTIVITY_EMAILS}},
    )
    return len([i for i in ids if i])


def _clean_metrics(doc: dict) -> dict:
    """Round money to 2dp, coerce counts to int, drop the group _id."""
    out = {}
    for k in _INT_KEYS:
        out[k] = int(doc.get(k) or 0)
    for k in _VALUE_KEYS:
        out[k] = round(float(doc.get(k) or 0), 2)
    return out


def _compute_analytics(granularity, start_date, end_date, created_by):
    """Run the full analytics aggregation and return the response dict.
    On bad input returns ("error", message); callers translate that to a 400.
    Shared by the JSON endpoint and the XLSX report endpoint."""
    fmt = "%Y-%m-%d" if granularity == "day" else "%Y-%m"

    # Users who place orders as customers themselves (self-service order form).
    customer_user_ids = [
        u["_id"] for u in users_collection.find({"role": "customer"}, {"_id": 1})
    ]

    # Orders that produced a paid (non-void) invoice — chain starts at orders.
    paid_order_ids = list(_build_paid_order_ids())

    # ── Base match ───────────────────────────────────────────────────────
    # Require a real date so period grouping works. We intentionally do NOT
    # filter `is_deleted` here: status='deleted' orders (all soft-deleted) are
    # counted, matching the /admin dashboard methodology. Draft/declined are
    # dropped later via the "real order" match.
    match_stage = {
        "created_at": {"$type": "date"},
    }
    date_bounds = {}
    if start_date:
        try:
            date_bounds["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            return ("error", "Invalid start_date")
    if end_date:
        try:
            # inclusive end-of-day
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            date_bounds["$lte"] = end_dt.replace(hour=23, minute=59, second=59)
        except ValueError:
            return ("error", "Invalid end_date")
    if date_bounds:
        match_stage["created_at"] = {"$type": "date", **date_bounds}

    pipeline = [
        {"$match": match_stage},
        # Stage A: cheap per-doc derived fields.
        {
            "$addFields": {
                "_prodStats": _PROD_STATS,
                "totalAmountNum": _safe_num("$total_amount"),
                # Estimate created = order was finalised into a Zoho estimate
                # (the step before it becomes an invoice).
                "estimateCreated": {
                    "$or": [
                        {"$eq": ["$estimate_created", True]},
                        {"$eq": ["$pre_order_estimate_created", True]},
                    ]
                },
                # Finalised = order was invoiced in Zoho.
                "finalised": {"$eq": ["$status", "invoiced"]},
                # Paid = order's estimate produced a paid (non-void) invoice.
                "paid": {"$in": ["$_id", paid_order_ids]},
                "isCustomerCreated": {"$in": ["$created_by", customer_user_ids]},
                "period": {"$dateToString": {"format": fmt, "date": "$created_at"}},
            }
        },
        # Stage B: fields derived from _prodStats / totalAmountNum.
        {
            "$addFields": {
                "totalQty": {"$sum": {"$map": {"input": "$_prodStats", "as": "s", "in": "$$s.qty"}}},
                "customerAddedItems": _added_agg("customer", "items"),
                "customerAddedQty": _added_agg("customer", "qty"),
                "customerAddedValue": _added_agg("customer", "value"),
                "salespersonAddedItems": _added_agg("sales_person", "items"),
                "salespersonAddedQty": _added_agg("sales_person", "qty"),
                "salespersonAddedValue": _added_agg("sales_person", "value"),
            }
        },
        # Stage C: "non-empty" = has real products or a positive value.
        # Abandoned "click create & leave" drafts fail both tests.
        {
            "$addFields": {
                "nonEmpty": {
                    "$or": [
                        {"$gt": ["$totalQty", 0]},
                        {"$gt": ["$totalAmountNum", 0]},
                    ]
                }
            }
        },
    ]

    # Optional created-by filter (affects every non-abandoned branch).
    creator_match = None
    if created_by == "customer":
        creator_match = {"isCustomerCreated": True}
    elif created_by == "sales_person":
        creator_match = {"isCustomerCreated": False}

    # A "real" order = non-empty AND not a never-placed draft / declined order.
    # (Matches the /admin dashboard's product-additions methodology so the
    # customer-vs-salesperson product counts reconcile.) The abandoned branch
    # deliberately keeps empty drafts — that's exactly what it measures.
    real_conds = [{"nonEmpty": True}, {"status": {"$nin": ["declined", "draft"]}}]
    abandoned_conds = [{"nonEmpty": False}]
    if creator_match:
        real_conds.append(creator_match)
        abandoned_conds.append(creator_match)
    non_empty_match = {"$and": real_conds}
    abandoned_match = (
        abandoned_conds[0] if len(abandoned_conds) == 1 else {"$and": abandoned_conds}
    )

    pipeline.append(
        {
            "$facet": {
                "summary": [{"$match": non_empty_match}, {"$group": _metrics_group(None)}],
                "periods": [
                    {"$match": non_empty_match},
                    {"$group": _metrics_group("$period")},
                    {"$sort": {"_id": 1}},
                ],
                "byAddedBy": [
                    {"$match": non_empty_match},
                    {"$unwind": "$_prodStats"},
                    {
                        "$group": {
                            "_id": "$_prodStats.added_by",
                            "lineItems": {"$sum": 1},
                            "qty": {"$sum": "$_prodStats.qty"},
                            "value": {"$sum": "$_prodStats.value"},
                        }
                    },
                    {"$sort": {"value": -1}},
                ],
                "byStatus": [
                    {"$match": non_empty_match},
                    {
                        "$group": {
                            "_id": {"$ifNull": ["$status", "unknown"]},
                            "count": {"$sum": 1},
                            "value": {"$sum": "$totalAmountNum"},
                        }
                    },
                    {"$sort": {"count": -1}},
                ],
                # Customers whose orders flow through the form (by customer_id).
                "byCustomer": [
                    {"$match": non_empty_match},
                    {
                        "$group": {
                            "_id": "$customer_id",
                            "customerName": {"$first": "$customer_name"},
                            "orders": {"$sum": 1},
                            "finalised": {"$sum": {"$cond": ["$finalised", 1, 0]}},
                            "paid": {"$sum": {"$cond": ["$paid", 1, 0]}},
                            "value": {"$sum": "$totalAmountNum"},
                            "paidValue": {"$sum": {"$cond": ["$paid", "$totalAmountNum", 0]}},
                            "customerAddedItems": {"$sum": "$customerAddedItems"},
                            "lastOrder": {"$max": "$created_at"},
                        }
                    },
                    {"$sort": {"orders": -1, "value": -1}},
                ],
                # Users who create orders (resolved to name/role in Python).
                "byCreator": [
                    {"$match": non_empty_match},
                    {
                        "$group": {
                            "_id": "$created_by",
                            "orders": {"$sum": 1},
                            "finalised": {"$sum": {"$cond": ["$finalised", 1, 0]}},
                            "paid": {"$sum": {"$cond": ["$paid", 1, 0]}},
                            "value": {"$sum": "$totalAmountNum"},
                            "paidValue": {"$sum": {"$cond": ["$paid", "$totalAmountNum", 0]}},
                            "customerAddedItems": {"$sum": "$customerAddedItems"},
                            "lastOrder": {"$max": "$created_at"},
                        }
                    },
                    {"$sort": {"orders": -1, "value": -1}},
                ],
                "abandoned": [
                    {"$match": abandoned_match},
                    {
                        "$group": {
                            "_id": None,
                            "total": {"$sum": 1},
                            "customerCreated": {
                                "$sum": {"$cond": ["$isCustomerCreated", 1, 0]}
                            },
                        }
                    },
                ],
            }
        }
    )

    result = list(orders_collection.aggregate(pipeline, allowDiskUse=True))
    facet = result[0] if result else {}

    # ── Assemble response ────────────────────────────────────────────────
    summary_doc = facet.get("summary", [{}])
    summary = _clean_metrics(summary_doc[0]) if summary_doc else _clean_metrics({})

    abandoned_doc = facet.get("abandoned", [])
    abandoned = abandoned_doc[0] if abandoned_doc else {}
    summary["abandonedOrders"] = int(abandoned.get("total") or 0)
    summary["abandonedCustomerOrders"] = int(abandoned.get("customerCreated") or 0)

    # Payments due = invoiced orders that aren't paid yet.
    summary["dueOrders"] = max(0, summary["finalisedOrders"] - summary["paidOrders"])
    summary["dueValue"] = round(
        max(0.0, summary["finalisedValue"] - summary["paidValue"]), 2
    )

    # Customer engagement from activity tracking (matches /admin/customer_activity):
    # customers who started (create_order) / ended (finalize_order) orders themselves.
    started_oids = _activity_order_ids("create_order")
    finalised_oids = _activity_order_ids("finalize_order")
    oid_date_q = {}
    if date_bounds:
        oid_date_q = {"created_at": {"$type": "date", **date_bounds}}

    def _count_ids(oids):
        if not oids:
            return 0
        q = {"_id": {"$in": oids}}
        q.update(oid_date_q)
        return orders_collection.count_documents(q)

    paid_set = set(paid_order_ids)
    summary["customerStartedOrders"] = _count_ids(started_oids)
    summary["customerActivityFinalised"] = _count_ids(finalised_oids)
    summary["customerActivityPaid"] = len([o for o in finalised_oids if o in paid_set])
    summary["customerSelfServiceCount"] = _activity_customer_count("finalize_order")

    periods = []
    for p in facet.get("periods", []):
        metrics = _clean_metrics(p)
        metrics["period"] = p.get("_id")
        periods.append(metrics)

    products_by_added_by = [
        {
            "addedBy": b.get("_id") or "other",
            "lineItems": int(b.get("lineItems") or 0),
            "qty": int(b.get("qty") or 0),
            "value": round(float(b.get("value") or 0), 2),
        }
        for b in facet.get("byAddedBy", [])
    ]

    by_status = [
        {
            "status": s.get("_id") or "unknown",
            "count": int(s.get("count") or 0),
            "value": round(float(s.get("value") or 0), 2),
        }
        for s in facet.get("byStatus", [])
    ]

    # ── Customers list (by customer_id; name comes off the order) ─────────
    customers_list = []
    for g in facet.get("byCustomer", []):
        cid = g.get("_id")
        last = g.get("lastOrder")
        customers_list.append(
            {
                "id": str(cid) if cid else None,
                "name": g.get("customerName") or "Unknown",
                "orders": int(g.get("orders") or 0),
                "finalised": int(g.get("finalised") or 0),
                "paid": int(g.get("paid") or 0),
                "value": round(float(g.get("value") or 0), 2),
                "paidValue": round(float(g.get("paidValue") or 0), 2),
                "productsAdded": int(g.get("customerAddedItems") or 0),
                "lastOrder": last.strftime("%Y-%m-%d") if last else None,
            }
        )

    # ── Salespeople list (resolve created_by -> user name/role) ──────────
    creator_groups = facet.get("byCreator", [])
    creator_ids = [g["_id"] for g in creator_groups if g.get("_id")]
    users_map = {
        u["_id"]: u
        for u in users_collection.find(
            {"_id": {"$in": creator_ids}},
            {"name": 1, "email": 1, "role": 1, "first_name": 1, "last_name": 1},
        )
    }

    salespeople_list = []
    customer_creators_list = []
    for g in creator_groups:
        uid = g.get("_id")
        u = users_map.get(uid) or {}
        role = u.get("role") or "unknown"
        name = (
            u.get("name")
            or f"{u.get('first_name', '')} {u.get('last_name', '')}".strip()
            or u.get("email")
            or "Unknown"
        )
        last = g.get("lastOrder")
        row = {
            "id": str(uid) if uid else None,
            "name": name,
            "email": u.get("email", ""),
            "role": role,
            "orders": int(g.get("orders") or 0),
            "finalised": int(g.get("finalised") or 0),
            "paid": int(g.get("paid") or 0),
            "value": round(float(g.get("value") or 0), 2),
            "paidValue": round(float(g.get("paidValue") or 0), 2),
            "lastOrder": last.strftime("%Y-%m-%d") if last else None,
        }
        if role == "customer":
            customer_creators_list.append(row)
        else:
            salespeople_list.append(row)

    return {
        "granularity": "day" if granularity == "day" else "month",
        "start_date": start_date,
        "end_date": end_date,
        "created_by": created_by,
        "summary": summary,
        "periods": periods,
        "productsByAddedBy": products_by_added_by,
        "byStatus": by_status,
        "customers": customers_list,
        "salespeople": salespeople_list,
        "customerCreators": customer_creators_list,
    }


@router.get("")
def get_order_analytics(
    granularity: str = Query("month", description="'month' or 'day'"),
    start_date: Optional[str] = Query(None, description="ISO date (YYYY-MM-DD), inclusive"),
    end_date: Optional[str] = Query(None, description="ISO date (YYYY-MM-DD), inclusive"),
    created_by: str = Query("all", description="'all' | 'customer' | 'sales_person'"),
):
    """Month-wise / day-wise order-form usage breakdown, excluding abandoned
    (empty, no-value) orders from the real metrics."""
    try:
        out = _compute_analytics(granularity, start_date, end_date, created_by)
        if isinstance(out, tuple) and out and out[0] == "error":
            return JSONResponse(status_code=400, content={"error": out[1]})
        return out
    except Exception as e:
        logger.error(f"Error in get_order_analytics: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ── XLSX report ────────────────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_sheet(ws, headers, rows):
    """Write a header row + data rows and auto-size columns.
    `rows` is a list of lists aligned with `headers`."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for column in ws.columns:
        max_len = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 45)


# section -> (sheet title, headers, row-builder taking the analytics dict)
def _periods_rows(data):
    return [
        [
            p["period"],
            p["totalOrders"],
            p["estimateOrders"],
            p["finalisedOrders"],
            p["paidOrders"],
            p["customerCreatedOrders"],
            p["customerFinalisedOrders"],
            p["salespersonCreatedOrders"],
            p["customerAddedItems"],
            p["salespersonAddedItems"],
            p["totalValue"],
            p["estimateValue"],
            p["finalisedValue"],
            p["paidValue"],
        ]
        for p in data["periods"]
    ]


_PERIODS_HEADERS = [
    "Period",
    "Orders",
    "Estimates Created",
    "Finalised (Invoiced)",
    "Paid",
    "Created by Customers",
    "Finalised by Customers",
    "Created by Salespeople",
    "Products Added by Customers",
    "Products Added by Salespeople",
    "Total Value",
    "Estimate Value",
    "Finalised Value",
    "Paid Value",
]


def _products_rows(data):
    return [
        [b["addedBy"], b["lineItems"], b["qty"], b["value"]]
        for b in data["productsByAddedBy"]
    ]


def _customers_rows(data):
    return [
        [
            c["name"],
            c["orders"],
            c["finalised"],
            c["paid"],
            c["productsAdded"],
            c["value"],
            c["paidValue"],
            c["lastOrder"] or "",
        ]
        for c in data["customers"]
    ]


def _salespeople_rows(data):
    return [
        [
            s["name"],
            s["email"],
            s["role"],
            s["orders"],
            s["finalised"],
            s["paid"],
            s["value"],
            s["paidValue"],
            s["lastOrder"] or "",
        ]
        for s in data["salespeople"] + data["customerCreators"]
    ]


_SECTIONS = {
    "periods": ("Period Breakdown", _PERIODS_HEADERS, _periods_rows),
    "products": (
        "Products Added",
        ["Added By", "Line Items", "Units", "Value"],
        _products_rows,
    ),
    "customers": (
        "Customers",
        [
            "Customer",
            "Orders",
            "Finalised",
            "Paid",
            "Self-Added Items",
            "Value",
            "Paid Value",
            "Last Order",
        ],
        _customers_rows,
    ),
    "salespeople": (
        "Salespeople",
        [
            "Name",
            "Email",
            "Role",
            "Orders",
            "Finalised",
            "Paid",
            "Value",
            "Paid Value",
            "Last Order",
        ],
        _salespeople_rows,
    ),
}


@router.get("/report")
def download_order_analytics_report(
    section: str = Query("all", description="periods | products | customers | salespeople | all"),
    granularity: str = Query("month", description="'month' or 'day'"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    created_by: str = Query("all"),
):
    """Download a section (or all sections) of the order analytics as XLSX."""
    try:
        data = _compute_analytics(granularity, start_date, end_date, created_by)
        if isinstance(data, tuple) and data and data[0] == "error":
            return JSONResponse(status_code=400, content={"error": data[1]})

        wanted = list(_SECTIONS.keys()) if section == "all" else [section]
        if any(s not in _SECTIONS for s in wanted):
            return JSONResponse(status_code=400, content={"error": "Invalid section"})

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sec in wanted:
            title, headers, builder = _SECTIONS[sec]
            ws = wb.create_sheet(title)
            _write_sheet(ws, headers, builder(data))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"order_analytics_{section}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    except Exception as e:
        logger.error(f"Error in download_order_analytics_report: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})
