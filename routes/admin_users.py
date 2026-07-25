from fastapi import APIRouter, HTTPException, Query, File, UploadFile, Body
from fastapi.responses import StreamingResponse
from ..config.root import get_database, serialize_mongo_document
from ..config.phone import normalize_indian_mobile
from bson.objectid import ObjectId
from passlib.hash import bcrypt
from datetime import datetime
from typing import Optional
import secrets
import string
import re
import openpyxl
from io import BytesIO

router = APIRouter()
db = get_database()


def hash_password(password: str) -> str:
    """Hash a password using bcrypt."""
    return bcrypt.hash(password)


def generate_password(length: int = 12) -> str:
    """Generate a secure random password."""
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    password = ''.join(secrets.choice(alphabet) for _ in range(length))
    return password


@router.get("")
def get_all_users(
    search: Optional[str] = Query(None, description="Search by name or email"),
    role: Optional[str] = Query(None, description="Filter by role"),
    status: Optional[str] = Query(None, description="Filter by status"),
    linked: Optional[str] = Query(
        None, description="Filter by customer link: 'linked' or 'unlinked'"
    ),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100)
):
    """
    Get all users with optional filtering and pagination.
    """
    # Base query = search + role, but NOT status. The stats cards are computed
    # against this so they stay stable while the user toggles the status filter.
    base_query = {}

    if search:
        base_query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
        ]

    if role:
        base_query["role"] = role

    # A user counts as "linked" only if customer_id is present and non-empty.
    if linked == "linked":
        base_query["customer_id"] = {"$nin": [None, ""], "$exists": True}
    elif linked == "unlinked":
        base_query["customer_id"] = {"$in": [None, ""]}

    query = dict(base_query)

    if status:
        query["status"] = status

    # Get total count for pagination
    total = db.users.count_documents(query)

    # Get paginated users
    skip = (page - 1) * per_page
    users_cursor = db.users.find(query).sort("created_at", -1).skip(skip).limit(per_page)
    users = serialize_mongo_document(list(users_cursor))

    # Remove password field from response
    for user in users:
        user.pop("password", None)

    # Get statistics, scoped to the current search/role filters so the cards
    # describe the same population the table is paginating through.
    stats = {
        "total": db.users.count_documents(base_query),
        "active": db.users.count_documents({**base_query, "status": "active"}),
        "inactive": db.users.count_documents({**base_query, "status": "inactive"}),
        "by_role": {}
    }

    # Count by role
    roles = ["admin", "sales_admin", "sales_person", "warehouse", "catalogue_manager", "hr", "customer"]
    for r in roles:
        stats["by_role"][r] = db.users.count_documents({"role": r})

    return {
        "users": users,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "stats": stats
    }


@router.get("/roles")
def get_available_roles():
    """Get list of available user roles."""
    return {
        "roles": [
            {"value": "admin", "label": "Admin"},
            {"value": "sales_admin", "label": "Sales Admin"},
            {"value": "sales_person", "label": "Sales Person"},
            {"value": "warehouse", "label": "Warehouse"},
            {"value": "catalogue_manager", "label": "Catalogue Manager"},
            {"value": "hr", "label": "HR"},
            {"value": "customer", "label": "Customer"},
        ]
    }


@router.get("/generate-password")
def generate_new_password():
    """Generate a new random password."""
    return {"password": generate_password()}


@router.get("/bulk-upload/template")
def download_bulk_upload_template():
    """Download the XLSX template for bulk customer upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["First Name", "Last Name", "Email", "Whatsapp Phone Number", "Zoho Customer Name"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    col_widths = [20, 20, 30, 25, 35]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_upload_template.xlsx"},
    )


@router.post("/bulk-upload/preview")
def preview_bulk_upload(file: UploadFile = File(...)):
    """Parse uploaded XLSX and match Zoho Customer Name against the customers collection."""
    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read file. Please upload a valid .xlsx file.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    expected = ["First Name", "Last Name", "Email", "Whatsapp Phone Number", "Zoho Customer Name"]
    if header != expected:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid template headers. Expected: {', '.join(expected)}. Got: {', '.join(header)}"
        )

    found = []
    not_found = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        first_name = str(row[0]).strip() if row[0] is not None else ""
        last_name = str(row[1]).strip() if row[1] is not None else ""
        email = str(row[2]).strip() if row[2] is not None else ""
        phone = str(row[3]).strip() if row[3] is not None else ""
        zoho_name = str(row[4]).strip() if row[4] is not None else ""

        entry = {
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "phone": phone,
            "zoho_customer_name": zoho_name,
            "row": i,
        }

        if not zoho_name:
            not_found.append({**entry, "reason": "No Zoho Customer Name provided"})
            continue

        customer = db.customers.find_one(
            {
                "$or": [
                    {"customer_name": {"$regex": re.escape(zoho_name), "$options": "i"}},
                    {"company_name": {"$regex": re.escape(zoho_name), "$options": "i"}},
                    {"contact_name": {"$regex": re.escape(zoho_name), "$options": "i"}},
                ]
            },
            {"_id": 0, "contact_id": 1, "contact_name": 1, "company_name": 1, "customer_name": 1},
        )

        if customer:
            matched_name = (
                customer.get("company_name")
                or customer.get("contact_name")
                or customer.get("customer_name")
                or zoho_name
            )
            found.append({
                **entry,
                "customer_id": customer.get("contact_id"),
                "matched_customer_name": matched_name,
            })
        else:
            not_found.append({**entry, "reason": "No matching customer found in database"})

    return {"found": found, "not_found": not_found}


@router.post("/bulk-upload/create")
def create_bulk_users(data: dict):
    """Create user accounts for confirmed entries from the bulk upload preview."""
    entries = data.get("entries", [])
    if not entries:
        raise HTTPException(status_code=400, detail="No entries provided.")

    created = []
    errors = []

    for entry in entries:
        email = (entry.get("email") or "").strip()
        first_name = (entry.get("first_name") or "").strip()
        last_name = (entry.get("last_name") or "").strip()
        phone = (entry.get("phone") or "").strip()
        customer_id = entry.get("customer_id") or ""
        matched_customer_name = entry.get("matched_customer_name") or ""

        if not email:
            errors.append({"entry": entry, "reason": "Email is required"})
            continue

        if db.users.find_one({"email": email}):
            errors.append({"entry": entry, "reason": f"Email {email} already exists"})
            continue

        name = f"{first_name} {last_name}".strip() or email

        plain_password = generate_password()
        try:
            phone_int = int(phone) if phone else 0
        except ValueError:
            phone_int = 0

        user_doc = {
            "name": name,
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "phone": phone_int,
            "role": "customer",
            "status": "active",
            "password": hash_password(plain_password),
            "customer_id": customer_id,
            "customer_name": matched_customer_name,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        }

        result = db.users.insert_one(user_doc)
        created.append({
            "user_id": str(result.inserted_id),
            "name": name,
            "email": email,
            "password": plain_password,
            "customer_name": matched_customer_name,
        })

    return {"created": created, "errors": errors, "total_created": len(created)}


@router.get("/by-customer/{contact_id}")
def get_user_by_customer(contact_id: str):
    """Get the user account linked to a Zoho customer by their contact_id."""
    user = db.users.find_one({"customer_id": contact_id, "role": "customer"})
    if not user:
        raise HTTPException(status_code=404, detail="No user account found for this customer")
    user = serialize_mongo_document(user)
    return {"user": {"email": user.get("email"), "name": user.get("name"), "_id": user.get("_id")}}


# ---------------------------------------------------------------------------
# Customer logins driven from /admin/customers (mirrors the flow on
# /admin/customer_requests, but keyed off the customer record itself).
# ---------------------------------------------------------------------------

@router.get("/customer-login/{contact_id}")
def get_customer_login(contact_id: str):
    """
    Login status for a customer, plus whether their number can take a WhatsApp
    message — so the UI can warn before anyone tries to create or send.
    """
    customer = db.customers.find_one(
        {"contact_id": contact_id},
        {"contact_id": 1, "contact_name": 1, "company_name": 1, "email": 1, "phone": 1, "mobile": 1},
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # A customer can carry more than one login (created by a salesperson, by the
    # admin, or self-registered), so return all of them rather than an arbitrary
    # find_one — which one you're looking at matters when sending.
    users = list(
        db.users.find(
            {"customer_id": contact_id, "role": "customer"},
            {"_id": 1, "name": 1, "email": 1, "phone": 1, "status": 1, "password": 1,
             "created_at": 1, "created_by_salesperson_id": 1},
        ).sort("created_at", 1)
    )

    logins = []
    for entry in users:
        item = serialize_mongo_document(entry)
        item["has_password"] = bool(item.pop("password", None))
        item["phone_info"] = normalize_indian_mobile(entry.get("phone"))
        item["created_by_salesperson"] = bool(item.pop("created_by_salesperson_id", None))
        logins.append(item)

    # Kept for callers that only care about "is there a login at all".
    user = users[0] if users else None
    login = logins[0] if logins else None

    # The login's own number wins once it exists; otherwise fall back to whatever
    # the customer record carries.
    raw_phone = (user or {}).get("phone") or customer.get("mobile") or customer.get("phone")
    resolved = normalize_indian_mobile(raw_phone)

    # Surface accounts belonging to SOMEONE ELSE that already claim this number or
    # email. Creating would be rejected anyway; showing who holds it up front means
    # the admin can go fix the right record instead of guessing at a 409.
    conflicts = []
    if not user:
        conflict_query = []
        if resolved["valid"]:
            conflict_query.append({"phone": int(resolved["phone"])})
            conflict_query.append({"phone": resolved["phone"]})
        customer_email = (customer.get("email") or "").strip()
        if customer_email:
            conflict_query.append({"email": customer_email})

        if conflict_query:
            for other in db.users.find(
                {"$and": [{"$or": conflict_query}, {"customer_id": {"$ne": contact_id}}]},
                {"_id": 1, "name": 1, "email": 1, "phone": 1, "role": 1,
                 "status": 1, "customer_id": 1, "customer_name": 1},
            ).limit(10):
                other_phone = normalize_indian_mobile(other.get("phone"))["phone"]
                reasons = []
                if resolved["valid"] and other_phone == resolved["phone"]:
                    reasons.append("mobile")
                if customer_email and (other.get("email") or "").strip().lower() == customer_email.lower():
                    reasons.append("email")

                linked_name = other.get("customer_name")
                if not linked_name and other.get("customer_id"):
                    linked = db.customers.find_one(
                        {"contact_id": other["customer_id"]},
                        {"company_name": 1, "contact_name": 1},
                    )
                    linked_name = (linked or {}).get("company_name") or (linked or {}).get("contact_name")

                conflicts.append({
                    "_id": str(other["_id"]),
                    "name": other.get("name"),
                    "email": other.get("email"),
                    "phone": other_phone or str(other.get("phone") or ""),
                    "role": other.get("role"),
                    "status": other.get("status"),
                    "customer_id": other.get("customer_id"),
                    "customer_name": linked_name,
                    "conflict_on": reasons,
                })

    return {
        "login": login,
        "logins": logins,
        "customer": {
            "contact_id": customer.get("contact_id"),
            "name": customer.get("company_name") or customer.get("contact_name"),
            "contact_name": customer.get("contact_name"),
            "email": customer.get("email"),
            "raw_phone": raw_phone,
        },
        "phone": resolved,
        "conflicts": conflicts,
    }


@router.post("/customer-login/{contact_id}")
def create_customer_login_for_customer(contact_id: str, payload: dict = Body(...)):
    """
    Create a customer login straight from the customer record. Password is
    optional — omit it for an OTP-only account.
    """
    customer = db.customers.find_one({"contact_id": contact_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    if db.users.find_one({"customer_id": contact_id, "role": "customer"}):
        raise HTTPException(status_code=409, detail="A login already exists for this customer")

    # The mobile is the only hard requirement: it is what OTP login runs on.
    # Email and password are both optional — without them the account is
    # WhatsApp-OTP-only, which is a complete, usable login.
    raw_phone = payload.get("phone") or customer.get("mobile") or customer.get("phone")
    resolved = normalize_indian_mobile(raw_phone)
    if not resolved["valid"]:
        raise HTTPException(status_code=400, detail=resolved["reason"])

    email = (payload.get("email") or customer.get("email") or "").strip()
    if email and db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")

    password = (payload.get("password") or "").strip()
    if password and not email:
        raise HTTPException(
            status_code=400,
            detail="An email is needed to set a password — leave the password blank for OTP-only login",
        )

    # A second account on the same number would make OTP login ambiguous.
    if db.users.find_one({"phone": int(resolved["phone"])}):
        raise HTTPException(
            status_code=409,
            detail="Another account already uses this mobile number",
        )

    name = (
        payload.get("name")
        or customer.get("contact_name")
        or customer.get("company_name")
        or ""
    ).strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    now = datetime.utcnow()
    doc = {
        "name": name,
        "phone": int(resolved["phone"]),
        "role": "customer",
        "status": "active",
        "customer_id": contact_id,
        "customer_name": customer.get("company_name") or customer.get("contact_name") or "",
        "created_at": now,
        "updated_at": now,
    }
    # Omit rather than store "" — an empty email would collide with every other
    # emailless account on the duplicate check above.
    if email:
        doc["email"] = email
    if password:
        doc["password"] = hash_password(password)

    result = db.users.insert_one(doc)
    has_password = "password" in doc

    return {
        "message": (
            "Customer login created successfully"
            if has_password
            else "Customer login created — the customer signs in with a WhatsApp OTP"
        ),
        "login": {
            "_id": str(result.inserted_id),
            "name": name,
            "email": email,
            "phone": int(resolved["phone"]),
            "status": "active",
            "has_password": has_password,
        },
    }


@router.post("/customer-login/{contact_id}/send")
def send_customer_login_link(contact_id: str, user_id: Optional[str] = Query(None)):
    """
    WhatsApp the customer their login link (UTILITY template, signed token in the
    button). Never carries a password or the number in plain text.

    `user_id` picks which login to send to when the customer has more than one;
    without it the oldest is used.
    """
    from ..config.whatsapp import send_whatsapp
    from .users import make_login_link_token
    from .customer_creation_requests import CUSTOMER_LOGIN_TEMPLATE_NAME

    if user_id:
        user = db.users.find_one(
            {"_id": ObjectId(user_id), "customer_id": contact_id, "role": "customer"}
        )
        if not user:
            raise HTTPException(status_code=404, detail="That login does not belong to this customer")
    else:
        user = db.users.find_one(
            {"customer_id": contact_id, "role": "customer"}, sort=[("created_at", 1)]
        )
    if not user:
        raise HTTPException(status_code=404, detail="No login exists for this customer yet")

    resolved = normalize_indian_mobile(user.get("phone"))
    if not resolved["valid"]:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot send on WhatsApp — {resolved['reason'][0].lower()}{resolved['reason'][1:]}",
        )
    phone10 = resolved["phone"]

    template = db.templates.find_one({"name": CUSTOMER_LOGIN_TEMPLATE_NAME})
    if not template:
        raise HTTPException(
            status_code=400,
            detail=(
                f"WhatsApp template '{CUSTOMER_LOGIN_TEMPLATE_NAME}' not found. "
                "Create and get it approved in /admin/templates first."
            ),
        )

    customer = db.customers.find_one({"contact_id": contact_id}, {"company_name": 1, "contact_name": 1})
    first_name = (user.get("name") or "there").split()[0]
    shop = (
        user.get("customer_name")
        or (customer or {}).get("company_name")
        or (customer or {}).get("contact_name")
        or "your shop"
    )

    response = send_whatsapp(
        phone10,
        template,
        {"name": first_name, "shop": shop, "button_url": make_login_link_token(phone10)},
    )
    if response is None:
        raise HTTPException(status_code=502, detail="Failed to send WhatsApp message")

    return {"message": f"Login link sent on WhatsApp to {phone10}"}


@router.get("/{user_id}")
def get_user(user_id: str):
    """Get a single user by ID."""
    user = db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user = serialize_mongo_document(user)
    user.pop("password", None)
    return {"user": user}


@router.post("")
def create_user(user_data: dict):
    """
    Create a new user with all required fields.
    Required fields: email, password, name, phone, role, status
    Optional fields: first_name, last_name, code, designation, department, customer_id
    """
    # Validate required fields
    required_fields = ["email", "password", "name", "phone", "role", "status"]
    missing_fields = [f for f in required_fields if not user_data.get(f)]
    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required fields: {', '.join(missing_fields)}"
        )

    # Check if email already exists
    existing_user = db.users.find_one({"email": user_data["email"]})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already exists")

    # Check if code already exists (if provided)
    if user_data.get("code"):
        existing_code = db.users.find_one({"code": user_data["code"]})
        if existing_code:
            raise HTTPException(status_code=400, detail="User code already exists")

    # Convert phone to integer
    try:
        user_data["phone"] = int(user_data["phone"])
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Phone must be a valid number")

    # Hash the password
    user_data["password"] = hash_password(user_data["password"])

    # Add timestamps
    user_data["created_at"] = datetime.utcnow()
    user_data["updated_at"] = datetime.utcnow()

    # Insert the user
    result = db.users.insert_one(user_data)

    return {
        "message": "User created successfully",
        "user_id": str(result.inserted_id)
    }


@router.put("/{user_id}")
def update_user(user_id: str, user_data: dict):
    """
    Update an existing user.
    If password is provided, it will be hashed before saving.
    """
    # Check if user exists
    existing_user = db.users.find_one({"_id": ObjectId(user_id)})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")

    # Remove _id from update data if present
    user_data.pop("_id", None)

    # If email is being changed, check for duplicates
    if user_data.get("email") and user_data["email"] != existing_user.get("email"):
        email_exists = db.users.find_one({
            "email": user_data["email"],
            "_id": {"$ne": ObjectId(user_id)}
        })
        if email_exists:
            raise HTTPException(status_code=400, detail="Email already exists")

    # If code is being changed, check for duplicates
    if user_data.get("code") and user_data["code"] != existing_user.get("code"):
        code_exists = db.users.find_one({
            "code": user_data["code"],
            "_id": {"$ne": ObjectId(user_id)}
        })
        if code_exists:
            raise HTTPException(status_code=400, detail="User code already exists")

    # Convert phone to integer if provided
    if "phone" in user_data and user_data["phone"] is not None:
        try:
            user_data["phone"] = int(user_data["phone"])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Phone must be a valid number")

    # Hash password if provided and not empty
    if user_data.get("password"):
        user_data["password"] = hash_password(user_data["password"])
    else:
        user_data.pop("password", None)  # Don't update password if not provided

    # Add updated timestamp
    user_data["updated_at"] = datetime.utcnow()

    # Filter out None values
    update_data = {k: v for k, v in user_data.items() if v is not None}

    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields provided for update")

    # Perform the update
    result = db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_data}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": "User updated successfully"}


@router.delete("/{user_id}")
def delete_user(user_id: str):
    """Delete a user by ID."""
    result = db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}


@router.post("/{user_id}/reset-password")
def reset_user_password(user_id: str, password_data: dict):
    """
    Reset a user's password (admin function).
    Expects: { "password": "new_password" }
    """
    if not password_data.get("password"):
        raise HTTPException(status_code=400, detail="Password is required")

    # Check if user exists
    user = db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Hash and update password
    hashed_password = hash_password(password_data["password"])

    result = db.users.update_one(
        {"_id": ObjectId(user_id)},
        {
            "$set": {
                "password": hashed_password,
                "updated_at": datetime.utcnow()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": "Password reset successfully"}


@router.put("/{user_id}/status")
def update_user_status(user_id: str, status_data: dict):
    """
    Update user status (active/inactive).
    Expects: { "status": "active" | "inactive" }
    """
    new_status = status_data.get("status")
    if new_status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="Status must be 'active' or 'inactive'")

    result = db.users.update_one(
        {"_id": ObjectId(user_id)},
        {
            "$set": {
                "status": new_status,
                "updated_at": datetime.utcnow()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": f"User status updated to {new_status}"}


@router.get("/search/customers")
def search_customers_for_assignment(
    search: str = Query(..., min_length=1, description="Search term for customer name")
):
    """
    Search customers from the contacts collection for assignment to a user.
    Returns contact_id and contact_name for selection.
    """
    customers = db.customers.find(
        {
            "contact_name": {"$regex": search, "$options": "i"},
            "status": "active"
        },
        {
            "_id": 1,
            "contact_id": 1,
            "contact_name": 1,
            "company_name": 1,
            "email": 1
        }
    ).limit(20)

    seen = set()
    results = []
    for customer in customers:
        contact_id = customer.get("contact_id")
        if contact_id in seen:
            continue
        seen.add(contact_id)
        results.append({
            "_id": str(customer.get("_id")),
            "contact_id": contact_id,
            "contact_name": customer.get("contact_name"),
            "company_name": customer.get("company_name"),
            "email": customer.get("email"),
            "display_name": customer.get("company_name") or customer.get("contact_name") or customer.get("email", "Unknown")
        })

    return {"customers": results}
